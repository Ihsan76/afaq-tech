import io

import openpyxl
from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from rest_framework.test import APITestCase

from apps.academics.models import Grade, Subject
from apps.schools.models import (
    DEFAULT_WEEK_START,
    DEFAULT_WORKING_DAYS,
    FAQ,
    AcademicYear,
    AnnouncementReadReceipt,
    Attachment,
    Attendance,
    DayOfWeek,
    FamilyLink,
    ParentTeacherTicket,
    Period,
    School,
    SchoolAnnouncement,
    SchoolBus,
    SchoolGrade,
    SchoolTeacher,
    Section,
    StudentEnrollment,
    SupportRequest,
    TeacherAssignment,
    TimetableSlot,
    WeeklyReport,
    WhatsAppNotificationLog,
)

User = get_user_model()

TEST_PASSWORD = 'Afaq#Secure2026!'

def make_user(email, role, name_ar, national_id=None, phone='', **extra):
    """Creates a test user with a policy-compliant password and full user info."""
    fields = {
        'role': role,
        'translations': {'ar': {'name': name_ar}},
        'is_verified': True,
        'phone': phone,
    }
    if national_id is not None:
        fields['national_id'] = national_id
    fields.update(extra)
    return User.objects.create_user(email=email, password=TEST_PASSWORD, **fields)


class SchoolsAPITestCase(APITestCase):
    def setUp(self):
        self.admin = make_user('admin@test.com', 'admin', 'مدير الاختبار')
        self.student = make_user('student@test.com', 'student', 'طالب الاختبار', national_id='1234567890', phone='0771000001')
        self.student2 = make_user('student2@test.com', 'student', 'طالب الاختبار الثاني', national_id='1234567891', phone='0771000002')
        self.teacher = make_user('teacher@test.com', 'teacher', 'معلم الاختبار', phone='0771000003')

        self.grade = Grade.objects.create(level=3, translations={'ar': {'name': 'الصف الثالث'}})
        self.subject = Subject.objects.create(translations={'ar': {'name': 'الرياضيات'}})
        self.year = AcademicYear.objects.create(name='2025/2026', is_current=True)

        self.school_a = School.objects.create(name='مدرسة أ', school_code='A1')
        self.school_b = School.objects.create(name='مدرسة ب', school_code='B1')

        SchoolGrade.objects.create(school=self.school_a, grade=self.grade)
        SchoolGrade.objects.create(school=self.school_b, grade=self.grade)

        self.section_a = Section.objects.create(
            school=self.school_a, grade=self.grade, academic_year=self.year, name='أ'
        )
        self.section_b = Section.objects.create(
            school=self.school_b, grade=self.grade, academic_year=self.year, name='ب'
        )

        StudentEnrollment.objects.create(student=self.student, section=self.section_a, academic_year=self.year)
        StudentEnrollment.objects.create(student=self.student2, section=self.section_b, academic_year=self.year)
        TeacherAssignment.objects.create(
            teacher=self.teacher, section=self.section_a, subject=self.subject, academic_year=self.year
        )

        self.ann_section = SchoolAnnouncement.objects.create(
            author=self.teacher, school=self.school_a, section=self.section_a,
            title='واجب رياضيات', content='حل صفحة 20'
        )
        self.ann_emergency = SchoolAnnouncement.objects.create(
            author=self.admin, school=self.school_b, section=self.section_b,
            title='حالة طارئة', content='إغلاق مبكر', is_emergency=True
        )

    def auth(self, user):
        self.client.force_authenticate(user=user)

    def results(self, res):
        return res.data.get('results', res.data)

    def test_student_sees_only_own_section_announcements(self):
        self.auth(self.student)
        res = self.client.get('/api/v1/schools/announcements/')
        self.assertEqual(res.status_code, 200)
        titles = [a['title'] for a in self.results(res)]
        self.assertIn('واجب رياضيات', titles)
        self.assertNotIn('حالة طارئة', titles)

    def test_teacher_sees_only_assigned_sections(self):
        self.auth(self.teacher)
        res = self.client.get('/api/v1/schools/sections/')
        self.assertEqual(res.status_code, 200)
        ids = [s['id'] for s in self.results(res)]
        self.assertIn(self.section_a.id, ids)
        self.assertNotIn(self.section_b.id, ids)

    def test_student_cannot_write_schools(self):
        self.auth(self.student)
        res = self.client.post('/api/v1/schools/schools/', {
            'name': 'مدرسة مخترقة', 'school_code': 'HACK'
        })
        self.assertEqual(res.status_code, 403)

    def test_teacher_can_create_announcement(self):
        self.auth(self.teacher)
        res = self.client.post('/api/v1/schools/announcements/', {
            'school': self.school_a.id,
            'section': self.section_a.id,
            'title': 'إعلان جديد', 'content': 'اختبار'
        })
        self.assertEqual(res.status_code, 201)

    def test_teacher_cannot_force_emergency(self):
        self.auth(self.teacher)
        res = self.client.post('/api/v1/schools/announcements/', {
            'school': self.school_a.id,
            'section': self.section_a.id,
            'title': 'محاولة طارئة', 'content': 'اختبار',
            'is_emergency': True
        })
        self.assertEqual(res.status_code, 201)
        self.assertFalse(res.data['is_emergency'])

    def test_admin_can_force_emergency(self):
        self.auth(self.admin)
        res = self.client.post('/api/v1/schools/announcements/', {
            'school': self.school_b.id,
            'section': self.section_b.id,
            'title': 'تنبيه طارئ', 'content': 'اختبار',
            'is_emergency': True
        })
        self.assertEqual(res.status_code, 201)
        self.assertTrue(res.data['is_emergency'])

    def test_tickets_scoped_to_participants(self):
        ParentTeacherTicket.objects.create(
            parent=self.student, teacher=self.teacher, student=self.student,
            title='استفسار عن الواجب'
        )
        ParentTeacherTicket.objects.create(
            parent=self.student2, teacher=self.teacher, student=self.student2,
            title='استفسار آخر'
        )
        self.auth(self.student)
        res = self.client.get('/api/v1/schools/tickets/')
        titles = [t['title'] for t in self.results(res)]
        self.assertIn('استفسار عن الواجب', titles)
        self.assertNotIn('استفسار آخر', titles)

    def test_my_context_student(self):
        self.auth(self.student)
        res = self.client.get('/api/v1/schools/my-context/')
        self.assertEqual(res.status_code, 200)
        self.assertEqual(res.data['role'], 'student')
        self.assertEqual(len(res.data['sections']), 1)
        self.assertEqual(res.data['sections'][0]['id'], self.section_a.id)
        self.assertEqual(len(res.data['announcements']), 1)
        self.assertEqual(len(res.data['tickets']), 0)

    def test_my_context_teacher(self):
        self.auth(self.teacher)
        res = self.client.get('/api/v1/schools/my-context/')
        self.assertEqual(res.status_code, 200)
        self.assertEqual(len(res.data['sections']), 1)
        self.assertEqual(len(res.data['announcements']), 1)

    def test_my_context_admin_sees_all(self):
        self.auth(self.admin)
        res = self.client.get('/api/v1/schools/my-context/', {'school': self.school_a.id})
        self.assertEqual(res.status_code, 200)
        self.assertEqual(len(res.data['sections']), 1)
        self.assertEqual(res.data['sections'][0]['id'], self.section_a.id)
        self.assertEqual(len(res.data['announcements']), 1)

    def test_ticket_reply_forbidden_for_outsider(self):
        ticket = ParentTeacherTicket.objects.create(
            parent=self.student, teacher=self.teacher, student=self.student,
            title='استفسار'
        )
        self.auth(self.student2)
        res = self.client.post(f'/api/v1/schools/tickets/{ticket.id}/add_message/', {
            'message': 'تدخل خارجي'
        })
        self.assertEqual(res.status_code, 404)


class SchoolsAdvancedFeaturesTestCase(APITestCase):
    def setUp(self):
        self.admin = make_user('admin@test.com', 'admin', 'مدير الاختبار')
        self.student = make_user('student@test.com', 'student', 'طالب الاختبار', national_id='9876543210', phone='0771000001')
        self.parent = make_user('parent@test.com', 'parent', 'ولي أمر الاختبار', phone='0771000002')
        self.teacher = make_user('teacher@test.com', 'teacher', 'معلم الاختبار', phone='0771000003')

        self.grade3 = Grade.objects.create(level=3, translations={'ar': {'name': 'الصف الثالث'}})
        self.grade4 = Grade.objects.create(level=4, translations={'ar': {'name': 'الصف الرابع'}})
        self.subject = Subject.objects.create(translations={'ar': {'name': 'الرياضيات'}})
        self.year_2025 = AcademicYear.objects.create(name='2025/2026', is_current=True)
        self.year_2026 = AcademicYear.objects.create(name='2026/2027', is_current=False)

        self.school = School.objects.create(name='مدرسة أ', school_code='A1')
        self.section3 = Section.objects.create(school=self.school, grade=self.grade3, academic_year=self.year_2025, name='أ')
        self.section4 = Section.objects.create(school=self.school, grade=self.grade4, academic_year=self.year_2026, name='أ')

        StudentEnrollment.objects.create(student=self.student, section=self.section3, academic_year=self.year_2025)
        TeacherAssignment.objects.create(teacher=self.teacher, section=self.section3, subject=self.subject, academic_year=self.year_2025)

        self.announcement = SchoolAnnouncement.objects.create(
            author=self.teacher, school=self.school, section=self.section3,
            title='واجب أسبوعي', content='حل التمارين'
        )

        FamilyLink.objects.create(parent=self.parent, student=self.student, relationship='والد')

    def auth(self, user):
        self.client.force_authenticate(user=user)

    def test_family_link_listed_for_parent(self):
        self.auth(self.parent)
        res = self.client.get('/api/v1/schools/family-links/')
        self.assertEqual(res.status_code, 200)
        data = res.data.get('results', res.data)
        self.assertEqual(len(data), 1)
        self.assertEqual(data[0]['student_email'], 'student@test.com')

    def test_family_link_creation_sets_parent(self):
        other = make_user('other@test.com', 'student', 'طالب آخر', national_id='5550001112')
        self.auth(self.parent)
        res = self.client.post('/api/v1/schools/family-links/', {'student': other.id})
        self.assertEqual(res.status_code, 201)
        self.assertEqual(res.data['parent'], self.parent.id)

    def test_family_link_duplicate_rejected(self):
        self.auth(self.parent)
        res = self.client.post('/api/v1/schools/family-links/', {'student': self.student.id})
        self.assertEqual(res.status_code, 400)

    def test_family_link_cannot_link_non_student(self):
        self.auth(self.parent)
        res = self.client.post('/api/v1/schools/family-links/', {'student': self.teacher.id})
        self.assertEqual(res.status_code, 400)

    def test_parent_context_has_children_and_announcements(self):
        self.auth(self.parent)
        res = self.client.get('/api/v1/schools/my-context/')
        self.assertEqual(res.status_code, 200)
        self.assertEqual(len(res.data['children']), 1)
        self.assertEqual(res.data['children'][0]['email'], 'student@test.com')
        self.assertEqual(len(res.data['family_links']), 1)
        self.assertEqual(len(res.data['announcements']), 1)

    def test_acknowledge_announcement(self):
        self.auth(self.student)
        res = self.client.post(f'/api/v1/schools/announcements/{self.announcement.id}/acknowledge/')
        self.assertEqual(res.status_code, 200)
        self.assertEqual(res.data['read_count'], 1)
        receipt = AnnouncementReadReceipt.objects.get(announcement=self.announcement, user=self.student)
        self.assertIsNotNone(receipt)

    def test_announcement_serializer_exposes_is_read(self):
        self.auth(self.student)
        res = self.client.get(f'/api/v1/schools/announcements/{self.announcement.id}/')
        self.assertEqual(res.status_code, 200)
        self.assertFalse(res.data['is_read'])
        self.client.post(f'/api/v1/schools/announcements/{self.announcement.id}/acknowledge/')
        res = self.client.get(f'/api/v1/schools/announcements/{self.announcement.id}/')
        self.assertTrue(res.data['is_read'])
        self.assertEqual(res.data['read_count'], 1)

    def test_promote_rolls_students_to_next_grade(self):
        self.auth(self.admin)
        res = self.client.post(f'/api/v1/schools/academic-years/{self.year_2025.id}/promote/', {
            'target_year_id': self.year_2026.id,
        })
        self.assertEqual(res.status_code, 200)
        self.assertEqual(len(res.data['promoted']), 1)
        new_enrollment = StudentEnrollment.objects.get(student=self.student, academic_year=self.year_2026)
        self.assertEqual(new_enrollment.section.grade, self.grade4)
        # archive preserved
        self.assertTrue(StudentEnrollment.objects.filter(student=self.student, academic_year=self.year_2025).exists())
        # year switching
        self.year_2025.refresh_from_db()
        self.year_2026.refresh_from_db()
        self.assertFalse(self.year_2025.is_current)
        self.assertTrue(self.year_2026.is_current)

    def test_transfer_student_to_other_section(self):
        self.auth(self.admin)
        other_section = Section.objects.create(school=self.school, grade=self.grade3, academic_year=self.year_2025, name='ب')
        res = self.client.post(f'/api/v1/schools/enrollments/{StudentEnrollment.objects.get(student=self.student).id}/transfer/', {
            'target_section_id': other_section.id,
        })
        self.assertEqual(res.status_code, 200)
        en = StudentEnrollment.objects.get(student=self.student, academic_year=self.year_2025)
        self.assertEqual(en.section, other_section)

    def test_transfer_by_code(self):
        self.auth(self.admin)
        self.student.national_id = '987654321'
        self.student.save()
        other_section = Section.objects.create(school=self.school, grade=self.grade3, academic_year=self.year_2025, name='ج')
        res = self.client.post('/api/v1/schools/enrollments/transfer_by_code/', {
            'national_id': '987654321',
            'target_section_id': other_section.id,
        })
        self.assertEqual(res.status_code, 200)
        en = StudentEnrollment.objects.get(student=self.student, academic_year=self.year_2025)
        self.assertEqual(en.section, other_section)

    def test_weekly_summary_for_parent(self):
        self.auth(self.parent)
        res = self.client.get('/api/v1/schools/weekly-summary/')
        self.assertEqual(res.status_code, 200)
        self.assertEqual(len(res.data['reports']), 1)
        report = res.data['reports'][0]
        self.assertEqual(report['student_email'], 'student@test.com')
        self.assertEqual(report['assignments_submitted'], 1)
        self.assertTrue(WeeklyReport.objects.filter(student=self.student).exists())

    def test_weekly_summary_for_student(self):
        self.auth(self.student)
        res = self.client.get('/api/v1/schools/weekly-summary/')
        self.assertEqual(res.status_code, 200)
        self.assertEqual(len(res.data['reports']), 1)

    def test_faq_list_public(self):
        FAQ.objects.create(question='كيف أسجل دخولي؟', answer='من صفحة تسجيل الدخول', sort_order=1)
        res = self.client.get('/api/v1/schools/faqs/')
        self.assertEqual(res.status_code, 200)
        data = res.data.get('results', res.data)
        self.assertEqual(len(data), 1)
        self.assertEqual(data[0]['question'], 'كيف أسجل دخولي؟')

    def test_support_request_creation(self):
        self.auth(self.student)
        res = self.client.post('/api/v1/schools/support/email/', {
            'subject': 'مشكلة في تسجيل الدخول',
            'message': 'لا أستطيع تسجيل الدخول لحسابي',
        })
        self.assertEqual(res.status_code, 201)
        self.assertEqual(SupportRequest.objects.count(), 1)
        self.assertEqual(SupportRequest.objects.first().user, self.student)

    def test_bulk_import_students(self):
        self.auth(self.admin)
        csv_content = (
            "email,name,national_id,phone,parent_email,school_code,grade_level,section_name,academic_year\r\n"
            "newstudent@test.com,طالب جديد,111222333,0777000000,parent2@test.com,A1,3,أ,2025/2026\r\n"
        ).encode()
        uploaded = SimpleUploadedFile('students.csv', csv_content, content_type='text/csv')
        res = self.client.post('/api/v1/schools/bulk/import/', {'kind': 'students', 'file': uploaded})
        self.assertEqual(res.status_code, 200)
        self.assertEqual(res.data['created'], 1)
        new_student = User.objects.get(email='newstudent@test.com')
        self.assertEqual(new_student.role, 'student')
        self.assertEqual(new_student.national_id, '111222333')
        self.assertTrue(StudentEnrollment.objects.filter(student=new_student).exists())
        parent2 = User.objects.get(email='parent2@test.com')
        self.assertEqual(parent2.role, 'parent')
        self.assertTrue(FamilyLink.objects.filter(parent=parent2, student=new_student).exists())

    def test_bulk_export_students_csv(self):
        self.auth(self.admin)
        res = self.client.get('/api/v1/schools/bulk/export/?kind=students&file_format=csv')
        self.assertEqual(res.status_code, 200)
        self.assertIn('text/csv', res['Content-Type'])
        content = res.content.decode('utf-8')
        self.assertIn('student@test.com', content)
        self.assertIn('A1', content)

    def test_bulk_import_schools_from_xlsx(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['رمز المؤسسة', 'اسم المؤسسة', 'المديرية', 'المحافظة', 'الإقليم', 'جنس المؤسس', 'نوع التعليم', 'العنوان'])
        ws.append(['990001', 'مدرسة الاستيراد التجريبية', 'قصبة عمان', 'العاصمة عمان', 'الوسط', 'مختلطة', 'المرحلة الأساسية', 'حي الرابية'])
        ws.append(['990002', 'مدرسة الاستيراد الثانوية', 'اربد الأولى', 'اربد', 'الشمال', 'اناث', 'المرحلة الثانوية', 'وسط البلد'])
        buffer = io.BytesIO()
        wb.save(buffer)
        uploaded = SimpleUploadedFile('schools.xlsx', buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        self.auth(self.admin)
        res = self.client.post('/api/v1/schools/bulk/import/', {'kind': 'schools', 'file': uploaded})
        self.assertEqual(res.status_code, 200)
        self.assertEqual(res.data['created'], 2)
        school = School.objects.get(school_code='990001')
        self.assertEqual(school.name, 'مدرسة الاستيراد التجريبية')
        self.assertEqual(school.governorate, 'العاصمة عمان')
        self.assertEqual(school.region, 'الوسط')
        self.assertEqual(school.gender, 'مختلطة')
        self.assertEqual(school.education_type, 'المرحلة الأساسية')
        self.assertEqual(school.translations['ar']['name'], 'مدرسة الاستيراد التجريبية')

    def test_bulk_import_schools_from_csv(self):
        csv_content = (
            "school_code,name,directorate,governorate\r\n"
            "990003,مدرسة من ملف نصي,معان,معان\r\n"
        ).encode()
        uploaded = SimpleUploadedFile('schools.csv', csv_content, content_type='text/csv')
        self.auth(self.admin)
        res = self.client.post('/api/v1/schools/bulk/import/', {'kind': 'schools', 'file': uploaded})
        self.assertEqual(res.status_code, 200)
        self.assertEqual(res.data['created'], 1)
        school = School.objects.get(school_code='990003')
        self.assertEqual(school.governorate, 'معان')

    def test_bulk_import_schools_skips_missing_code(self):
        csv_content = (
            "school_code,name\r\n"
            ",مدرسة بدون رمز\r\n"
        ).encode()
        uploaded = SimpleUploadedFile('schools.csv', csv_content, content_type='text/csv')
        self.auth(self.admin)
        res = self.client.post('/api/v1/schools/bulk/import/', {'kind': 'schools', 'file': uploaded})
        self.assertEqual(res.status_code, 200)
        self.assertEqual(len(res.data['errors']), 1)
        self.assertEqual(School.objects.filter(name='مدرسة بدون رمز').count(), 0)

    def test_bulk_export_schools_csv(self):
        School.objects.create(name='مدرسة التصدير', school_code='990100', governorate='الزرقاء')
        self.auth(self.admin)
        res = self.client.get('/api/v1/schools/bulk/export/?kind=schools&file_format=csv')
        self.assertEqual(res.status_code, 200)
        self.assertIn('text/csv', res['Content-Type'])
        content = res.content.decode('utf-8')
        self.assertIn('990100', content)
        self.assertIn('الزرقاء', content)

    def test_voice_transcribe_mock_fallback(self):
        self.auth(self.student)
        from django.core.files.uploadedfile import SimpleUploadedFile
        audio = SimpleUploadedFile('a.webm', b'\x1aE\xdf\xa3fakeaudio', content_type='audio/webm')
        res = self.client.post('/api/v1/schools/voice/transcribe/', {'audio': audio}, format='multipart')
        self.assertEqual(res.status_code, 200)
        self.assertIn('text', res.data)
        self.assertTrue(res.data['text'])

    def test_teacher_uploads_lesson_attachment(self):
        self.auth(self.teacher)
        file = SimpleUploadedFile('lesson.pdf', b'%PDF-1.4 lesson content', content_type='application/pdf')
        res = self.client.post('/api/v1/schools/attachments/', {
            'kind': 'lesson', 'title': 'شرح الكسور', 'description': 'ملف شرح',
            'section': self.section3.id, 'file': file,
        }, format='multipart')
        self.assertEqual(res.status_code, 201)
        self.assertEqual(res.data['uploader'], self.teacher.id)
        self.assertEqual(res.data['uploader_name'], 'معلم الاختبار')
        self.assertEqual(res.data['kind'], 'lesson')
        self.assertEqual(res.data['review_status'], 'pending')
        self.assertEqual(res.data['file_name'], 'lesson.pdf')
        self.assertTrue(res.data['file_url'])

    def test_student_uploads_homework_submission(self):
        self.auth(self.student)
        file = SimpleUploadedFile('homework.jpg', b'\xff\xd8\xff\xe0 fake image', content_type='image/jpeg')
        res = self.client.post('/api/v1/schools/attachments/', {
            'kind': 'submission', 'title': 'حل التمارين', 'file': file,
        }, format='multipart')
        self.assertEqual(res.status_code, 201)
        self.assertEqual(res.data['kind'], 'submission')
        self.assertEqual(res.data['uploader_name'], 'طالب الاختبار')

    def test_upload_requires_file(self):
        self.auth(self.teacher)
        res = self.client.post('/api/v1/schools/attachments/', {'kind': 'lesson'})
        self.assertEqual(res.status_code, 400)

    def test_student_cannot_upload_to_unrelated_section(self):
        self.auth(self.student)
        file = SimpleUploadedFile('x.png', b'png', content_type='image/png')
        other_section = Section.objects.create(school=self.school, grade=self.grade3, academic_year=self.year_2025, name='ب')
        res = self.client.post('/api/v1/schools/attachments/', {
            'kind': 'homework', 'section': other_section.id, 'file': file,
        }, format='multipart')
        self.assertEqual(res.status_code, 400)

    def test_admin_can_review_attachment(self):
        self.auth(self.student)
        file = SimpleUploadedFile('hw.pdf', b'pdf', content_type='application/pdf')
        res = self.client.post('/api/v1/schools/attachments/', {
            'kind': 'homework', 'title': 'واجب', 'file': file,
        }, format='multipart')
        self.assertEqual(res.status_code, 201)
        attachment_id = res.data['id']

        self.auth(self.teacher)
        res = self.client.post(f'/api/v1/schools/attachments/{attachment_id}/review/', {'review_status': 'approved'})
        self.assertEqual(res.status_code, 403)

        self.auth(self.admin)
        res = self.client.post(f'/api/v1/schools/attachments/{attachment_id}/review/', {
            'review_status': 'approved', 'review_notes': 'ممتاز',
        })
        self.assertEqual(res.status_code, 200)
        self.assertEqual(res.data['review_status'], 'approved')
        self.assertEqual(res.data['review_notes'], 'ممتاز')
        attachment = Attachment.objects.get(id=attachment_id)
        self.assertEqual(attachment.reviewed_by, self.admin)

    def test_admin_my_context_lists_attachments(self):
        self.auth(self.student)
        file = SimpleUploadedFile('a.png', b'pngdata', content_type='image/png')
        res = self.client.post('/api/v1/schools/attachments/', {'kind': 'lesson', 'file': file}, format='multipart')
        self.assertEqual(res.status_code, 201)

        self.auth(self.admin)
        res = self.client.get('/api/v1/schools/my-context/')
        self.assertEqual(res.status_code, 200)
        self.assertGreaterEqual(len(res.data['attachments']), 1)

    def test_analytics_reports_attachment_counts(self):
        self.auth(self.student)
        file = SimpleUploadedFile('b.png', b'pngdata', content_type='image/png')
        self.client.post('/api/v1/schools/attachments/', {'kind': 'submission', 'file': file}, format='multipart')

        self.auth(self.admin)
        res = self.client.get('/api/v1/schools/analytics/')
        self.assertEqual(res.status_code, 200)
        self.assertGreaterEqual(res.data['attachments_total'], 1)
        self.assertGreaterEqual(res.data['attachments_pending_review'], 1)


class AttendanceFeatureTestCase(APITestCase):
    def setUp(self):
        self.admin = make_user('admin@test.com', 'admin', 'مدير الاختبار')
        self.student = make_user('student@test.com', 'student', 'طالب الاختبار', national_id='1234567890', phone='0771000001')
        self.parent = make_user('parent@test.com', 'parent', 'ولي أمر الاختبار', phone='0772000001')
        self.teacher = make_user('teacher@test.com', 'teacher', 'معلم الاختبار', phone='0773000001')

        self.grade = Grade.objects.create(level=5, translations={'ar': {'name': 'الصف الخامس'}})
        self.subject = Subject.objects.create(translations={'ar': {'name': 'العلوم'}})
        self.year = AcademicYear.objects.create(name='2025/2026', is_current=True)
        self.school = School.objects.create(name='مدرسة الحضور', school_code='AT1')
        self.section_a = Section.objects.create(school=self.school, grade=self.grade, academic_year=self.year, name='أ')
        self.section_b = Section.objects.create(school=self.school, grade=self.grade, academic_year=self.year, name='ب')

        StudentEnrollment.objects.create(student=self.student, section=self.section_a, academic_year=self.year)
        TeacherAssignment.objects.create(teacher=self.teacher, section=self.section_a, subject=self.subject, academic_year=self.year)
        FamilyLink.objects.create(parent=self.parent, student=self.student, relationship='والد')

        self.target_date = '2026-08-03'  # a Monday (weekday, not skipped by the command)

    def auth(self, user):
        self.client.force_authenticate(user=user)

    def results(self, res):
        return res.data.get('results', res.data)

    def test_teacher_bulk_records_attendance(self):
        self.auth(self.teacher)
        res = self.client.post('/api/v1/schools/attendances/bulk_record/', {
            'section': self.section_a.id,
            'date': self.target_date,
            'records': [{'student': self.student.id, 'status': 'present'}],
        }, format='json')
        self.assertEqual(res.status_code, 200)
        self.assertEqual(res.data['created'], 1)
        attendance = Attendance.objects.get(student=self.student, date=self.target_date)
        self.assertEqual(attendance.status, 'present')
        self.assertEqual(attendance.recorded_by, self.teacher)
        self.assertEqual(attendance.section, self.section_a)

    def test_teacher_cannot_record_unassigned_section(self):
        self.auth(self.teacher)
        res = self.client.post('/api/v1/schools/attendances/bulk_record/', {
            'section': self.section_b.id,
            'date': self.target_date,
            'records': [{'student': self.student.id, 'status': 'present'}],
        }, format='json')
        self.assertEqual(res.status_code, 403)

    def test_absent_triggers_whatsapp_and_inapp_alert(self):
        self.auth(self.teacher)
        from apps.notifications.models import Notification
        res = self.client.post('/api/v1/schools/attendances/bulk_record/', {
            'section': self.section_a.id,
            'date': self.target_date,
            'records': [{'student': self.student.id, 'status': 'absent'}],
        }, format='json')
        self.assertEqual(res.status_code, 200)
        self.assertEqual(res.data['absent_alerts_sent'], 1)
        self.assertTrue(WhatsAppNotificationLog.objects.filter(recipient_phone=self.parent.phone).exists())
        self.assertTrue(Notification.objects.filter(user=self.parent, type='absence').exists())

    def test_bulk_record_is_idempotent(self):
        self.auth(self.teacher)
        payload = {
            'section': self.section_a.id,
            'date': self.target_date,
            'records': [{'student': self.student.id, 'status': 'present'}],
        }
        self.client.post('/api/v1/schools/attendances/bulk_record/', payload, format='json')
        res = self.client.post('/api/v1/schools/attendances/bulk_record/', payload, format='json')
        self.assertEqual(res.status_code, 200)
        self.assertEqual(res.data['updated'], 1)
        self.assertEqual(Attendance.objects.filter(student=self.student, date=self.target_date).count(), 1)

    def test_teacher_list_scoped_to_assigned_sections(self):
        self.auth(self.teacher)
        Attendance.objects.create(student=self.student, section=self.section_a, school=self.school, date=self.target_date, status='present')
        other = make_user('other@test.com', 'student', 'طالب آخر', phone='0779000001')
        StudentEnrollment.objects.create(student=other, section=self.section_b, academic_year=self.year)
        Attendance.objects.create(student=other, section=self.section_b, school=self.school, date=self.target_date, status='present')

        res = self.client.get('/api/v1/schools/attendances/')
        self.assertEqual(res.status_code, 200)
        students = {a['student'] for a in self.results(res)}
        self.assertIn(self.student.id, students)
        self.assertNotIn(other.id, students)

    def test_enrollments_filtered_by_section(self):
        self.auth(self.teacher)
        other = make_user('other@test.com', 'student', 'طالب آخر', phone='0779000001')
        StudentEnrollment.objects.create(student=other, section=self.section_b, academic_year=self.year)

        res = self.client.get('/api/v1/schools/enrollments/', {'section': self.section_a.id})
        self.assertEqual(res.status_code, 200)
        data = self.results(res)
        self.assertEqual(len(data), 1)
        self.assertEqual(data[0]['student'], self.student.id)
        self.assertEqual(data[0]['section'], self.section_a.id)

    def test_parent_views_child_attendance(self):
        self.auth(self.teacher)
        self.client.post('/api/v1/schools/attendances/bulk_record/', {
            'section': self.section_a.id,
            'date': self.target_date,
            'records': [{'student': self.student.id, 'status': 'absent'}],
        }, format='json')
        self.auth(self.parent)
        res = self.client.get(f'/api/v1/schools/attendances/?date={self.target_date}')
        self.assertEqual(res.status_code, 200)
        data = self.results(res)
        self.assertEqual(len(data), 1)
        self.assertEqual(data[0]['student_email'], 'student@test.com')
        self.assertEqual(data[0]['status'], 'absent')

    def test_student_own_attendance_in_my_context(self):
        self.auth(self.teacher)
        self.client.post('/api/v1/schools/attendances/bulk_record/', {
            'section': self.section_a.id,
            'date': self.target_date,
            'records': [{'student': self.student.id, 'status': 'present'}],
        }, format='json')
        self.auth(self.student)
        res = self.client.get('/api/v1/schools/my-context/')
        self.assertEqual(res.status_code, 200)
        self.assertEqual(len(res.data['attendance']), 1)
        self.assertEqual(res.data['attendance'][0]['status'], 'present')

    def test_weekly_summary_uses_real_attendance(self):
        from datetime import date, timedelta
        monday = date.today() - timedelta(days=date.today().weekday())
        self.auth(self.teacher)
        self.client.post('/api/v1/schools/attendances/bulk_record/', {
            'section': self.section_a.id,
            'date': monday.isoformat(),
            'records': [{'student': self.student.id, 'status': 'present'}],
        }, format='json')
        self.auth(self.parent)
        res = self.client.get('/api/v1/schools/weekly-summary/')
        self.assertEqual(res.status_code, 200)
        self.assertEqual(res.data['reports'][0]['attendance_rate'], 100.0)

    def test_analytics_reports_attendance_counts(self):
        from django.utils import timezone
        self.auth(self.teacher)
        self.client.post('/api/v1/schools/attendances/bulk_record/', {
            'section': self.section_a.id,
            'date': timezone.localdate().isoformat(),
            'records': [{'student': self.student.id, 'status': 'absent'}],
        }, format='json')
        self.auth(self.admin)
        res = self.client.get('/api/v1/schools/analytics/')
        self.assertEqual(res.status_code, 200)
        self.assertGreaterEqual(res.data['attendance_today'], 1)
        self.assertGreaterEqual(res.data['absent_today'], 1)

    def test_send_absence_alerts_command(self):
        from django.core.management import call_command
        call_command('send_absence_alerts', date=self.target_date)
        attendance = Attendance.objects.get(student=self.student, date=self.target_date)
        self.assertEqual(attendance.status, 'absent')
        self.assertTrue(WhatsAppNotificationLog.objects.filter(recipient_phone=self.parent.phone).exists())
        from apps.notifications.models import Notification
        self.assertTrue(Notification.objects.filter(user=self.parent, type='absence').exists())

    def test_send_absence_alerts_is_idempotent(self):
        from django.core.management import call_command
        call_command('send_absence_alerts', date=self.target_date)
        call_command('send_absence_alerts', date=self.target_date)
        self.assertEqual(Attendance.objects.filter(student=self.student, date=self.target_date).count(), 1)
        self.assertEqual(WhatsAppNotificationLog.objects.count(), 1)

    def test_send_absence_alerts_dry_run(self):
        from django.core.management import call_command
        call_command('send_absence_alerts', date=self.target_date, dry_run=True)
        self.assertFalse(Attendance.objects.filter(student=self.student, date=self.target_date).exists())
        self.assertFalse(WhatsAppNotificationLog.objects.exists())


class PeriodGenerationTestCase(APITestCase):
    """Tests for period schedule generation, archiving, and restore."""

    def setUp(self):
        self.admin = make_user('periodadmin@test.com', 'admin', 'مدير النظام')
        self.manager = make_user('periodmanager@test.com', 'school_admin', 'مدير المدرسة', phone='0772000001')
        self.school = School.objects.create(name='مدرسة الحصص', school_code='PR1', manager=self.manager)

    def auth(self, user):
        self.client.force_authenticate(user=user)

    def test_generate_creates_periods_with_correct_times(self):
        self.auth(self.manager)
        res = self.client.post('/api/v1/schools/periods/generate/', {
            'school_id': self.school.id,
            'start_time': '08:00',
            'period_duration_min': 45,
            'break_duration_min': 10,
            'long_break_duration_min': 30,
            'long_break_after_period': 3,
            'total_periods': 4,
        }, format='json')
        self.assertEqual(res.status_code, 201, res.data)
        self.assertEqual(len(res.data), 4)
        self.assertEqual(res.data[0]['period_number'], 1)
        self.assertEqual(res.data[0]['start_time'], '08:00:00')
        self.assertEqual(res.data[0]['end_time'], '08:45:00')
        # After period 1 a short break (10m) => period 2 at 08:55
        self.assertEqual(res.data[1]['start_time'], '08:55:00')
        self.assertEqual(res.data[1]['end_time'], '09:40:00')
        # After period 2 a short break (10m) => period 3 at 09:50
        self.assertEqual(res.data[2]['start_time'], '09:50:00')
        # After period 3 the LONG break (30m) => period 4 at 11:05
        self.assertEqual(res.data[3]['start_time'], '11:05:00')
        self.assertEqual(res.data[3]['end_time'], '11:50:00')
        self.assertTrue(all(p['is_active'] for p in res.data))
        self.assertTrue(all(p['is_break'] is False for p in res.data))

    def test_generate_archives_previous_schedule_and_reuses_period_number(self):
        self.auth(self.manager)
        first = self.client.post('/api/v1/schools/periods/generate/', {
            'school_id': self.school.id, 'start_time': '08:00', 'total_periods': 3,
        }, format='json')
        self.assertEqual(first.status_code, 201)
        gen1 = first.data[0]['generation']

        second = self.client.post('/api/v1/schools/periods/generate/', {
            'school_id': self.school.id, 'start_time': '09:00', 'total_periods': 3,
        }, format='json')
        self.assertEqual(second.status_code, 201)
        gen2 = second.data[0]['generation']
        self.assertNotEqual(gen1, gen2)
        # Old schedule archived, new one active
        self.assertEqual(Period.objects.filter(school=self.school, generation=gen1, is_active=False).count(), 3)
        self.assertEqual(Period.objects.filter(school=self.school, generation=gen2, is_active=True).count(), 3)
        self.assertEqual(second.data[0]['start_time'], '09:00:00')

    def test_generate_requires_school_ownership(self):
        self.auth(self.manager)
        other = School.objects.create(name='مدرسة أخرى', school_code='PR2')
        res = self.client.post('/api/v1/schools/periods/generate/', {
            'school_id': other.id, 'start_time': '08:00', 'total_periods': 3,
        }, format='json')
        self.assertEqual(res.status_code, 400)

    def test_generate_validates_inputs(self):
        self.auth(self.manager)
        res = self.client.post('/api/v1/schools/periods/generate/', {
            'school_id': self.school.id, 'start_time': '08:00', 'period_duration_min': 0, 'total_periods': 0,
        }, format='json')
        self.assertEqual(res.status_code, 400)
        bad_time = self.client.post('/api/v1/schools/periods/generate/', {
            'school_id': self.school.id, 'start_time': 'bad', 'total_periods': 2,
        }, format='json')
        self.assertEqual(bad_time.status_code, 400)

    def test_default_list_shows_active_periods_only(self):
        self.auth(self.manager)
        self.client.post('/api/v1/schools/periods/generate/', {
            'school_id': self.school.id, 'start_time': '08:00', 'total_periods': 2,
        }, format='json')
        self.client.post('/api/v1/schools/periods/generate/', {
            'school_id': self.school.id, 'start_time': '09:00', 'total_periods': 2,
        }, format='json')
        res = self.client.get(f'/api/v1/schools/periods/?school={self.school.id}')
        self.assertEqual(res.status_code, 200)
        periods = res.data.get('results', res.data)
        self.assertEqual(len(periods), 2)
        self.assertEqual(periods[0]['start_time'], '09:00:00')

    def test_archives_lists_generations_with_metadata(self):
        self.auth(self.manager)
        self.client.post('/api/v1/schools/periods/generate/', {
            'school_id': self.school.id, 'start_time': '08:00', 'total_periods': 3,
        }, format='json')
        self.client.post('/api/v1/schools/periods/generate/', {
            'school_id': self.school.id, 'start_time': '09:00', 'total_periods': 3,
        }, format='json')
        res = self.client.get(f'/api/v1/schools/periods/archives/?school_id={self.school.id}')
        self.assertEqual(res.status_code, 200)
        self.assertEqual(len(res.data), 1)  # only the first generation is archived
        self.assertEqual(res.data[0]['count'], 3)
        self.assertEqual(res.data[0]['archived_by_name'], 'مدير المدرسة')

    def test_restore_swaps_generations(self):
        self.auth(self.manager)
        first = self.client.post('/api/v1/schools/periods/generate/', {
            'school_id': self.school.id, 'start_time': '08:00', 'total_periods': 3,
        }, format='json')
        gen1 = first.data[0]['generation']
        second = self.client.post('/api/v1/schools/periods/generate/', {
            'school_id': self.school.id, 'start_time': '09:00', 'total_periods': 3,
        }, format='json')
        gen2 = second.data[0]['generation']

        res = self.client.post('/api/v1/schools/periods/restore/', {
            'school_id': self.school.id, 'generation': gen1,
        }, format='json')
        self.assertEqual(res.status_code, 200, res.data)
        self.assertEqual(len(res.data), 3)
        self.assertEqual(res.data[0]['start_time'], '08:00:00')
        # gen1 active, gen2 archived now
        self.assertTrue(Period.objects.filter(school=self.school, generation=gen1, is_active=True).exists())
        self.assertTrue(Period.objects.filter(school=self.school, generation=gen2, is_active=False).exists())

    def test_restore_missing_generation_returns_404(self):
        self.auth(self.manager)
        res = self.client.post('/api/v1/schools/periods/restore/', {
            'school_id': self.school.id, 'generation': 999,
        }, format='json')
        self.assertEqual(res.status_code, 404)


class SectionClassTeacherAndOrderingTestCase(APITestCase):
    def setUp(self):
        self.manager = make_user('manager@test.com', 'school_admin', 'مدير المدرسة')
        self.teacher = make_user('teacher@test.com', 'teacher', 'معلم المدرسة')
        self.other_teacher = make_user('other@test.com', 'teacher', 'معلم خارج المدرسة')
        self.year = AcademicYear.objects.create(name='2025/2026', is_current=True)

        self.grade1 = Grade.objects.create(level=1, translations={'ar': {'name': 'الأول'}})
        self.grade2 = Grade.objects.create(level=2, translations={'ar': {'name': 'الثاني'}})
        self.grade3 = Grade.objects.create(level=3, translations={'ar': {'name': 'الثالث'}})

        self.school = School.objects.create(name='مدرسة', school_code='S1', manager=self.manager)
        SchoolTeacher.objects.create(school=self.school, teacher=self.teacher)
        for g in (self.grade1, self.grade2, self.grade3):
            SchoolGrade.objects.create(school=self.school, grade=g)

        # Create sections out of order (by pk) to verify grade-level ordering
        self.s3 = Section.objects.create(school=self.school, grade=self.grade3, academic_year=self.year, name='ب')
        self.s1 = Section.objects.create(school=self.school, grade=self.grade1, academic_year=self.year, name='أ')
        self.s2 = Section.objects.create(school=self.school, grade=self.grade2, academic_year=self.year, name='أ')

    def _auth(self, user):
        self.client.force_authenticate(user=user)

    def test_sections_ordered_by_grade_level_then_name(self):
        self._auth(self.manager)
        res = self.client.get('/api/v1/schools/sections/', {'school': self.school.id})
        self.assertEqual(res.status_code, 200, res.data)
        names = [s['id'] for s in res.data['results']]
        self.assertEqual(names, [self.s1.id, self.s2.id, self.s3.id])
    def test_set_class_teacher_and_get_name(self):
        self._auth(self.manager)
        res = self.client.patch(f'/api/v1/schools/sections/{self.s1.id}/', {
            'class_teacher': self.teacher.id,
        }, format='json')
        self.assertEqual(res.status_code, 200, res.data)
        self.assertEqual(res.data['class_teacher'], self.teacher.id)
        self.assertEqual(res.data['class_teacher_name'], 'معلم المدرسة')

    def test_reject_teacher_outside_school(self):
        self._auth(self.manager)
        res = self.client.patch(f'/api/v1/schools/sections/{self.s1.id}/', {
            'class_teacher': self.other_teacher.id,
        }, format='json')
        self.assertEqual(res.status_code, 400, res.data)

    def test_clear_class_teacher(self):
        Section.objects.filter(pk=self.s1.id).update(class_teacher=self.teacher)
        self._auth(self.manager)
        res = self.client.patch(f'/api/v1/schools/sections/{self.s1.id}/', {
            'class_teacher': None,
        }, format='json')
        self.assertEqual(res.status_code, 200, res.data)
        self.assertIsNone(res.data['class_teacher'])
        self.assertEqual(res.data['class_teacher_name'], '')

    def test_enrollment_includes_student_name(self):
        student = make_user('s@test.com', 'student', 'طالب الشعبة')
        StudentEnrollment.objects.create(student=student, section=self.s1, academic_year=self.year)
        self._auth(self.manager)
        res = self.client.get('/api/v1/schools/enrollments/', {'section': self.s1.id})
        self.assertEqual(res.status_code, 200, res.data)
        rows = res.data['results'] if isinstance(res.data, dict) and 'results' in res.data else res.data
        self.assertEqual(rows[0]['student_name'], 'طالب الشعبة')


class SchoolManagerStudentManagementTestCase(APITestCase):
    def setUp(self):
        self.admin = make_user('admin@test.com', 'admin', 'مدير النظام')
        self.manager = make_user('manager@test.com', 'school_admin', 'مدير المدرسة')
        self.other_manager = make_user('other-manager@test.com', 'school_admin', 'مدير مدرسة أخرى')
        self.student = make_user('student@test.com', 'student', 'طالب أول')
        self.other_student = make_user('other-student@test.com', 'student', 'طالب ثان')
        self.teacher = make_user('teacher@test.com', 'teacher', 'معلم المدرسة')

        self.year = AcademicYear.objects.create(name='2025/2026', is_current=True)
        self.next_year = AcademicYear.objects.create(name='2026/2027', is_current=False)

        self.grade1 = Grade.objects.create(level=1, translations={'ar': {'name': 'الأول'}})
        self.grade2 = Grade.objects.create(level=2, translations={'ar': {'name': 'الثاني'}})
        self.grade3 = Grade.objects.create(level=3, translations={'ar': {'name': 'الثالث'}})

        self.school = School.objects.create(name='المدرسة الأولى', school_code='SCH1', manager=self.manager)
        SchoolTeacher.objects.create(school=self.school, teacher=self.teacher)
        SchoolGrade.objects.create(school=self.school, grade=self.grade1)
        SchoolGrade.objects.create(school=self.school, grade=self.grade2)

        self.other_school = School.objects.create(name='المدرسة الثانية', school_code='SCH2', manager=self.other_manager)
        SchoolGrade.objects.create(school=self.other_school, grade=self.grade1)

        self.sec_a = Section.objects.create(school=self.school, grade=self.grade1, academic_year=self.year, name='أ')
        self.sec_b = Section.objects.create(school=self.school, grade=self.grade1, academic_year=self.year, name='ب')
        self.sec_grade2 = Section.objects.create(school=self.school, grade=self.grade2, academic_year=self.year, name='أ')
        self.other_sec = Section.objects.create(school=self.other_school, grade=self.grade1, academic_year=self.year, name='أ')

        StudentEnrollment.objects.create(student=self.student, section=self.sec_a, academic_year=self.year)
        StudentEnrollment.objects.create(student=self.other_student, section=self.other_sec, academic_year=self.year)

    def _auth(self, user):
        self.client.force_authenticate(user=user)

    def test_non_manager_cannot_write_enrollments(self):
        for user in (self.student, self.teacher):
            self._auth(user)
            res = self.client.post('/api/v1/schools/enrollments/', {
                'student': self.student.id, 'section': self.sec_a.id, 'academic_year': self.year.id,
            })
            self.assertEqual(res.status_code, 403, f'{user.role} should be forbidden')
            en = StudentEnrollment.objects.get(student=self.student, academic_year=self.year)
            res = self.client.delete(f'/api/v1/schools/enrollments/{en.id}/')
            self.assertEqual(res.status_code, 403, f'{user.role} should be forbidden to delete')

    def test_school_admin_enrolls_new_student(self):
        self._auth(self.manager)
        res = self.client.post(f'/api/v1/schools/sections/{self.sec_b.id}/enroll/', {
            'name': 'طالب جديد', 'email': 'new@test.com', 'phone': '0771000000',
        })
        self.assertEqual(res.status_code, 200, res.data)
        self.assertTrue(res.data['created_account'])
        self.assertFalse(res.data['moved'])
        user = User.objects.get(email='new@test.com')
        self.assertEqual(user.role, 'student')
        self.assertFalse(user.has_usable_password())
        en = StudentEnrollment.objects.get(student=user, academic_year=self.year)
        self.assertEqual(en.section, self.sec_b)

    def test_school_admin_enroll_existing_moves_with_confirmation(self):
        self._auth(self.manager)
        res = self.client.post(f'/api/v1/schools/sections/{self.sec_b.id}/enroll/', {
            'name': 'طالب أول', 'email': 'student@test.com',
        })
        self.assertEqual(res.status_code, 200, res.data)
        self.assertFalse(res.data['created_account'])
        self.assertTrue(res.data['moved'])
        self.assertEqual(res.data['moved_from'], 'أ')
        en = StudentEnrollment.objects.get(student=self.student, academic_year=self.year)
        self.assertEqual(en.section, self.sec_b)

    def test_enroll_creates_parent_link(self):
        self._auth(self.manager)
        res = self.client.post(f'/api/v1/schools/sections/{self.sec_b.id}/enroll/', {
            'name': 'طالب مع والد', 'email': 'kid@test.com', 'parent_email': 'parent@test.com',
        })
        self.assertEqual(res.status_code, 200, res.data)
        parent = User.objects.get(email='parent@test.com')
        self.assertEqual(parent.role, 'parent')
        self.assertTrue(FamilyLink.objects.filter(parent=parent, student__email='kid@test.com').exists())

    def test_manager_cannot_enroll_other_school(self):
        self._auth(self.manager)
        res = self.client.post(f'/api/v1/schools/sections/{self.other_sec.id}/enroll/', {
            'name': 'طالب', 'email': 'x@test.com',
        })
        self.assertEqual(res.status_code, 404, res.data)
        self.assertFalse(User.objects.filter(email='x@test.com').exists())

    def test_transfer_same_school_same_grade_allowed(self):
        self._auth(self.manager)
        en = StudentEnrollment.objects.get(student=self.student, academic_year=self.year)
        res = self.client.post(f'/api/v1/schools/enrollments/{en.id}/transfer/', {
            'target_section_id': self.sec_b.id,
        })
        self.assertEqual(res.status_code, 200, res.data)
        en.refresh_from_db()
        self.assertEqual(en.section, self.sec_b)

    def test_transfer_cross_grade_denied_for_manager(self):
        self._auth(self.manager)
        en = StudentEnrollment.objects.get(student=self.student, academic_year=self.year)
        res = self.client.post(f'/api/v1/schools/enrollments/{en.id}/transfer/', {
            'target_section_id': self.sec_grade2.id,
        })
        self.assertEqual(res.status_code, 400, res.data)
        en.refresh_from_db()
        self.assertEqual(en.section, self.sec_a)

    def test_transfer_cross_school_denied_for_manager(self):
        self._auth(self.manager)
        en = StudentEnrollment.objects.get(student=self.student, academic_year=self.year)
        res = self.client.post(f'/api/v1/schools/enrollments/{en.id}/transfer/', {
            'target_section_id': self.other_sec.id,
        })
        self.assertEqual(res.status_code, 400, res.data)
        en.refresh_from_db()
        self.assertEqual(en.section, self.sec_a)

    def test_manager_cannot_touch_other_school_enrollment(self):
        self._auth(self.manager)
        other_en = StudentEnrollment.objects.get(student=self.other_student, academic_year=self.year)
        res = self.client.post(f'/api/v1/schools/enrollments/{other_en.id}/transfer/', {
            'target_section_id': self.sec_a.id,
        })
        self.assertEqual(res.status_code, 404, res.data)
        res = self.client.delete(f'/api/v1/schools/enrollments/{other_en.id}/')
        self.assertEqual(res.status_code, 404, res.data)

    def test_admin_transfer_cross_grade_allowed(self):
        self._auth(self.admin)
        en = StudentEnrollment.objects.get(student=self.student, academic_year=self.year)
        res = self.client.post(f'/api/v1/schools/enrollments/{en.id}/transfer/', {
            'target_section_id': self.sec_grade2.id,
        })
        self.assertEqual(res.status_code, 200, res.data)
        en.refresh_from_db()
        self.assertEqual(en.section, self.sec_grade2)

    def test_delete_enrollment_own_school(self):
        self._auth(self.manager)
        en = StudentEnrollment.objects.get(student=self.student, academic_year=self.year)
        res = self.client.delete(f'/api/v1/schools/enrollments/{en.id}/')
        self.assertEqual(res.status_code, 204, res.data)
        self.assertFalse(StudentEnrollment.objects.filter(student=self.student, academic_year=self.year).exists())

    def test_promote_scoped_to_school(self):
        self._auth(self.manager)
        res = self.client.post(f'/api/v1/schools/academic-years/{self.year.id}/promote/', {
            'school_id': self.school.id, 'target_year_id': self.next_year.id,
        })
        self.assertEqual(res.status_code, 200, res.data)
        self.assertEqual(len(res.data['promoted']), 1)
        # own school promoted to grade2
        new_en = StudentEnrollment.objects.get(student=self.student, academic_year=self.next_year)
        self.assertEqual(new_en.section.grade, self.grade2)
        # other school untouched
        self.assertFalse(StudentEnrollment.objects.filter(student=self.other_student, academic_year=self.next_year).exists())

    def test_promote_copies_class_teacher_and_capacity(self):
        self.sec_a.class_teacher = self.teacher
        self.sec_a.capacity = 42
        self.sec_a.save(update_fields=['class_teacher', 'capacity'])
        self._auth(self.manager)
        res = self.client.post(f'/api/v1/schools/academic-years/{self.year.id}/promote/', {
            'school_id': self.school.id, 'target_year_id': self.next_year.id,
        })
        self.assertEqual(res.status_code, 200, res.data)
        new_sec = Section.objects.get(school=self.school, grade=self.grade2, academic_year=self.next_year, name='أ')
        self.assertEqual(new_sec.class_teacher, self.teacher)
        self.assertEqual(new_sec.capacity, 42)

    def test_promote_manager_requires_own_school(self):
        self._auth(self.manager)
        res = self.client.post(f'/api/v1/schools/academic-years/{self.year.id}/promote/', {
            'target_year_id': self.next_year.id,
        })
        self.assertEqual(res.status_code, 400, res.data)
        res = self.client.post(f'/api/v1/schools/academic-years/{self.year.id}/promote/', {
            'school_id': self.other_school.id, 'target_year_id': self.next_year.id,
        })
        self.assertEqual(res.status_code, 403, res.data)

    def test_school_admin_sees_all_years(self):
        self._auth(self.manager)
        res = self.client.get('/api/v1/schools/academic-years/')
        self.assertEqual(res.status_code, 200, res.data)
        years = res.data['results'] if isinstance(res.data, dict) and 'results' in res.data else res.data
        ids = {y['id'] for y in years}
        self.assertIn(self.year.id, ids)
        self.assertIn(self.next_year.id, ids)

    def test_school_admin_cannot_create_year(self):
        self._auth(self.manager)
        res = self.client.post('/api/v1/schools/academic-years/', {
            'name': '2030/2031', 'is_current': False,
        })
        self.assertEqual(res.status_code, 403, res.data)

    def test_sections_filter_by_year_and_grade(self):
        self._auth(self.manager)
        res = self.client.get('/api/v1/schools/sections/', {
            'school': self.school.id, 'academic_year': self.year.id, 'grade': self.grade1.id,
        })
        self.assertEqual(res.status_code, 200, res.data)
        ids = {s['id'] for s in res.data['results']}
        self.assertEqual(ids, {self.sec_a.id, self.sec_b.id})
        self.assertNotIn(self.sec_grade2.id, ids)


class ExcelAndClassTeacherTestCase(SchoolManagerStudentManagementTestCase):
    """School-manager/class-teacher Excel import-export + national-ID conflict handling."""

    def setUp(self):
        super().setUp()
        self.sec_a.class_teacher = self.teacher
        self.sec_a.save(update_fields=['class_teacher'])

    def _make_xlsx(self, headers, rows):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers)
        for row in rows:
            ws.append(row)
        buffer = io.BytesIO()
        wb.save(buffer)
        return SimpleUploadedFile(
            'students.xlsx', buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    def _read_xlsx(self, content):
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        return [list(r) for r in ws.iter_rows(values_only=True)]

    def _import_file(self, user, file, **extra):
        self._auth(user)
        data = {'kind': 'students', 'file': file}
        data.update(extra)
        return self.client.post('/api/v1/schools/bulk/import/', data, format='multipart')

    # ---- School-manager Excel import scoping ----

    def test_manager_import_xlsx_scoped_to_own_school(self):
        file = self._make_xlsx(
            ['اسم الطالب', 'الرقم الوطني', 'البريد', 'الهاتف', 'الصف', 'الشعبة', 'السنة الدراسية', 'الرمز المدرسي'],
            [['طالب استيراد', '9900111111', 'imp1@test.com', '0790000001', '1', 'ج', '2025/2026', 'SCH2']],
        )
        res = self._import_file(self.manager, file, school_id=self.school.id)
        self.assertEqual(res.status_code, 200, res.data)
        self.assertEqual(res.data['created'], 1)
        self.assertEqual(res.data['errors'], [])
        user = User.objects.get(email='imp1@test.com')
        self.assertEqual(user.national_id, '9900111111')
        en = StudentEnrollment.objects.get(student=user, academic_year=self.year)
        self.assertEqual(en.section.school, self.school)
        self.assertEqual(en.section.name, 'ج')
        # the file's SCH2 (another school) must have been ignored
        self.assertFalse(Section.objects.filter(school=self.other_school, name='ج', academic_year=self.year).exists())

    def test_manager_import_requires_own_school_id(self):
        file = self._make_xlsx(['اسم الطالب', 'البريد'], [['طالب', 'no-school@test.com']])
        res = self._import_file(self.manager, file)
        self.assertEqual(res.status_code, 403)
        self.assertFalse(User.objects.filter(email='no-school@test.com').exists())

    def test_manager_cannot_import_other_school(self):
        file = self._make_xlsx(['اسم الطالب', 'البريد'], [['طالب', 'other@test.com']])
        res = self._import_file(self.manager, file, school_id=self.other_school.id)
        self.assertEqual(res.status_code, 403)
        self.assertFalse(User.objects.filter(email='other@test.com').exists())

    def test_manager_cannot_import_teachers(self):
        file = self._make_xlsx(['اسم المعلم', 'البريد'], [['معلم', 't@test.com']])
        res = self._import_file(self.manager, file, school_id=self.school.id, kind='teachers')
        self.assertEqual(res.status_code, 403)

    def test_import_matches_existing_student_by_national_id(self):
        existing = make_user('migrate@test.com', 'student', 'طالب منقول', national_id='9900222222')
        StudentEnrollment.objects.create(student=existing, section=self.other_sec, academic_year=self.year)
        file = self._make_xlsx(
            ['اسم الطالب', 'الرقم الوطني', 'البريد', 'الصف', 'الشعبة', 'السنة الدراسية'],
            [['طالب منقول', '9900222222', 'different@test.com', '1', 'أ', '2025/2026']],
        )
        res = self._import_file(self.manager, file, school_id=self.school.id)
        self.assertEqual(res.status_code, 200, res.data)
        self.assertEqual(res.data['updated'], 1)
        self.assertFalse(User.objects.filter(email='different@test.com').exists())
        # already enrolled for the same year in another school: import is additive and keeps that enrollment
        enrollments = StudentEnrollment.objects.filter(student=existing, academic_year=self.year)
        self.assertEqual(enrollments.count(), 1)
        self.assertEqual(enrollments.first().section, self.other_sec)

    def test_import_enrolls_existing_student_without_enrollment(self):
        existing = make_user('no-enroll@test.com', 'student', 'طالب غير مسجل', national_id='9900777777')
        file = self._make_xlsx(
            ['اسم الطالب', 'الرقم الوطني', 'الصف', 'الشعبة', 'السنة الدراسية'],
            [['طالب غير مسجل', '9900777777', '1', 'أ', '2025/2026']],
        )
        res = self._import_file(self.manager, file, school_id=self.school.id)
        self.assertEqual(res.status_code, 200, res.data)
        self.assertEqual(res.data['updated'], 1)
        en = StudentEnrollment.objects.get(student=existing, academic_year=self.year)
        self.assertEqual(en.section.school, self.school)
        self.assertEqual(en.section, self.sec_a)

    def test_import_national_id_only_creates_placeholder_email(self):
        file = self._make_xlsx(
            ['اسم الطالب', 'الرقم الوطني', 'الصف', 'الشعبة', 'السنة الدراسية'],
            [['طالب بلا بريد', '9900333333', '1', 'ب', '2025/2026']],
        )
        res = self._import_file(self.manager, file, school_id=self.school.id)
        self.assertEqual(res.status_code, 200, res.data)
        self.assertEqual(res.data['created'], 1)
        user = User.objects.get(national_id='9900333333')
        self.assertEqual(user.email, 'student.9900333333@student.local')
        self.assertEqual(StudentEnrollment.objects.get(student=user).section, self.sec_b)

    # ---- School-manager Excel export scoping ----

    def test_manager_export_xlsx_scoped_to_school(self):
        self._auth(self.manager)
        res = self.client.get('/api/v1/schools/bulk/export/', {
            'kind': 'students', 'school_id': self.school.id,
        })
        self.assertEqual(res.status_code, 200)
        self.assertIn('spreadsheetml.sheet', res['Content-Type'])
        rows = self._read_xlsx(res.content)
        self.assertEqual(rows[0][0], 'البريد الإلكتروني')
        flat = '\n'.join('\t'.join(str(c or '') for c in r) for r in rows)
        self.assertIn('student@test.com', flat)
        self.assertIn('طالب أول', flat)
        self.assertNotIn('other-student@test.com', flat)

    def test_manager_export_template(self):
        self._auth(self.manager)
        res = self.client.get('/api/v1/schools/bulk/export/', {
            'kind': 'students', 'school_id': self.school.id, 'template': 1,
        })
        self.assertEqual(res.status_code, 200)
        rows = self._read_xlsx(res.content)
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[1][1], 'محمد أحمد')
        self.assertEqual(rows[1][2], '1000000000')

    def test_manager_cannot_export_other_school(self):
        self._auth(self.manager)
        res = self.client.get('/api/v1/schools/bulk/export/', {
            'kind': 'students', 'school_id': self.other_school.id,
        })
        self.assertEqual(res.status_code, 403)

    def test_manager_cannot_export_schools(self):
        self._auth(self.manager)
        res = self.client.get('/api/v1/schools/bulk/export/', {
            'kind': 'schools', 'school_id': self.school.id,
        })
        self.assertEqual(res.status_code, 403)

    # ---- National-ID conflict on enroll ----

    def test_enroll_cross_school_national_id_conflict(self):
        cross = make_user('x-school@test.com', 'student', 'طالب مدرسة أخرى', national_id='9900444444')
        StudentEnrollment.objects.create(student=cross, section=self.other_sec, academic_year=self.year)
        self._auth(self.manager)
        res = self.client.post(f'/api/v1/schools/sections/{self.sec_a.id}/enroll/', {
            'name': 'طالب مدرسة أخرى', 'national_id': '9900444444',
        })
        self.assertEqual(res.status_code, 409, res.data)
        self.assertEqual(res.data['conflict'], 'national_id')
        self.assertEqual(res.data['student']['email'], 'x-school@test.com')
        self.assertEqual(res.data['school']['id'], self.other_school.id)
        # not moved yet
        self.assertEqual(StudentEnrollment.objects.get(student=cross, academic_year=self.year).section, self.other_sec)

    def test_enroll_cross_school_national_id_confirmed_moves(self):
        cross = make_user('x-school@test.com', 'student', 'طالب مدرسة أخرى', national_id='9900444444')
        StudentEnrollment.objects.create(student=cross, section=self.other_sec, academic_year=self.year)
        self._auth(self.manager)
        res = self.client.post(f'/api/v1/schools/sections/{self.sec_a.id}/enroll/', {
            'name': 'طالب مدرسة أخرى', 'national_id': '9900444444', 'confirm': True,
        })
        self.assertEqual(res.status_code, 200, res.data)
        self.assertEqual(res.data['school_moved_from'], self.other_school.name)
        self.assertEqual(StudentEnrollment.objects.get(student=cross, academic_year=self.year).section, self.sec_a)

    def test_enroll_same_school_national_id_no_conflict(self):
        self.student.national_id = '9900555555'
        self.student.save(update_fields=['national_id'])
        self._auth(self.manager)
        res = self.client.post(f'/api/v1/schools/sections/{self.sec_b.id}/enroll/', {
            'name': 'طالب أول', 'national_id': '9900555555',
        })
        self.assertEqual(res.status_code, 200, res.data)
        self.assertFalse(res.data['created_account'])
        self.assertTrue(res.data['moved'])
        self.assertIsNone(res.data.get('school_moved_from'))
        self.assertEqual(StudentEnrollment.objects.get(student=self.student, academic_year=self.year).section, self.sec_b)

    # ---- Class-teacher (مربي الصف) permissions ----

    def test_class_teacher_sees_only_own_class_students(self):
        self._auth(self.teacher)
        res = self.client.get('/api/v1/schools/enrollments/')
        self.assertEqual(res.status_code, 200, res.data)
        rows = res.data['results'] if isinstance(res.data, dict) and 'results' in res.data else res.data
        emails = {r['student_email'] for r in rows}
        self.assertIn('student@test.com', emails)
        self.assertNotIn('other-student@test.com', emails)
        # a section the teacher does not mentor is not visible
        res = self.client.get('/api/v1/schools/enrollments/', {'section': self.sec_b.id})
        rows = res.data['results'] if isinstance(res.data, dict) and 'results' in res.data else res.data
        self.assertEqual(rows, [])

    def test_class_teacher_can_add_student_to_own_class(self):
        self._auth(self.teacher)
        res = self.client.post(f'/api/v1/schools/sections/{self.sec_a.id}/enroll/', {
            'name': 'طالب مربي الصف', 'email': 'classkid@test.com',
        })
        self.assertEqual(res.status_code, 200, res.data)
        self.assertTrue(User.objects.filter(email='classkid@test.com').exists())

    def test_class_teacher_cannot_add_to_unmentored_section(self):
        self._auth(self.teacher)
        res = self.client.post(f'/api/v1/schools/sections/{self.sec_b.id}/enroll/', {
            'name': 'طالب', 'email': 'nope@test.com',
        })
        self.assertEqual(res.status_code, 400)
        self.assertFalse(User.objects.filter(email='nope@test.com').exists())

    def test_class_teacher_records_attendance_own_class(self):
        self._auth(self.teacher)
        res = self.client.post('/api/v1/schools/attendances/bulk_record/', {
            'section': self.sec_a.id,
            'date': '2025-09-01',
            'records': [{'student': self.student.id, 'status': 'present'}],
        }, format='json')
        self.assertEqual(res.status_code, 200, res.data)
        self.assertEqual(Attendance.objects.get(student=self.student, date='2025-09-01').section, self.sec_a)

    def test_class_teacher_cannot_record_unmentored_section(self):
        self._auth(self.teacher)
        res = self.client.post('/api/v1/schools/attendances/bulk_record/', {
            'section': self.sec_b.id,
            'date': '2025-09-01',
            'records': [{'student': self.student.id, 'status': 'present'}],
        }, format='json')
        self.assertEqual(res.status_code, 403)

    def test_class_teacher_import_export_scoped_to_section(self):
        file = self._make_xlsx(
            ['اسم الطالب', 'الرقم الوطني', 'البريد'],
            [['طالب صف مربي', '9900666666', 'classimport@test.com']],
        )
        res = self._import_file(self.teacher, file, section_id=self.sec_a.id)
        self.assertEqual(res.status_code, 200, res.data)
        self.assertEqual(res.data['created'], 1)
        user = User.objects.get(email='classimport@test.com')
        self.assertEqual(StudentEnrollment.objects.get(student=user, academic_year=self.year).section, self.sec_a)

        self._auth(self.teacher)
        res = self.client.get('/api/v1/schools/bulk/export/', {
            'kind': 'students', 'section_id': self.sec_a.id,
        })
        self.assertEqual(res.status_code, 200)
        rows = self._read_xlsx(res.content)
        flat = '\n'.join('\t'.join(str(c or '') for c in r) for r in rows)
        self.assertIn('student@test.com', flat)
        self.assertIn('classimport@test.com', flat)
        self.assertNotIn('other-student@test.com', flat)

    def test_sections_filter_by_class_teacher(self):
        self._auth(self.admin)
        res = self.client.get('/api/v1/schools/sections/', {'class_teacher': self.teacher.id})
        self.assertEqual(res.status_code, 200, res.data)
        ids = {s['id'] for s in res.data['results']}
        self.assertEqual(ids, {self.sec_a.id})


class SchoolCalendarTestCase(APITestCase):
    """Tests for per-school week start / working days configuration."""

    def setUp(self):
        self.admin = make_user('caladmin@test.com', 'admin', 'مدير النظام')
        self.manager = make_user('calmanager@test.com', 'school_admin', 'مدير المدرسة', phone='0773000001')
        self.grade = Grade.objects.create(level=4, translations={'ar': {'name': 'الصف الرابع'}})
        self.subject = Subject.objects.create(translations={'ar': {'name': 'العلوم'}})
        self.year = AcademicYear.objects.create(name='2026/2027', is_current=True)
        self.school = School.objects.create(
            name='مدرسة التقويم', school_code='CAL-1', manager=self.manager,
        )
        SchoolGrade.objects.create(school=self.school, grade=self.grade)
        self.section = Section.objects.create(
            school=self.school, grade=self.grade, academic_year=self.year, name='أ'
        )
        self.period = Period.objects.create(
            school=self.school, name='الحصة 1', period_number=1,
            start_time='08:00', end_time='08:45',
        )

    def _auth(self, user):
        self.client.force_authenticate(user=user)

    def test_defaults_are_jordanian(self):
        self.assertEqual(DayOfWeek.SUNDAY, 7)
        self.assertEqual(DayOfWeek.MONDAY, 1)
        self.assertEqual(DEFAULT_WEEK_START, 7)
        self.assertEqual(DEFAULT_WORKING_DAYS, [7, 1, 2, 3, 4])
        self.assertEqual(self.school.week_start, 7)
        self.assertEqual(list(self.school.working_days), [7, 1, 2, 3, 4])

    def test_manager_can_update_calendar(self):
        self._auth(self.manager)
        res = self.client.patch(
            f'/api/v1/schools/schools/{self.school.id}/calendar/',
            {'week_start': 1, 'working_days': [1, 2, 3, 4, 5]},
            format='json',
        )
        self.assertEqual(res.status_code, 200, res.data)
        self.school.refresh_from_db()
        self.assertEqual(self.school.week_start, 1)
        self.assertEqual(list(self.school.working_days), [1, 2, 3, 4, 5])
        self.assertEqual(res.data['week_start'], 1)

    def test_other_manager_cannot_update_calendar(self):
        outsider = make_user('caloutsider@test.com', 'school_admin', 'مدير آخر', phone='0773000002')
        self._auth(outsider)
        res = self.client.patch(
            f'/api/v1/schools/schools/{self.school.id}/calendar/',
            {'week_start': 1, 'working_days': [1, 2, 3, 4, 5]},
            format='json',
        )
        self.assertIn(res.status_code, [403, 404])

    def test_invalid_days_rejected(self):
        self._auth(self.manager)
        res = self.client.patch(
            f'/api/v1/schools/schools/{self.school.id}/calendar/',
            {'week_start': 9, 'working_days': [1, 2, 3, 4, 5]},
            format='json',
        )
        self.assertEqual(res.status_code, 400)
        res = self.client.patch(
            f'/api/v1/schools/schools/{self.school.id}/calendar/',
            {'week_start': 1, 'working_days': []},
            format='json',
        )
        self.assertEqual(res.status_code, 400)
        res = self.client.patch(
            f'/api/v1/schools/schools/{self.school.id}/calendar/',
            {'week_start': 2, 'working_days': [1, 3, 4, 5]},
            format='json',
        )
        self.assertEqual(res.status_code, 400)

    def test_auto_schedule_respects_working_days(self):
        teacher = make_user('caltacher@test.com', 'teacher', 'معلم', phone='0773000003')
        subject2 = Subject.objects.create(translations={'ar': {'name': 'الرياضيات'}})
        TeacherAssignment.objects.create(
            teacher=teacher, section=self.section, subject=self.subject, academic_year=self.year,
        )
        TeacherAssignment.objects.create(
            teacher=teacher, section=self.section, subject=subject2, academic_year=self.year,
        )
        self.school.rooms.create(name='قاعة 1', code='R1')
        self._auth(self.admin)
        # Gulf week: Monday-Friday, starts Monday.
        self.school.week_start = 1
        self.school.working_days = [1, 2, 3, 4, 5]
        self.school.save()
        res = self.client.post('/api/v1/schools/timetable-slots/auto_schedule/', {
            'school_id': self.school.id, 'academic_year_id': self.year.id,
        }, format='json')
        self.assertEqual(res.status_code, 200, res.data)
        days = set(TimetableSlot.objects.filter(section=self.section).values_list('day_of_week', flat=True))
        self.assertTrue(days.issubset({1, 2, 3, 4, 5}))
        self.assertGreater(len(days), 1)
        self.assertNotIn(7, days)  # Sunday not a working day

    def test_attendance_warning_on_non_working_day(self):
        student = make_user('calstudent@test.com', 'student', 'طالب', national_id='1234567892', phone='0773000004')
        StudentEnrollment.objects.create(student=student, section=self.section, academic_year=self.year)
        self.school.week_start = 1
        self.school.working_days = [1, 2, 3, 4, 5]
        self.school.save()
        self._auth(self.manager)
        # 2026-08-16 is a Sunday (non-working for this school).
        res = self.client.post('/api/v1/schools/attendances/', {
            'student': student.id, 'section': self.section.id,
            'school': self.school.id,
            'date': '2026-08-16', 'status': 'present',
        }, format='json')
        self.assertEqual(res.status_code, 201, res.data)
        self.assertIn('warning', res.data)
        # Still recorded (warning only).
        self.assertTrue(Attendance.objects.filter(student=student, date='2026-08-16').exists())

    def test_bulk_attendance_warning_on_non_working_day(self):
        student = make_user('calbulk@test.com', 'student', 'طالب', national_id='1234567893', phone='0773000005')
        StudentEnrollment.objects.create(student=student, section=self.section, academic_year=self.year)
        self.school.week_start = 1
        self.school.working_days = [1, 2, 3, 4, 5]
        self.school.save()
        self._auth(self.manager)
        res = self.client.post('/api/v1/schools/attendances/bulk_record/', {
            'section': self.section.id,
            'date': '2026-08-16',  # Sunday
            'records': [{'student': student.id, 'status': 'absent'}],
        }, format='json')
        self.assertEqual(res.status_code, 200, res.data)
        self.assertIsNotNone(res.data['warning'])

    def test_day_display_localized(self):
        TimetableSlot.objects.create(
            school=self.school, academic_year=self.year, section=self.section,
            day_of_week=7, period=self.period, subject=self.subject,
            teacher=self.admin, room=None,
        )
        self._auth(self.manager)
        res = self.client.get('/api/v1/schools/timetable-slots/', {'locale': 'en'})
        self.assertEqual(res.status_code, 200)
        self.assertEqual(res.data['results'][0]['day_display'], 'Sunday')
        res = self.client.get('/api/v1/schools/timetable-slots/', {'locale': 'ar'})
        self.assertEqual(res.data['results'][0]['day_display'], 'الأحد')


class SchoolAuxiliaryFeatureGatingTestCase(APITestCase):
    def setUp(self):
        self.admin = make_user('admin@test.com', 'admin', 'مدير النظام')
        self.free_manager = make_user('freemanager@test.com', 'school_admin', 'مدير مجاني', subscription_plan='free')
        self.school_manager = make_user('schoolmanager@test.com', 'school_admin', 'مدير مدرسي', subscription_plan='school')
        
        self.free_school = School.objects.create(name='مدرسة مجانية', school_code='FREE1', manager=self.free_manager)
        self.school_school = School.objects.create(name='مدرسة مدرسية', school_code='SCHOOL1', manager=self.school_manager)
        self.grade = Grade.objects.create(level=1, translations={'ar': {'name': 'الأول'}})
        self.year = AcademicYear.objects.create(name='2025/2026', is_current=True)
        self.section = Section.objects.create(school=self.school_school, grade=self.grade, academic_year=self.year, name='أ')
        self.bus = SchoolBus.objects.create(school=self.school_school, bus_number='101', driver_name='سائق')

    def auth(self, user):
        self.client.force_authenticate(user=user)

    def test_free_manager_cannot_create_auxiliary_modules(self):
        self.auth(self.free_manager)
        
        # 1. Fees
        res = self.client.post('/api/v1/schools/fees/', {'school': self.free_school.id, 'title': 'رسوم', 'amount': 100})
        self.assertEqual(res.status_code, 403)
        self.assertIn('subscription_required', res.data.get('error', ''))

        # 2. Buses
        res = self.client.post('/api/v1/schools/buses/', {'school': self.free_school.id, 'bus_number': '102', 'driver_name': 'سائق'})
        self.assertEqual(res.status_code, 403)

        # 3. Books
        res = self.client.post('/api/v1/schools/books/', {'school': self.free_school.id, 'title': 'كتاب'})
        self.assertEqual(res.status_code, 403)

    def test_school_manager_can_create_auxiliary_modules(self):
        self.auth(self.school_manager)
        
        # 1. Fees
        res = self.client.post('/api/v1/schools/fees/', {'school': self.school_school.id, 'title': 'رسوم دراسية', 'amount': 250})
        self.assertEqual(res.status_code, 201)

        # 2. Buses
        res = self.client.post('/api/v1/schools/buses/', {'school': self.school_school.id, 'bus_number': '202', 'driver_name': 'سائق 2'})
        self.assertEqual(res.status_code, 201)
        bus_id = res.data['id']

        # 3. Routes
        res = self.client.post('/api/v1/schools/bus-routes/', {'bus': bus_id, 'route_name': 'خط 1'})
        self.assertEqual(res.status_code, 201)

        # 4. Books
        res = self.client.post('/api/v1/schools/books/', {'school': self.school_school.id, 'title': 'رياضيات الصف الأول', 'total_copies': 10})
        self.assertEqual(res.status_code, 201)
        book_id = res.data['id']

        # 5. Library Lendings
        res = self.client.post('/api/v1/schools/library-lendings/', {'book': book_id, 'borrower_name': 'طالب'})
        self.assertEqual(res.status_code, 201)

    def test_admin_override_can_create_auxiliary_modules(self):
        self.auth(self.admin)
        res = self.client.post('/api/v1/schools/fees/', {'school': self.free_school.id, 'title': 'رسوم أدمن', 'amount': 50})
        self.assertEqual(res.status_code, 201)
