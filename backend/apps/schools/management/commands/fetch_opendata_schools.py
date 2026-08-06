"""جلب بيانات المدارس من بوابة البيانات الحكومية المفتوحة (opendata.gov.jo) واستيرادها.

المصدر: مجموعة بيانات "بيانات المدارس" لوزارة التربية والتعليم الأردنية.
الاستخدام:
  python manage.py fetch_opendata_schools            # تحميل واستيراد من البوابة
  python manage.py fetch_opendata_schools --file path.xlsx   # استيراد من ملف محلي
"""
import io
import urllib.request

from django.core.management.base import BaseCommand

from apps.schools.models import School

BULK_BATCH = 2000

DATASET_URL = (
    'https://opendata.gov.jo/dataset/MIGRATED-162-2018/'
    'resource/3d84b559-727b-4935-b45c-c0346b553663/download/-.xlsx'
)

HEADER_MAP = {
    'رمز المؤسسة': 'school_code',
    'اسم المؤسسة': 'name',
    'المديرية': 'directorate',
    'المحافظة': 'governorate',
    'الإقليم': 'region',
    'جنس المؤسس': 'gender',
    'نوع التعليم': 'education_type',
    'العنوان': 'address',
}


class Command(BaseCommand):
    help = 'Fetch schools from the Jordanian open-data portal (opendata.gov.jo) and import them.'

    def add_arguments(self, parser):
        parser.add_argument('--url', default=DATASET_URL, help='XLSX download URL for the schools dataset.')
        parser.add_argument('--file', default='', help='Local XLSX path to import instead of downloading.')
        parser.add_argument('--dry-run', action='store_true', help='Count rows without writing to the database.')

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError:
            self.stderr.write(self.style.ERROR('openpyxl is required (pip install openpyxl)'))
            return

        local_path = options['file']
        if local_path:
            self.stdout.write(f'Reading local file {local_path} ...')
            with open(local_path, 'rb') as fh:
                raw = fh.read()
        else:
            url = options['url']
            self.stdout.write(f'Downloading dataset from {url} ...')
            raw = urllib.request.urlopen(url, timeout=120).read()
            self.stdout.write(f'Downloaded {len(raw) / 1024:.0f} KB')

        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        header = [str(c).strip() if c is not None else '' for c in next(rows)]

        created = updated = skipped = 0
        records = []
        for values in rows:
            record = {header[i]: (values[i] if i < len(values) else None) for i in range(len(header))}
            school_code = str(record.get('رمز المؤسسة') or '').strip()
            name = str(record.get('اسم المؤسسة') or '').strip()
            if not school_code or not name:
                skipped += 1
                continue
            data = {'name': name, 'translations': {'ar': {'name': name}}}
            for src, field in HEADER_MAP.items():
                if field == 'school_code':
                    continue
                value = record.get(src)
                if value is None:
                    continue
                data[field] = str(value).strip() if not isinstance(value, float) or not value.is_integer() else str(int(value))
            records.append((school_code, data))

        if options['dry_run']:
            self.stdout.write(self.style.SUCCESS(f'Dry run: {len(records)} rows ready, {skipped} skipped.'))
            return

        existing = {
            s.school_code: s
            for s in School.objects.filter(school_code__in=[code for code, _ in records])
        }
        to_create = []
        to_update = []
        for code, data in records:
            obj = existing.get(code)
            if obj:
                for field, value in data.items():
                    setattr(obj, field, value)
                to_update.append(obj)
            else:
                to_create.append(School(school_code=code, **data))

        for i in range(0, len(to_create), BULK_BATCH):
            School.objects.bulk_create(to_create[i:i + BULK_BATCH], batch_size=BULK_BATCH)
        created = len(to_create)

        update_fields = ['name', 'directorate', 'governorate', 'region', 'gender', 'education_type', 'address', 'translations']
        for i in range(0, len(to_update), BULK_BATCH):
            School.objects.bulk_update(to_update[i:i + BULK_BATCH], update_fields, batch_size=BULK_BATCH)
        updated = len(to_update)

        self.stdout.write(self.style.SUCCESS(
            f'Done: {created} created, {updated} updated, {skipped} skipped.'
        ))
