import csv
import io
import json
from datetime import date, timedelta

from django.conf import settings
from django.db import transaction
from django.db.models import Count, Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from rest_framework import permissions, serializers, status, viewsets
from rest_framework.decorators import action
from rest_framework.pagination import PageNumberPagination
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.users.models import User
from apps.users.services import RoleService

from .absence import notify_absence
from .models import (
    DEFAULT_WEEK_START,
    DEFAULT_WORKING_DAYS,
    FAQ,
    AcademicYear,
    AnnouncementReadReceipt,
    Assignment,
    AssignmentSubmission,
    Attachment,
    Attendance,
    Book,
    BusLocationLog,
    BusRoute,
    DayOfWeek,
    DeviceEvent,
    FamilyLink,
    GradeCategory,
    GradeEntry,
    LibraryLending,
    ParentTeacherTicket,
    Period,
    Room,
    School,
    SchoolAnnouncement,
    SchoolBus,
    SchoolDevice,
    SchoolFee,
    SchoolGrade,
    SchoolManagerRequest,
    SchoolStaff,
    SchoolSubjectPeriod,
    SchoolTeacher,
    Section,
    StudentBusAssignment,
    StudentEnrollment,
    StudentFeeAssignment,
    SupportRequest,
    TeacherAssignment,
    TimetableSlot,
    UserAISetting,
    WeeklyReport,
    WhatsAppNotificationLog,
)
from .serializers import (
    AcademicYearSerializer,
    AssignmentSerializer,
    AssignmentSubmissionSerializer,
    AttachmentSerializer,
    AttendanceSerializer,
    BookSerializer,
    BusRouteSerializer,
    FamilyLinkSerializer,
    FAQSerializer,
    GradeCategorySerializer,
    GradeEntrySerializer,
    LibraryLendingSerializer,
    ParentTeacherTicketSerializer,
    PeriodSerializer,
    RoomSerializer,
    SchoolAnnouncementSerializer,
    SchoolBusSerializer,
    SchoolDeviceSerializer,
    SchoolFeeSerializer,
    SchoolGradeSerializer,
    SchoolManagerRequestCreateSerializer,
    SchoolManagerRequestReviewSerializer,
    SchoolManagerRequestSerializer,
    SchoolSerializer,
    SchoolStaffSerializer,
    SchoolSubjectPeriodSerializer,
    SchoolTeacherCreateSerializer,
    SchoolTeacherSerializer,
    SectionSerializer,
    StudentBusAssignmentSerializer,
    StudentEnrollmentSerializer,
    StudentFeeAssignmentSerializer,
    SupportRequestSerializer,
    TeacherAssignmentSerializer,
    TimetableSlotSerializer,
    WeeklyReportSerializer,
)
from .whatsapp import send_whatsapp_alert


def is_admin(user):
    """System admin or dev-team member with schools-section access."""
    if not (user and user.is_authenticated):
        return False
    if user.is_staff or user.is_superuser:
        return True
    from apps.users.permissions import SECTION_ROLES
    allowed = SECTION_ROLES.get('schools', {'admin'})
    user_roles = RoleService.get_role_names(user)
    return any(r in allowed for r in user_roles)


def is_teacher(user):
    return bool(user and user.is_authenticated and RoleService.has_role(user, 'teacher'))


def is_school_admin(user):
    """School manager: full permissions on their own school(s) only."""
    return bool(user and user.is_authenticated and RoleService.has_role(user, 'school_admin'))


def ordered_working_days(week_start, working_days):
    """Return working days (ISO 1=Mon..7=Sun) sorted starting from the given week_start day."""
    ordered = sorted(working_days, key=lambda d: (d - week_start) % 7)
    return [d for d in ordered if d in DayOfWeek.values]


def school_week_days(school):
    """Resolve a school's working days ordered from its week start (fallback to defaults)."""
    days = list(school.working_days or []) if hasattr(school, 'working_days') else []
    if not days:
        days = list(DEFAULT_WORKING_DAYS)
    return ordered_working_days(school.week_start or DEFAULT_WEEK_START, days)


def week_start_date_for(school, date):
    """Return the date of the week's first day for the given school (ISO week_start)."""
    offset = (date.isoweekday() - (school.week_start or DEFAULT_WEEK_START)) % 7
    return date - timedelta(days=offset)


def working_day_warning(school, date):
    """Return an Arabic warning string if `date` is not a working day for the school, else None."""
    if not school:
        return None
    days = list(school.working_days or DEFAULT_WORKING_DAYS)
    if date.isoweekday() not in days:
        return 'هذا اليوم ليس من أيام الدوام الدراسي للمدرسة (تحذير فقط).'
    return None


def user_section_ids(user):
    """Returns a set of section ids the user is linked to, or None for admins (no filtering).
    Teachers are linked to the sections they teach (TeacherAssignment) plus the sections
    they mentor as class teacher (مربي الصف)."""
    if not user or not user.is_authenticated:
        return set()
    if RoleService.has_role(user, 'teacher'):
        section_ids = set(TeacherAssignment.objects.filter(teacher=user).values_list('section_id', flat=True))
        section_ids |= set(Section.objects.filter(class_teacher=user).values_list('id', flat=True))
        return section_ids
    if RoleService.has_role(user, 'student'):
        return set(StudentEnrollment.objects.filter(student=user).values_list('section_id', flat=True))
    if RoleService.has_role(user, 'parent'):
        child_ids = FamilyLink.objects.filter(parent=user).values_list('student_id', flat=True)
        return set(StudentEnrollment.objects.filter(student_id__in=child_ids).values_list('section_id', flat=True))
    if RoleService.has_role(user, 'school_admin'):
        return set(Section.objects.filter(school__manager=user).values_list('id', flat=True))
    return set()


def class_teacher_section_ids(user):
    """Section ids where the user is the class mentor (مربي الصف)."""
    if not (user and user.is_authenticated) or not RoleService.has_role(user, 'teacher'):
        return set()
    return set(Section.objects.filter(class_teacher=user).values_list('id', flat=True))


def user_school_ids(user):
    if not user or not user.is_authenticated:
        return set()
    if RoleService.has_role(user, 'school_admin'):
        return set(user.managed_schools.values_list('id', flat=True))
    section_ids = user_section_ids(user)
    if section_ids is None:
        return None
    return set(Section.objects.filter(id__in=section_ids).values_list('school_id', flat=True))


class IsAdminOrReadOnly(permissions.BasePermission):
    """Read for anyone; writes for system admins or school managers (scoped via get_queryset)."""

    def has_permission(self, request, view):
        if request.method in permissions.SAFE_METHODS:
            return True
        return is_admin(request.user) or is_school_admin(request.user)


class CanManageAnnouncements(permissions.BasePermission):
    """Allows any authenticated user to read; admins, school managers and teachers can create announcements."""

    def has_permission(self, request, view):
        if request.method in permissions.SAFE_METHODS:
            return bool(request.user and request.user.is_authenticated)
        return is_admin(request.user) or is_teacher(request.user) or is_school_admin(request.user)


class CanManageEnrollments(permissions.BasePermission):
    """Enrollment writes (create/update/delete/transfer) only for system admins or
    school managers, and only within the school(s) they manage."""

    def has_permission(self, request, view):
        if request.method in permissions.SAFE_METHODS:
            return bool(request.user and request.user.is_authenticated)
        if not (request.user and request.user.is_authenticated):
            return False
        if is_admin(request.user):
            return True
        if not is_school_admin(request.user):
            return False
        if view.action == 'create':
            raw = request.data.get('section')
            if isinstance(raw, dict):
                raw = raw.get('id')
            if raw is None:
                return False
            return Section.objects.filter(id=raw, school__manager=request.user).exists()
        # Detail actions (update/destroy/transfer) rely on has_object_permission.
        return True

    def has_object_permission(self, request, view, obj):
        if request.method in permissions.SAFE_METHODS:
            return bool(request.user and request.user.is_authenticated)
        if is_admin(request.user):
            return True
        if is_school_admin(request.user):
            school_ids = user_school_ids(request.user)
            return obj.section.school_id in school_ids if school_ids else False
        return False


class IsAdminOrReadOnlyYears(permissions.BasePermission):
    """Academic years are global: anyone may read; only system admins may create/update/delete."""

    def has_permission(self, request, view):
        if request.method in permissions.SAFE_METHODS:
            return True
        return is_admin(request.user)


class IsAdminRole(permissions.BasePermission):
    """Allows access only for users with the admin role (or staff), consistent with is_admin()."""

    def has_permission(self, request, view):
        return is_admin(request.user)


class IsAdminOrSchoolManager(permissions.BasePermission):
    """Allows system admins unrestricted; school managers when `school_id` targets one
    of their schools; class teachers when `section_id` targets a section they mentor.

    Used by the bulk import/export endpoints so school staff can exchange student
    data via Excel without crossing school boundaries.
    """

    def has_permission(self, request, view):
        if not (request.user and request.user.is_authenticated):
            return False
        if is_admin(request.user):
            return True
        if is_school_admin(request.user):
            raw = self._scoped_id(request, 'school_id')
            if not raw:
                return False
            school_ids = user_school_ids(request.user)
            return bool(school_ids) and int(raw) in school_ids
        if is_teacher(request.user):
            raw = self._scoped_id(request, 'section_id')
            if not raw:
                return False
            return Section.objects.filter(id=int(raw), class_teacher=request.user).exists()
        return False

    @staticmethod
    def _scoped_id(request, name):
        raw = request.data.get(name) if request.method == 'POST' else request.query_params.get(name)
        if raw in (None, ''):
            return None
        try:
            return int(raw)
        except (TypeError, ValueError):
            return None


class SchoolPagination(PageNumberPagination):
    page_size = 20
    page_size_query_param = 'page_size'
    max_page_size = 200


class SchoolViewSet(viewsets.ModelViewSet):
    queryset = School.objects.all()
    serializer_class = SchoolSerializer
    permission_classes = [IsAdminOrReadOnly]
    pagination_class = SchoolPagination

    def get_queryset(self):
        if is_admin(self.request.user):
            qs = School.objects.all()
        else:
            school_ids = user_school_ids(self.request.user)
            if not school_ids:
                qs = School.objects.none()
            else:
                qs = School.objects.filter(id__in=school_ids)

        search = self.request.query_params.get('search')
        if search:
            qs = qs.filter(
                Q(name__icontains=search) |
                Q(school_code__icontains=search) |
                Q(directorate__icontains=search) |
                Q(governorate__icontains=search) |
                Q(region__icontains=search) |
                Q(address__icontains=search)
            )
        directorate = self.request.query_params.get('directorate')
        if directorate:
            qs = qs.filter(directorate__icontains=directorate)

        ordering = self.request.query_params.get('ordering', 'id')
        if ordering not in ['id', '-id', 'name', '-name', 'school_code', '-school_code', 'directorate', '-directorate']:
            ordering = 'id'
        return qs.order_by(ordering)

    def create(self, request, *args, **kwargs):
        # Only system admins may create schools (school managers can't create new ones)
        if not is_admin(request.user):
            return Response({'detail': 'Only system admins can create schools'}, status=status.HTTP_403_FORBIDDEN)
        return super().create(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        if not is_admin(request.user):
            return Response({'detail': 'Only system admins can delete schools'}, status=status.HTTP_403_FORBIDDEN)
        return super().destroy(request, *args, **kwargs)

    @action(detail=True, methods=['patch'], permission_classes=[permissions.IsAuthenticated])
    def calendar(self, request, pk=None):
        """Update a school's week start and working days (system admin or the school's manager)."""
        school = self.get_object()
        if not is_admin(request.user) and not (is_school_admin(request.user) and school.manager_id == request.user.id):
            return Response({'detail': 'Permission denied for this school'}, status=status.HTTP_403_FORBIDDEN)

        week_start = request.data.get('week_start', school.week_start)
        working_days = request.data.get('working_days')
        if working_days is not None:
            try:
                working_days = [int(d) for d in working_days]
            except (TypeError, ValueError):
                return Response({'error': 'working_days must be a list of day numbers (1=Mon..7=Sun)'}, status=status.HTTP_400_BAD_REQUEST)
        else:
            working_days = list(school.working_days or [])

        valid_days = set(DayOfWeek.values)
        if week_start not in valid_days:
            return Response({'error': 'week_start must be in 1..7 (ISO)'}, status=status.HTTP_400_BAD_REQUEST)
        if not working_days or any(d not in valid_days for d in working_days):
            return Response({'error': 'working_days must be a non-empty list of day numbers in 1..7'}, status=status.HTTP_400_BAD_REQUEST)
        if week_start not in working_days:
            return Response({'error': 'week_start must be one of the working days'}, status=status.HTTP_400_BAD_REQUEST)

        school.week_start = week_start
        school.working_days = list(dict.fromkeys(working_days))
        school.save(update_fields=['week_start', 'working_days'])
        return Response(SchoolSerializer(school, context=self.get_serializer_context()).data)

    def perform_update(self, serializer):
        old_manager_id = self.get_object().manager_id
        school = serializer.save()
        new_manager_id = school.manager_id
        if old_manager_id == new_manager_id:
            return
        if new_manager_id:
            new_manager = school.manager
            if new_manager.role not in User.ADMIN_ROLES and not RoleService.has_role(new_manager, 'school_admin'):
                new_manager.role = 'school_admin'
                new_manager.save(update_fields=['role'])
        if old_manager_id:
            old_manager = User.objects.filter(pk=old_manager_id).first()
            if old_manager and RoleService.has_role(old_manager, 'school_admin') and not old_manager.managed_schools.exists():
                old_manager.role = 'student'
                old_manager.save(update_fields=['role'])


class AcademicYearViewSet(viewsets.ModelViewSet):
    queryset = AcademicYear.objects.all()
    serializer_class = AcademicYearSerializer
    permission_classes = [IsAdminOrReadOnlyYears]

    def get_queryset(self):
        if is_admin(self.request.user) or is_school_admin(self.request.user):
            return AcademicYear.objects.all()
        section_ids = user_section_ids(self.request.user)
        if not section_ids:
            return AcademicYear.objects.none()
        year_ids = Section.objects.filter(id__in=section_ids).values_list('academic_year_id', flat=True)
        return AcademicYear.objects.filter(id__in=year_ids)

    @action(detail=True, methods=['post'], permission_classes=[permissions.IsAuthenticated])
    def promote(self, request, pk=None):
        """Annual promotion (الترفيع السنوي): moves enrollments of this year
        into the next academic year and next grade, keeping history archived.
        Also migrates teacher assignments and optionally flips is_current.
        Supports dry_run=true to preview without executing."""
        source_year = self.get_object()

        user = request.user
        school_id = request.data.get('school_id')
        if not is_admin(user):
            if not is_school_admin(user):
                return Response({'error': 'غير مصرح لك بإجراء الترفيع'}, status=status.HTTP_403_FORBIDDEN)
            if not school_id:
                return Response({'error': 'school_id is required for school managers'}, status=status.HTTP_400_BAD_REQUEST)
            school_ids = user_school_ids(user)
            if int(school_id) not in school_ids:
                return Response({'error': 'يمكنك ترفيع مدرستك فقط'}, status=status.HTTP_403_FORBIDDEN)

        target_year_id = request.data.get('target_year_id')
        dry_run = request.data.get('dry_run', False)
        if not target_year_id:
            return Response({'error': 'target_year_id is required'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            target_year = AcademicYear.objects.get(id=target_year_id)
        except AcademicYear.DoesNotExist:
            return Response({'error': 'target year not found'}, status=status.HTTP_404_NOT_FOUND)

        enrollments = StudentEnrollment.objects.filter(
            academic_year=source_year,
        ).select_related('student', 'section', 'section__school', 'section__grade')
        if school_id:
            enrollments = enrollments.filter(section__school_id=school_id)

        promoted = []
        skipped = []
        sections_created = 0
        teachers_migrated = 0

        def _do_promote():
            nonlocal sections_created, teachers_migrated
            for en in enrollments:
                new_grade = self._next_grade(en.section.grade)
                if new_grade is None:
                    skipped.append({
                        'student': en.student.email,
                        'student_name': en.student.translations.get('ar', {}).get('name', en.student.email),
                        'reason': 'graduated — no next grade',
                    })
                    continue
                new_section, was_created = Section.objects.get_or_create(
                    school=en.section.school,
                    grade=new_grade,
                    academic_year=target_year,
                    name=en.section.name,
                    defaults={'class_teacher': en.section.class_teacher, 'capacity': en.section.capacity},
                )
                if was_created:
                    sections_created += 1
                _, was_enrollment_created = StudentEnrollment.objects.get_or_create(
                    student=en.student,
                    academic_year=target_year,
                    defaults={'section': new_section},
                )
                promoted.append({
                    'student': en.student.email,
                    'student_name': en.student.translations.get('ar', {}).get('name', en.student.email),
                    'from_section': str(en.section),
                    'to_section': str(new_section),
                })

            # Migrate teacher assignments from source year to target year
            source_assignments = TeacherAssignment.objects.filter(
                academic_year=source_year,
            ).select_related('teacher', 'section', 'subject', 'section__school', 'section__grade')
            if school_id:
                source_assignments = source_assignments.filter(section__school_id=school_id)

            for ta in source_assignments:
                new_grade = self._next_grade(ta.section.grade)
                if new_grade is None:
                    continue
                new_section, _ = Section.objects.get_or_create(
                    school=ta.section.school,
                    grade=new_grade,
                    academic_year=target_year,
                    name=ta.section.name,
                    defaults={'class_teacher': ta.section.class_teacher, 'capacity': ta.section.capacity},
                )
                _, was_ta_created = TeacherAssignment.objects.get_or_create(
                    teacher=ta.teacher,
                    section=new_section,
                    subject=ta.subject,
                    academic_year=target_year,
                )
                if was_ta_created:
                    teachers_migrated += 1

            # Flip is_current flags (inside atomic block for safety)
            if source_year.is_current and not target_year.is_current:
                source_year.is_current = False
                source_year.save(update_fields=['is_current'])
                target_year.is_current = True
                target_year.save(update_fields=['is_current'])

        if dry_run:
            with transaction.atomic():
                sid = transaction.savepoint()
                _do_promote()
                transaction.savepoint_rollback(sid)
        else:
            with transaction.atomic():
                _do_promote()

        return Response({
            'dry_run': dry_run,
            'promoted': promoted,
            'skipped': skipped,
            'sections_created': sections_created,
            'teachers_migrated': teachers_migrated,
            'source_year': source_year.name,
            'target_year': target_year.name,
        })

    @action(detail=True, methods=['post'], permission_classes=[permissions.IsAuthenticated])
    def archive(self, request, pk=None):
        """Archive an academic year: marks it as not current (not current = archived)."""
        year = self.get_object()
        user = request.user
        if not is_admin(user) and not is_school_admin(user):
            return Response({'error': 'غير مصرح'}, status=status.HTTP_403_FORBIDDEN)

        year.is_current = False
        year.save(update_fields=['is_current'])
        return Response({'archived': year.name, 'is_current': year.is_current})

    @action(detail=False, methods=['get'], permission_classes=[permissions.IsAuthenticated])
    def stats(self, request):
        """Return promotion statistics for a given year: enrollment counts, section counts, teacher counts."""
        year_id = request.query_params.get('year')
        school_id = request.query_params.get('school')
        if not year_id:
            return Response({'error': 'year param is required'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            year = AcademicYear.objects.get(id=year_id)
        except AcademicYear.DoesNotExist:
            return Response({'error': 'year not found'}, status=status.HTTP_404_NOT_FOUND)

        enrollments = StudentEnrollment.objects.filter(academic_year=year)
        sections = Section.objects.filter(academic_year=year)
        teachers = TeacherAssignment.objects.filter(academic_year=year)
        if school_id:
            enrollments = enrollments.filter(section__school_id=school_id)
            sections = sections.filter(school_id=school_id)
            teachers = teachers.filter(section__school_id=school_id)

        return Response({
            'year': year.name,
            'enrollments_count': enrollments.count(),
            'sections_count': sections.count(),
            'teachers_count': teachers.values('teacher').distinct().count(),
            'is_current': year.is_current,
        })

    @staticmethod
    def _next_grade(grade):
        from apps.academics.models import Grade
        try:
            return Grade.objects.filter(level=grade.level + 1).first()
        except Grade.DoesNotExist:
            return None


class SectionViewSet(viewsets.ModelViewSet):
    queryset = Section.objects.all()
    serializer_class = SectionSerializer
    permission_classes = [IsAdminOrReadOnly]

    def get_queryset(self):
        from django.db.models import F
        if is_admin(self.request.user):
            qs = Section.objects.all()
        else:
            school_ids = user_school_ids(self.request.user)
            if not school_ids:
                return Section.objects.none()
            qs = Section.objects.filter(school_id__in=school_ids)
        school_id = self.request.query_params.get('school')
        if school_id:
            qs = qs.filter(
                school_id=school_id,
                grade__school_offers__school_id=school_id,
            )
        else:
            # If no school param specified for non-admin, ensure grade__school_offers matches school_id
            qs = qs.filter(grade__school_offers__school_id=F('school_id'))
        academic_year_id = self.request.query_params.get('academic_year')
        if academic_year_id:
            qs = qs.filter(academic_year_id=academic_year_id)
        grade_id = self.request.query_params.get('grade')
        if grade_id:
            qs = qs.filter(grade_id=grade_id)
        class_teacher = self.request.query_params.get('class_teacher')
        if class_teacher:
            qs = qs.filter(class_teacher_id=class_teacher)
        return qs.select_related('school', 'grade', 'academic_year').order_by('school_id', 'grade__level', 'name').annotate(
            students_count_annotated=Count('students', distinct=True)
        )

    def perform_create(self, serializer):
        if not is_admin(self.request.user):
            school = serializer.validated_data.get('school')
            if not school or school.manager_id != self.request.user.id:
                raise serializers.ValidationError('يمكنك إنشاء شعب في مدرستك فقط')
        serializer.save()

    def perform_update(self, serializer):
        if not is_admin(self.request.user):
            school = serializer.validated_data.get('school') or self.get_object().school
            if school.manager_id != self.request.user.id:
                raise serializers.ValidationError('يمكنك تعديل شعب مدرستك فقط')
        serializer.save()

    @action(detail=True, methods=['post'], url_path='enroll', permission_classes=[permissions.IsAuthenticated])
    def enroll_student(self, request, pk=None):
        """Add a student to this section (school-manager or class-teacher self-service).
        Creates the student account if missing and optionally a parent link.
        The student is resolved by email OR national ID. If the student already exists
        and is enrolled in ANOTHER school, a 409 conflict is returned (unless
        `confirm=true`) so the requester can review the student's name and school.
        If the student is already enrolled this year in another section (same school),
        they are moved here and the response reports `moved_from`."""
        section = self.get_object()
        if (not is_admin(request.user)
                and section.school.manager_id != request.user.id
                and not (is_teacher(request.user) and section.class_teacher_id == request.user.id)):
            raise serializers.ValidationError('يمكنك إضافة طلاب إلى شعب مدرستك فقط')

        email = (request.data.get('email') or '').strip().lower() or None
        name = (request.data.get('name') or '').strip()
        national_id = (request.data.get('national_id') or '').strip() or None
        if not name or (not email and not national_id):
            return Response({'error': 'name and (email or national_id) are required'}, status=status.HTTP_400_BAD_REQUEST)
        phone = (request.data.get('phone') or '').strip() or None
        parent_email = (request.data.get('parent_email') or '').strip().lower() or None
        confirm = str(request.data.get('confirm', '')).lower() in ('1', 'true', 'yes')

        user, created = self._resolve_enroll_student(email, national_id, name, phone, request)
        if user is None:
            return Response({'error': 'name and (email or national_id) are required'}, status=status.HTTP_400_BAD_REQUEST)

        if not created:
            latest = StudentEnrollment.objects.filter(student=user).select_related('section__school').order_by('-id').first()
            if latest and latest.section.school_id != section.school_id and not confirm:
                return Response({
                    'conflict': 'national_id',
                    'student': {
                        'id': user.id,
                        'email': user.email,
                        'name': user.translations.get('ar', {}).get('name') or user.email,
                    },
                    'school': {'id': latest.section.school_id, 'name': latest.section.school.name},
                }, status=status.HTTP_409_CONFLICT)

        if parent_email:
            parent, created = User.objects.get_or_create(
                email=parent_email,
                defaults={'username': parent_email, 'role': 'parent', 'translations': {'ar': {'name': parent_email}}},
            )
            if created:
                RoleService.assign_role(parent, 'parent')
            elif parent.role not in ('parent', 'admin', 'teacher'):
                parent.role = 'parent'
                parent.save(update_fields=['role'])
                RoleService.assign_role(parent, 'parent')
            FamilyLink.objects.get_or_create(parent=parent, student=user)

        enrollment, was_created = StudentEnrollment.objects.get_or_create(
            student=user,
            academic_year=section.academic_year,
            defaults={'section': section},
        )
        moved_from = None
        school_moved_from = None
        if not was_created and enrollment.section_id != section.id:
            moved_from = enrollment.section.name
            if enrollment.section.school_id != section.school_id:
                school_moved_from = enrollment.section.school.name
            enrollment.section = section
            enrollment.save(update_fields=['section'])

        return Response({
            'status': 'enrolled',
            'created_account': created,
            'moved': bool(moved_from),
            'moved_from': moved_from,
            'school_moved_from': school_moved_from,
            'student': {
                'id': user.id,
                'email': user.email,
                'name': user.translations.get('ar', {}).get('name') or user.email,
            },
            'enrollment': StudentEnrollmentSerializer(enrollment, context={'request': request}).data,
        })

    def _resolve_enroll_student(self, email, national_id, name, phone, request):
        """Find or create a student by national ID or email. Returns (user, created)."""
        user = None
        if national_id:
            user = User.objects.filter(national_id=national_id, role='student').first()
        if user is None and email:
            user = User.objects.filter(email=email, role='student').first()
        if user:
            changed = False
            if national_id and not user.national_id:
                user.national_id = national_id
                changed = True
            if phone and not user.phone:
                user.phone = phone
                changed = True
            if not RoleService.has_role(user, 'student'):
                user.role = 'student'
                RoleService.assign_role(user, 'student')
                changed = True
            translations = dict(user.translations or {})
            ar = dict(translations.get('ar') or {})
            if name and not ar.get('name'):
                ar['name'] = name
                translations['ar'] = ar
                user.translations = translations
                changed = True
            if changed:
                user.save()
            return user, False
        if not email:
            email = f'student-{national_id}@student.local'
        user = User.objects.create_user(
            email=email,
            username=email,
            role='student',
            national_id=national_id,
            phone=phone or '',
            translations={'ar': {'name': name}},
            is_verified=True,
        )
        user.set_unusable_password()
        user.save(update_fields=['password'])
        RoleService.assign_role(user, 'student')
        return user, True


SECTION_LETTERS = ['أ', 'ب', 'ج', 'د', 'هـ', 'و', 'ز', 'ح', 'ط', 'ي', 'ك', 'ل', 'م', 'ن', 'س', 'ع', 'ف', 'ص', 'ق', 'ر', 'ش', 'ت', 'ث', 'خ', 'ذ', 'ض', 'ظ', 'غ']


class SchoolGradeViewSet(viewsets.ModelViewSet):
    """Which grades a school offers, with auto-generated sections per grade."""
    queryset = SchoolGrade.objects.all()
    serializer_class = SchoolGradeSerializer
    permission_classes = [IsAdminOrReadOnly]

    def get_queryset(self):
        if is_admin(self.request.user):
            qs = SchoolGrade.objects.all()
        else:
            school_ids = user_school_ids(self.request.user)
            qs = SchoolGrade.objects.filter(school_id__in=school_ids) if school_ids else SchoolGrade.objects.none()
        school_id = self.request.query_params.get('school')
        if school_id:
            qs = qs.filter(school_id=school_id)
        return qs.select_related('school', 'grade')

    def _check_school_access(self, serializer):
        if is_admin(self.request.user):
            return serializer.validated_data.get('school')
        school = serializer.validated_data.get('school')
        if not school or school.manager_id != self.request.user.id:
            raise serializers.ValidationError({'school': 'يمكنك إدارة صفوف مدرستك فقط'})
        return school

    @transaction.atomic
    def perform_create(self, serializer):
        self._check_school_access(serializer)
        obj = serializer.save()
        self._sync_sections(obj)

    @transaction.atomic
    def perform_update(self, serializer):
        if is_admin(self.request.user):
            school = serializer.validated_data.get('school') or self.get_object().school
        else:
            school = self.get_object().school
            if school.manager_id != self.request.user.id:
                raise serializers.ValidationError({'school': 'يمكنك إدارة صفوف مدرستك فقط'})
        obj = serializer.save()
        self._sync_sections(obj)

    @transaction.atomic
    def perform_destroy(self, instance):
        if not is_admin(self.request.user) and instance.school.manager_id != self.request.user.id:
            raise serializers.ValidationError({'school': 'يمكنك حذف صفوف مدرستك فقط'})
        instance.delete()

    def _sync_sections(self, obj, academic_year=None):
        """Reconcile Section rows so the school/grade has exactly section_count sections."""
        from apps.schools.models import Section as SectionModel

        academic_year = academic_year or AcademicYear.objects.filter(is_current=True).first()
        if academic_year is None:
            academic_year = AcademicYear.objects.first()
        if academic_year is None:
            return

        existing = SectionModel.objects.filter(
            school=obj.school, grade=obj.grade, academic_year=academic_year,
        ).order_by('id')
        existing_ids = list(existing.values_list('id', flat=True))

        target = obj.section_count if obj.is_active else 0

        # Delete surplus sections (only those without enrollments/assignments)
        surplus = SectionModel.objects.filter(id__in=existing_ids[target:])
        for sec in surplus:
            if sec.students.exists() or sec.teachers.exists():
                continue
            sec.delete()

        # Create missing sections with Arabic letters
        created = 0
        for i in range(target):
            if i < len(existing_ids):
                continue
            if created >= len(SECTION_LETTERS):
                break
            SectionModel.objects.get_or_create(
                school=obj.school, grade=obj.grade, academic_year=academic_year,
                name=SECTION_LETTERS[i],
            )
            created += 1


class SchoolSubjectPeriodViewSet(viewsets.ModelViewSet):
    """Weekly periods per subject within an offered grade of a school."""
    queryset = SchoolSubjectPeriod.objects.all()
    serializer_class = SchoolSubjectPeriodSerializer
    permission_classes = [IsAdminOrReadOnly]

    def get_queryset(self):
        if is_admin(self.request.user):
            qs = SchoolSubjectPeriod.objects.all()
        else:
            school_ids = user_school_ids(self.request.user)
            qs = SchoolSubjectPeriod.objects.filter(school_id__in=school_ids) if school_ids else SchoolSubjectPeriod.objects.none()
        school_id = self.request.query_params.get('school')
        if school_id:
            qs = qs.filter(
                school_id=school_id,
                # Only grades the school actually offers.
                grade__school_offers__school_id=school_id,
            )
        return qs.select_related('school', 'grade', 'subject')

    def _check_school_access(self):
        if is_admin(self.request.user):
            return
        school_id = self.request.data.get('school')
        if school_id and str(school_id) not in {str(s) for s in user_school_ids(self.request.user)}:
            raise serializers.ValidationError({'school': 'يمكنك إدارة صفوف مدرستك فقط'})

    def perform_create(self, serializer):
        self._check_school_access()
        serializer.save()

    def perform_update(self, serializer):
        if not is_admin(self.request.user) and self.get_object().school.manager_id != self.request.user.id:
            raise serializers.ValidationError({'school': 'يمكنك إدارة صفوف مدرستك فقط'})
        serializer.save()

    def perform_destroy(self, instance):
        if not is_admin(self.request.user) and instance.school.manager_id != self.request.user.id:
            raise serializers.ValidationError({'school': 'يمكنك حذف صفوف مدرستك فقط'})
        instance.delete()


class SchoolTeacherViewSet(viewsets.ModelViewSet):
    """Teacher accounts linked to a school, so a school manager can manage their teachers."""
    queryset = SchoolTeacher.objects.all()
    serializer_class = SchoolTeacherSerializer
    permission_classes = [IsAdminOrReadOnly]

    def get_queryset(self):
        if is_admin(self.request.user):
            qs = SchoolTeacher.objects.all()
        else:
            school_ids = user_school_ids(self.request.user)
            qs = SchoolTeacher.objects.filter(school_id__in=school_ids) if school_ids else SchoolTeacher.objects.none()
        school_id = self.request.query_params.get('school')
        if school_id:
            qs = qs.filter(school_id=school_id)
        return qs.select_related('school', 'teacher')

    def _check_school_access(self, serializer):
        if is_admin(self.request.user):
            return serializer.validated_data.get('school')
        school = serializer.validated_data.get('school')
        if not school or school.manager_id != self.request.user.id:
            raise serializers.ValidationError({'school': 'يمكنك إدارة معلمي مدرستك فقط'})
        return school

    def create(self, request, *args, **kwargs):
        """Create (or reuse) a teacher account and link it to the school."""
        from django.contrib.auth import get_user_model

        input_serializer = SchoolTeacherCreateSerializer(data=request.data)
        input_serializer.is_valid(raise_exception=True)
        data = input_serializer.validated_data

        school = self._check_school_access(input_serializer)
        email = (data.get('teacher_email') or '').strip().lower()
        name = (data.get('teacher_name') or '').strip()
        password = data.get('password') or get_user_model().objects.make_random_password()

        teacher = User.objects.filter(email=email).first()
        if teacher is None:
            teacher = get_user_model()(
                email=email,
                username=email,
                role=get_user_model().Role.TEACHER,
                is_verified=True,
            )
            teacher.set_password(password)
            teacher.translations = {'ar': {'name': name}} if name else {}
            teacher.save()
        elif teacher.role != get_user_model().Role.TEACHER:
            teacher.role = get_user_model().Role.TEACHER
            if name:
                teacher.translations['ar'] = {'name': name}
            teacher.save()

        link, _ = SchoolTeacher.objects.get_or_create(school=school, teacher=teacher)

        out = SchoolTeacherSerializer(link, context=self.get_serializer_context())
        return Response(out.data, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        if not is_admin(self.request.user) and instance.school.manager_id != self.request.user.id:
            raise serializers.ValidationError({'school': 'يمكنك تعديل معلمي مدرستك فقط'})

        name = (request.data.get('teacher_name') or '').strip()
        if name:
            translations = dict(instance.teacher.translations or {})
            translations['ar'] = {'name': name}
            instance.teacher.translations = translations
            instance.teacher.save(update_fields=['translations'])

        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        return Response(serializer.data)

    def perform_update(self, serializer):
        if not is_admin(self.request.user):
            school = self.get_object().school
            if school.manager_id != self.request.user.id:
                raise serializers.ValidationError({'school': 'يمكنك تعديل معلمي مدرستك فقط'})
        serializer.save()

    def perform_destroy(self, instance):
        if not is_admin(self.request.user) and instance.school.manager_id != self.request.user.id:
            raise serializers.ValidationError({'school': 'يمكنك حذف معلمي مدرستك فقط'})
        instance.delete()


class SchoolStaffViewSet(viewsets.ModelViewSet):
    """Non-teaching staff (accountant, transport officer, librarian) linked to a school."""
    serializer_class = SchoolStaffSerializer
    permission_classes = [IsAdminOrReadOnly]

    def get_queryset(self):
        qs = SchoolStaff.objects.select_related('school', 'user').all()
        school_id = self.request.query_params.get('school')
        if school_id:
            qs = qs.filter(school_id=school_id)
        if is_admin(self.request.user):
            return qs
        if is_school_admin(self.request.user):
            managed_ids = self.request.user.managed_schools.values_list('id', flat=True)
            return qs.filter(school_id__in=managed_ids)
        return SchoolStaff.objects.none()

    def perform_create(self, serializer):
        school = serializer.validated_data['school']
        if not is_admin(self.request.user) and school.manager_id != self.request.user.id:
            raise serializers.ValidationError({'school': 'يمكنك إدارة طاقم مدرستك فقط'})
        serializer.save()

    def perform_update(self, serializer):
        school = self.get_object().school
        if not is_admin(self.request.user) and school.manager_id != self.request.user.id:
            raise serializers.ValidationError({'school': 'يمكنك إدارة طاقم مدرستك فقط'})
        serializer.save()

    def perform_destroy(self, instance):
        if not is_admin(self.request.user) and instance.school.manager_id != self.request.user.id:
            raise serializers.ValidationError({'school': 'يمكنك حذف طاقم مدرستك فقط'})
        instance.delete()


class TeacherAssignmentViewSet(viewsets.ModelViewSet):
    queryset = TeacherAssignment.objects.all()
    serializer_class = TeacherAssignmentSerializer
    permission_classes = [IsAdminOrReadOnly]

    def get_queryset(self):
        if is_admin(self.request.user):
            qs = TeacherAssignment.objects.all()
        else:
            section_ids = user_section_ids(self.request.user)
            if section_ids is None:
                section_ids = set()
            if is_teacher(self.request.user):
                qs = TeacherAssignment.objects.filter(teacher=self.request.user)
            elif RoleService.has_role(self.request.user, 'student'):
                qs = TeacherAssignment.objects.filter(section_id__in=section_ids)
            else:
                qs = TeacherAssignment.objects.none()
        school_id = self.request.query_params.get('school')
        if school_id:
            qs = qs.filter(section__school_id=school_id)
        return qs

    def perform_create(self, serializer):
        if is_admin(self.request.user):
            serializer.save()
            return
        section = serializer.validated_data.get('section')
        if not section or section.school.manager_id != self.request.user.id:
            raise serializers.ValidationError({'section': 'يمكنك إسناد معلمين في شعب مدرستك فقط'})
        serializer.save()

    def perform_update(self, serializer):
        if is_admin(self.request.user):
            serializer.save()
            return
        assignment = self.get_object()
        if assignment.section.school.manager_id != self.request.user.id:
            raise serializers.ValidationError({'section': 'يمكنك تعديل إسناد مدرستك فقط'})
        serializer.save()

    def perform_destroy(self, instance):
        if is_admin(self.request.user):
            instance.delete()
            return
        if instance.section.school.manager_id != self.request.user.id:
            raise serializers.ValidationError({'section': 'يمكنك حذف إسناد مدرستك فقط'})
        instance.delete()


class StudentEnrollmentViewSet(viewsets.ModelViewSet):
    queryset = StudentEnrollment.objects.all()
    serializer_class = StudentEnrollmentSerializer
    permission_classes = [CanManageEnrollments]

    def get_queryset(self):
        if is_admin(self.request.user):
            qs = StudentEnrollment.objects.all()
        elif RoleService.has_role(self.request.user, 'school_admin'):
            school_ids = user_school_ids(self.request.user)
            qs = StudentEnrollment.objects.filter(section__school_id__in=school_ids) if school_ids else StudentEnrollment.objects.none()
        else:
            section_ids = user_section_ids(self.request.user)
            if section_ids is None:
                section_ids = set()
            if is_teacher(self.request.user):
                qs = StudentEnrollment.objects.filter(section_id__in=section_ids)
            elif RoleService.has_role(self.request.user, 'student'):
                qs = StudentEnrollment.objects.filter(student=self.request.user)
            else:
                qs = StudentEnrollment.objects.none()
        section = self.request.query_params.get('section')
        if section:
            qs = qs.filter(section_id=section)
        school_id = self.request.query_params.get('school')
        if school_id:
            qs = qs.filter(section__school_id=school_id)
        return qs

    @action(detail=True, methods=['post'], permission_classes=[CanManageEnrollments])
    def transfer(self, request, pk=None):
        """Transfer a student to another section (same school/year/grade for school
        managers; unrestricted for system admins), keeping archive."""
        enrollment = self.get_object()
        target_section_id = request.data.get('target_section_id')
        if not target_section_id:
            return Response({'error': 'target_section_id is required'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            target_section = Section.objects.get(id=target_section_id)
        except Section.DoesNotExist:
            return Response({'error': 'target section not found'}, status=status.HTTP_404_NOT_FOUND)

        if (not is_admin(request.user)
                and (target_section.school_id != enrollment.section.school_id
                     or target_section.academic_year_id != enrollment.section.academic_year_id
                     or target_section.grade_id != enrollment.section.grade_id)):
            return Response(
                    {'error': 'يمكن نقل الطالب ضمن نفس المدرسة والسنة الدراسية والصف فقط'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        old_section = enrollment.section
        enrollment.section = target_section
        enrollment.save(update_fields=['section'])
        return Response({
            'status': 'transferred',
            'student': enrollment.student.email,
            'from_section': old_section.name,
            'to_section': target_section.name,
        })

    @action(detail=False, methods=['post'], permission_classes=[IsAdminRole])
    def transfer_by_code(self, request):
        """Transfer by national ID / transfer code: find the student and move to the target section."""
        national_id = request.data.get('national_id')
        transfer_code = request.data.get('transfer_code')
        target_section_id = request.data.get('target_section_id')

        student = None
        if national_id:
            student = User.objects.filter(
                national_id=national_id, role='student',
            ).first() if hasattr(User, 'national_id') else None
        if not student and transfer_code:
            student = User.objects.filter(
                email=transfer_code, role='student',
            ).first()
        if not student:
            return Response({'error': 'Student not found by the given code'}, status=status.HTTP_404_NOT_FOUND)

        try:
            target_section = Section.objects.get(id=target_section_id)
        except (Section.DoesNotExist, TypeError):
            return Response({'error': 'target_section_id is required and must be valid'}, status=status.HTTP_400_BAD_REQUEST)

        enrollment, created = StudentEnrollment.objects.get_or_create(
            student=student,
            academic_year=target_section.academic_year,
            defaults={'section': target_section},
        )
        if not created:
            enrollment.section = target_section
            enrollment.save(update_fields=['section'])
        return Response({
            'status': 'transferred',
            'student': student.email,
            'to_section': str(target_section),
        })


class SchoolAnnouncementViewSet(viewsets.ModelViewSet):
    queryset = SchoolAnnouncement.objects.all()
    serializer_class = SchoolAnnouncementSerializer
    permission_classes = [CanManageAnnouncements]

    def get_queryset(self):
        if is_admin(self.request.user):
            qs = SchoolAnnouncement.objects.all()
        else:
            section_ids = user_section_ids(self.request.user)
            school_ids = user_school_ids(self.request.user)
            if not section_ids and not school_ids:
                return SchoolAnnouncement.objects.none()
            qs = SchoolAnnouncement.objects.filter(
                Q(section_id__in=section_ids) | Q(school_id__in=school_ids, section__isnull=True)
            )
        school_id = self.request.query_params.get('school')
        if school_id:
            qs = qs.filter(school_id=school_id)
        return qs

    def perform_create(self, serializer):
        if not is_admin(self.request.user) and RoleService.has_role(self.request.user, 'school_admin'):
            school = serializer.validated_data.get('school')
            if not school or school.manager_id != self.request.user.id:
                raise serializers.ValidationError({'school': 'يمكنك النشر في مدرستك فقط'})
        is_emergency = serializer.validated_data.get('is_emergency', False)
        if is_emergency and not is_admin(self.request.user) and not RoleService.has_role(self.request.user, 'school_admin'):
            serializer.validated_data['is_emergency'] = False
        announcement = serializer.save(author=self.request.user)
        from apps.notifications.services import notify_many
        students = StudentEnrollment.objects.filter(
            section=announcement.section
        ) if announcement.section else StudentEnrollment.objects.filter(
            section__school=announcement.school
        )
        student_ids = list(students.values_list('student_id', flat=True))
        parent_ids = list(
            FamilyLink.objects.filter(student_id__in=student_ids).values_list('parent_id', flat=True)
        )
        recipients = set(student_ids) | set(parent_ids)
        recipients.discard(announcement.author_id)
        if recipients:
            notify_many(
                recipients,
                type='announcement',
                title={'ar': 'إعلان جديد', 'en': 'New announcement'},
                body={
                    'ar': f"{announcement.school.name} — {announcement.title}",
                    'en': f"{announcement.school.name} — {announcement.title}",
                },
                link='/school',
                icon='🏫',
            )
        if announcement.is_emergency:
            enrollments = StudentEnrollment.objects.filter(section=announcement.section) if announcement.section else StudentEnrollment.objects.filter(section__school=announcement.school)
            for en in enrollments:
                if en.student.phone:
                    send_whatsapp_alert(en.student.phone, f"تنبيه طارئ من مدرسة {announcement.school.name}:\n{announcement.title}\n{announcement.content}")

    @action(detail=True, methods=['post'], permission_classes=[permissions.IsAuthenticated])
    def acknowledge(self, request, pk=None):
        """تأكيد القراءة: marks the announcement as read by the current user (parent/student)."""
        announcement = self.get_object()
        receipt, created = AnnouncementReadReceipt.objects.get_or_create(
            announcement=announcement,
            user=request.user,
        )
        return Response({
            'status': 'acknowledged',
            'read_count': announcement.read_receipts.count(),
        })


class AttendanceViewSet(viewsets.ModelViewSet):
    """سجل الحضور/الغياب اليومي، يسجّله المعلم أو الإدارة وتصل تنبيهات الغياب لولي الأمر."""

    queryset = Attendance.objects.select_related('student', 'section', 'section__school', 'school', 'recorded_by')
    serializer_class = AttendanceSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if is_admin(user):
            qs = self.queryset
        elif RoleService.has_role(user, 'school_admin'):
            school_ids = user_school_ids(user)
            qs = self.queryset.filter(school_id__in=school_ids) if school_ids else self.queryset.none()
        elif RoleService.has_role(user, 'teacher'):
            section_ids = user_section_ids(user)
            qs = self.queryset.filter(section_id__in=section_ids) if section_ids else self.queryset.none()
        elif RoleService.has_role(user, 'parent'):
            child_ids = FamilyLink.objects.filter(parent=user).values_list('student_id', flat=True)
            qs = self.queryset.filter(student_id__in=child_ids)
        elif RoleService.has_role(user, 'student'):
            qs = self.queryset.filter(student=user)
        else:
            qs = self.queryset.none()

        date_filter = self.request.query_params.get('date')
        if date_filter:
            qs = qs.filter(date=date_filter)
        section = self.request.query_params.get('section')
        if section:
            qs = qs.filter(section_id=section)
        student = self.request.query_params.get('student')
        if student:
            qs = qs.filter(student_id=student)
        school_id = self.request.query_params.get('school')
        if school_id:
            qs = qs.filter(school_id=school_id)
        return qs.distinct()

    def _can_manage_section(self, user, section):
        """Teachers (assigned to the section or its class mentor) and school admins
        (of that school) may record."""
        if is_admin(user):
            return True
        if RoleService.has_role(user, 'school_admin'):
            return section.school.manager_id == user.id
        if RoleService.has_role(user, 'teacher'):
            return (TeacherAssignment.objects.filter(teacher=user, section=section).exists()
                    or section.class_teacher_id == user.id)
        return False

    def perform_create(self, serializer):
        section = serializer.validated_data.get('section')
        if section and not self._can_manage_section(self.request.user, section):
            raise serializers.ValidationError({'section': 'أنت غير مخوَّل بالتسجيل في هذه الشعبة'})
        attendance = serializer.save(
            recorded_by=self.request.user,
            school=section.school if section else serializer.validated_data.get('school'),
        )
        notify_absence(attendance)

    def create(self, request, *args, **kwargs):
        """Record attendance, appending a non-blocking warning if the date is not a working day."""
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        section = serializer.validated_data.get('section')
        warning = None
        if section:
            att_date = serializer.validated_data.get('date') or timezone.localdate()
            warning = working_day_warning(section.school, att_date)
        self.perform_create(serializer)
        data = serializer.data
        if warning:
            data = {**data, 'warning': warning}
        return Response(data, status=status.HTTP_201_CREATED, headers=self.get_success_headers(serializer.data))

    @action(detail=False, methods=['post'])
    def bulk_record(self, request):
        """تسجيل جماعي للحضور/الغياب لشعبة في يوم واحد.

        Body: {section: id, date: "YYYY-MM-DD", records: [{student: id, status: "present|absent"}]}
        """
        section_id = request.data.get('section')
        try:
            section = Section.objects.get(id=section_id)
        except (Section.DoesNotExist, TypeError, ValueError):
            return Response({'error': 'section is required and must be valid'}, status=status.HTTP_400_BAD_REQUEST)

        if not self._can_manage_section(request.user, section):
            return Response({'error': 'أنت غير مخوَّل بالتسجيل في هذه الشعبة'}, status=status.HTTP_403_FORBIDDEN)

        target_date_raw = request.data.get('date')
        if target_date_raw:
            try:
                target_date = date.fromisoformat(str(target_date_raw))
            except ValueError:
                return Response({'error': 'date must be YYYY-MM-DD'}, status=status.HTTP_400_BAD_REQUEST)
        else:
            target_date = timezone.localdate()
        records = request.data.get('records') or []
        if not isinstance(records, list) or not records:
            return Response({'error': 'records must be a non-empty list'}, status=status.HTTP_400_BAD_REQUEST)

        created = []
        updated = []
        absent_alerts = 0
        warning = working_day_warning(section.school, target_date)
        with transaction.atomic():
            for record in records:
                student_id = record.get('student')
                status_value = record.get('status')
                if status_value not in Attendance.Status.values:
                    return Response(
                        {'error': f"status must be one of {list(Attendance.Status.values)}"},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                attendance, was_created = Attendance.objects.update_or_create(
                    student_id=student_id,
                    date=target_date,
                    defaults={
                        'section': section,
                        'school': section.school,
                        'status': status_value,
                        'recorded_by': request.user,
                    },
                )
                if was_created:
                    created.append(student_id)
                else:
                    updated.append(student_id)
                if status_value == Attendance.Status.ABSENT:
                    absent_alerts += 1
                    notify_absence(attendance)

        return Response({
            'status': 'ok',
            'date': target_date.isoformat(),
            'section': section.id,
            'created': len(created),
            'updated': len(updated),
            'absent_alerts_sent': absent_alerts,
            'warning': warning,
        })


class ParentTeacherTicketViewSet(viewsets.ModelViewSet):
    queryset = ParentTeacherTicket.objects.all()
    serializer_class = ParentTeacherTicketSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        if is_admin(self.request.user):
            qs = ParentTeacherTicket.objects.all()
        else:
            qs = ParentTeacherTicket.objects.filter(
                parent=self.request.user,
            ) | ParentTeacherTicket.objects.filter(
                student=self.request.user,
            ) | ParentTeacherTicket.objects.filter(
                teacher=self.request.user,
            )
        school_id = self.request.query_params.get('school')
        if school_id:
            student_ids = StudentEnrollment.objects.filter(
                section__school_id=school_id
            ).values_list('student_id', flat=True)
            qs = qs.filter(student_id__in=student_ids)
        return qs

    def perform_create(self, serializer):
        serializer.save(parent=self.request.user)

    @action(detail=True, methods=['post'])
    def add_message(self, request, pk=None):
        ticket = self.get_object()
        if request.user not in (ticket.parent, ticket.teacher, ticket.student) and not is_admin(request.user):
            return Response({'error': 'Not allowed to reply on this ticket'}, status=status.HTTP_403_FORBIDDEN)
        text = request.data.get('message')
        if not text:
            return Response({'error': 'Message text is required'}, status=status.HTTP_400_BAD_REQUEST)

        messages_list = ticket.messages or []
        messages_list.append({
            'sender': request.user.email,
            'role': request.user.role,
            'text': text,
            'timestamp': timezone.now().isoformat(),
        })
        ticket.messages = messages_list
        ticket.save()
        from apps.notifications.services import notify
        for participant in {ticket.parent, ticket.teacher, ticket.student}:
            if participant and participant.id != request.user.id:
                notify(
                    participant,
                    type='ticket',
                    title={'ar': 'رد جديد على تذكرتك', 'en': 'New reply on your ticket'},
                    body={
                        'ar': f"{request.user.email}: {text[:80]}",
                        'en': f"{request.user.email}: {text[:80]}",
                    },
                    link='/school',
                    icon='💬',
                )
        return Response(ParentTeacherTicketSerializer(ticket).data)


class FamilyLinkViewSet(viewsets.ModelViewSet):
    queryset = FamilyLink.objects.all()
    serializer_class = FamilyLinkSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        if is_admin(self.request.user):
            return FamilyLink.objects.all()
        return FamilyLink.objects.filter(parent=self.request.user)

    def perform_create(self, serializer):
        student = serializer.validated_data.get('student')
        if is_admin(self.request.user):
            return self._save(serializer)
        if not student or not RoleService.has_role(student, 'student'):
            raise serializers.ValidationError({'student': 'Invalid student account'})
        return self._save(serializer)

    def _save(self, serializer):
        try:
            serializer.save(parent=self.request.user)
        except Exception as exc:
            from django.db import IntegrityError
            if isinstance(exc, IntegrityError):
                raise serializers.ValidationError({'student': 'This student is already linked to you'}) from None
            raise


class FAQViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = FAQ.objects.filter(is_active=True)
    serializer_class = FAQSerializer
    permission_classes = [permissions.AllowAny]


class AttachmentViewSet(viewsets.ModelViewSet):
    """رفع صور/ملفات لشرح الدروس أو الواجبات المنزلية، مع متابعة إدارية للحالة."""

    queryset = Attachment.objects.all()
    serializer_class = AttachmentSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if is_admin(user):
            qs = Attachment.objects.all()
        elif RoleService.has_role(user, 'teacher'):
            section_ids = user_section_ids(user)
            qs = Attachment.objects.filter(Q(uploader=user) | Q(section_id__in=section_ids))
        elif RoleService.has_role(user, 'parent'):
            child_ids = FamilyLink.objects.filter(parent=user).values_list('student_id', flat=True)
            qs = Attachment.objects.filter(
                Q(uploader_id__in=child_ids) |
                Q(section_id__in=user_section_ids(user))
            )
        else:
            qs = Attachment.objects.filter(
                Q(uploader=user) |
                Q(section_id__in=user_section_ids(user))
            )
        kind = self.request.query_params.get('kind')
        if kind in Attachment.Kind.values:
            qs = qs.filter(kind=kind)
        status_filter = self.request.query_params.get('review_status')
        if status_filter in Attachment.ReviewStatus.values:
            qs = qs.filter(review_status=status_filter)
        school_id = self.request.query_params.get('school')
        if school_id:
            qs = qs.filter(Q(school_id=school_id) | Q(section__school_id=school_id))
        return qs.distinct()

    def perform_create(self, serializer):
        file = self.request.FILES.get('file')
        if not file:
            raise serializers.ValidationError({'file': 'file is required'})
        dangerous_exts = ['.html', '.htm', '.py', '.js', '.php', '.sh', '.exe', '.bat', '.cmd', '.svg']
        filename_lower = file.name.lower()
        if any(filename_lower.endswith(ext) for ext in dangerous_exts) or getattr(file, 'content_type', '') == 'text/html':
            raise serializers.ValidationError({'file': 'This file type is not allowed for security reasons.'})
        section_id = serializer.validated_data.get('section')
        if section_id:
            allowed = user_section_ids(self.request.user)
            if not is_admin(self.request.user) and section_id.id not in allowed:
                raise serializers.ValidationError({'section': 'You are not allowed to upload to this section'})
        serializer.save(
            uploader=self.request.user,
            file_name=file.name,
            mime_type=getattr(file, 'content_type', ''),
            file_size=file.size,
        )

    @action(detail=True, methods=['get'], permission_classes=[permissions.IsAuthenticated])
    def download(self, request, pk=None):
        attachment = self.get_object()
        if not attachment.file:
            return Response({'error': 'File not found'}, status=status.HTTP_404_NOT_FOUND)
        from django.http import FileResponse
        response = FileResponse(attachment.file.open('rb'), content_type=attachment.mime_type or 'application/octet-stream')
        response['Content-Disposition'] = f'attachment; filename="{attachment.file_name or "attachment"}"'
        return response

    @action(detail=True, methods=['post'], permission_classes=[IsAdminRole])
    def review(self, request, pk=None):
        """متابعة إدارية: اعتماد أو رفض المرفق مع ملاحظات."""
        attachment = self.get_object()
        new_status = request.data.get('review_status')
        if new_status not in Attachment.ReviewStatus.values:
            return Response({'error': 'review_status must be approved or rejected'}, status=status.HTTP_400_BAD_REQUEST)
        attachment.review_status = new_status
        attachment.review_notes = request.data.get('review_notes', '')
        attachment.reviewed_by = request.user
        attachment.reviewed_at = timezone.now()
        attachment.save(update_fields=['review_status', 'review_notes', 'reviewed_by', 'reviewed_at'])
        return Response(AttachmentSerializer(attachment, context=self.get_serializer_context()).data)


class SupportRequestCreateView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        subject = request.data.get('subject')
        message = request.data.get('message')
        if not subject or not message:
            return Response({'error': 'subject and message are required'}, status=status.HTTP_400_BAD_REQUEST)
        support = SupportRequest.objects.create(
            user=request.user,
            subject=subject,
            message=message,
        )
        return Response(SupportRequestSerializer(support).data, status=status.HTTP_201_CREATED)


class WeeklySummaryAPIView(APIView):
    """Automated weekly summary (التقارير الأسبوعية) for the parent/student's children or sections."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        user = request.user
        today = date.today()

        if RoleService.has_role(user, 'parent'):
            children = User.objects.filter(
                Q(linked_guardians__parent=user),
            ).distinct()
            students = children
        elif RoleService.has_role(user, 'student'):
            students = User.objects.filter(pk=user.pk)
        elif is_teacher(user) or is_admin(user):
            section_ids = user_section_ids(user)
            students = User.objects.filter(
                school_enrollments__section_id__in=section_ids,
            ).distinct() if section_ids else User.objects.none()
        else:
            students = User.objects.none()

        reports = []
        for student in students:
            student_enrollments = StudentEnrollment.objects.filter(student=student).select_related('section')
            if not student_enrollments.exists():
                continue
            section_ids = student_enrollments.values_list('section_id', flat=True)
            section = student_enrollments.first().section
            school = section.school if section else None
            week_start = week_start_date_for(school, today)
            week_end = week_start + timedelta(days=7)
            week_announcements = SchoolAnnouncement.objects.filter(
                section_id__in=section_ids,
                created_at__date__gte=week_start,
                created_at__date__lt=week_end,
            )
            open_tickets = ParentTeacherTicket.objects.filter(student=student, status__in=['open', 'in_progress'])
            week_attendance = Attendance.objects.filter(
                student=student,
                date__gte=week_start,
                date__lt=week_end,
            )
            if week_attendance.exists():
                present = week_attendance.filter(status='present').count()
                attendance_rate = present / week_attendance.count() * 100.0
            else:
                attendance_rate = 0.0

            topics_needing_support = []
            for ticket in ParentTeacherTicket.objects.filter(student=student).exclude(subject=None)[:5]:
                subject = ticket.subject
                if subject:
                    name = subject.translations.get('ar', {}).get('name', '')
                    if name and name not in topics_needing_support:
                        topics_needing_support.append(name)

            summary_text = (
                f"خلال أسبوع {week_start:%d/%m/%Y} تلقى الطالب {len(week_announcements)} تنبيهاً/واجباً، "
                f"ومازالت {open_tickets.count()} من الاستفسارات مفتوحة. نسبة الحضور المقدّرة {attendance_rate:.0f}%."
            )

            report, _ = WeeklyReport.objects.update_or_create(
                student=student,
                week_start=week_start,
                defaults={
                    'parent': user if RoleService.has_role(user, 'parent') else None,
                    'summary': summary_text,
                    'assignments_submitted': len(week_announcements),
                    'attendance_rate': attendance_rate,
                    'topics_needing_support': topics_needing_support,
                },
            )
            reports.append(WeeklyReportSerializer(report).data)

        return Response({
            'week_start': week_start.isoformat(),
            'reports': reports,
        })


class BulkImportView(APIView):
    """Bulk import (الاستيراد الجماعي) of schools/students/teachers from CSV/Excel.

    School managers and class teachers may only import students, scoped to their own
    school (or their class-teacher section); the file's school_code column is ignored
    for them and their school/section is forced instead.
    """

    permission_classes = [IsAdminOrSchoolManager]

    STUDENT_ALIASES = {
        'email': ('email', 'البريد', 'البريد الإلكتروني', 'الايميل'),
        'name': ('name', 'اسم الطالب', 'الاسم', 'student name', 'full name'),
        'national_id': ('national_id', 'الرقم الوطني', 'رقم الهوية'),
        'phone': ('phone', 'الهاتف', 'رقم الهاتف', 'جوال', 'phone number'),
        'parent_email': ('parent_email', 'بريد ولي الأمر', 'البريد الإلكتروني لولي الأمر', 'parent email'),
        'grade_level': ('grade_level', 'الصف', 'المستوى', 'الصف الدراسي', 'المرحلة', 'grade'),
        'section_name': ('section_name', 'الشعبة', 'اسم الشعبة', 'section'),
        'academic_year': ('academic_year', 'السنة الدراسية', 'العام الدراسي', 'academic year'),
        'school_code': ('school_code', 'الرمز المدرسي', 'رمز المدرسة', 'school code'),
    }

    def post(self, request):
        kind = request.data.get('kind', 'students')
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'file is required'}, status=status.HTTP_400_BAD_REQUEST)

        rows = self._read_rows(file)
        if rows is None:
            return Response({'error': 'file must be CSV or XLSX'}, status=status.HTTP_400_BAD_REQUEST)

        scope = self._resolve_scope(request)
        if not is_admin(request.user) and kind != 'students':
            return Response({'error': 'school managers and class teachers may only import students'}, status=status.HTTP_403_FORBIDDEN)

        results = {'created': 0, 'updated': 0, 'skipped': 0, 'errors': []}

        if kind == 'schools':
            self._import_schools(rows, results)
        elif kind == 'students':
            self._import_students(rows, results, school_id=scope.get('school_id'), section=scope.get('section'))
        elif kind == 'teachers':
            self._import_teachers(rows, results)
        else:
            return Response({'error': 'kind must be schools, students or teachers'}, status=status.HTTP_400_BAD_REQUEST)

        return Response(results)

    def _resolve_scope(self, request):
        """Determine the forced school/section for the current user (non-admins)."""
        user = request.user
        if is_admin(user):
            return {}
        if is_school_admin(user):
            school_ids = user_school_ids(user)
            raw = request.data.get('school_id')
            try:
                school_id = int(raw)
            except (TypeError, ValueError):
                school_id = None
            if school_id is None or not school_ids or school_id not in school_ids:
                raise serializers.ValidationError({'school_id': 'school_id is required and must be one of your schools'})
            return {'school_id': school_id}
        if is_teacher(user):
            raw = request.data.get('section_id')
            try:
                section = Section.objects.get(id=int(raw), class_teacher=user)
            except (Section.DoesNotExist, TypeError, ValueError) as exc:
                raise serializers.ValidationError({'section_id': 'section_id is required and must be a section you mentor'}) from exc
            return {'school_id': section.school_id, 'section': section}
        raise serializers.ValidationError('Not allowed to import')

    @staticmethod
    def _read_rows(file):
        """Convert an uploaded CSV or XLSX file into a list of dict rows (first row = headers)."""
        name = (file.name or '').lower()
        if name.endswith('.xlsx') or name.endswith('.xls'):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
                ws = wb[wb.sheetnames[0]]
                iter_rows = ws.iter_rows(values_only=True)
                try:
                    header = [str(c).strip() if c is not None else '' for c in next(iter_rows)]
                except StopIteration:
                    return []
                rows = []
                for values in iter_rows:
                    row = {header[i]: (values[i] if i < len(values) else None) for i in range(len(header))}
                    rows.append(row)
                return rows
            except ImportError:
                raise serializers.ValidationError({'file': 'XLSX support requires openpyxl'}) from None
            except Exception as exc:
                raise serializers.ValidationError({'file': f'could not parse XLSX: {exc}'}) from None

        try:
            decoded = file.read().decode('utf-8-sig')
        except UnicodeDecodeError:
            decoded = file.read().decode('utf-8')
        return list(csv.DictReader(io.StringIO(decoded)))

    def _import_schools(self, rows, results):
        """Import schools from the Jordanian open-data portal (opendata.gov.jo) dataset."""
        mapping = {
            'رمز المؤسسة': 'school_code',
            'اسم المؤسسة': 'name',
            'المديرية': 'directorate',
            'المحافظة': 'governorate',
            'الإقليم': 'region',
            'جنس المؤسس': 'gender',
            'نوع التعليم': 'education_type',
            'العنوان': 'address',
        }
        for raw in rows:
            school_code = self._cell(raw, 'رمز المؤسسة') or self._cell(raw, 'school_code')
            name = self._cell(raw, 'اسم المؤسسة') or self._cell(raw, 'name')
            if not school_code or not name:
                results['errors'].append({'row': raw, 'error': 'school_code and name are required'})
                continue
            data = {'name': name, 'school_code': school_code}
            for src, field in mapping.items():
                if field == 'school_code':
                    continue
                value = self._cell(raw, src) or self._cell(raw, field)
                if value:
                    data[field] = value
            data['translations'] = {'ar': {'name': name}}
            _, created = School.objects.update_or_create(
                school_code=school_code,
                defaults=data,
            )
            if created:
                results['created'] += 1
            else:
                results['updated'] += 1

    @staticmethod
    def _cell(row, header):
        value = row.get(header) if isinstance(row, dict) else None
        if value is None:
            return ''
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    @classmethod
    def _row_value(cls, row, aliases):
        """First non-empty cell across the given column aliases (Arabic/English)."""
        for alias in aliases:
            value = cls._cell(row, alias)
            if value:
                return value
        return ''

    def _import_students(self, rows, results, school_id=None, section=None):
        for row in rows:
            email = self._row_value(row, self.STUDENT_ALIASES['email']).lower() or None
            name = self._row_value(row, self.STUDENT_ALIASES['name'])
            national_id = self._row_value(row, self.STUDENT_ALIASES['national_id']) or None
            phone = self._row_value(row, self.STUDENT_ALIASES['phone'])
            parent_email = self._row_value(row, self.STUDENT_ALIASES['parent_email']).lower() or None

            if not email and not national_id:
                results['errors'].append({'row': row, 'error': 'email or national_id is required'})
                continue
            if not name:
                results['errors'].append({'row': row, 'error': 'name is required'})
                continue

            user, created = self._get_or_create_student(email, national_id, name, phone)
            if created:
                results['created'] += 1
            else:
                results['updated'] += 1

            if parent_email:
                parent, _ = User.objects.get_or_create(
                    email=parent_email,
                    defaults={'username': parent_email, 'role': 'parent', 'translations': {'ar': {'name': parent_email}}},
                )
                if parent.role not in ('parent', 'admin', 'teacher'):
                    parent.role = 'parent'
                    parent.save(update_fields=['role'])
                FamilyLink.objects.get_or_create(parent=parent, student=user)

            self._enroll_by_row(user, row, results, school_id=school_id, section=section)

    def _get_or_create_student(self, email, national_id, name, phone):
        """Resolve a student by national ID first, then email; create if missing."""
        user = None
        if national_id:
            user = User.objects.filter(national_id=national_id, role='student').first()
        if user is None and email:
            user = User.objects.filter(email=email, role='student').first()
        if user:
            changed = False
            if national_id and not user.national_id:
                user.national_id = national_id
                changed = True
            if phone and not user.phone:
                user.phone = phone
                changed = True
            if not RoleService.has_role(user, 'student'):
                user.role = 'student'
                changed = True
            translations = dict(user.translations or {})
            ar = dict(translations.get('ar') or {})
            if name and not ar.get('name'):
                ar['name'] = name
                translations['ar'] = ar
                user.translations = translations
                changed = True
            if changed:
                user.save()
            return user, False
        if not email:
            base = f'student.{national_id}' if national_id else f'student-{User.objects.count() + 1}'
            email = f'{base}@student.local'
            suffix = 1
            while User.objects.filter(email=email).exists():
                email = f'{base}.{suffix}@student.local'
                suffix += 1
        user = User.objects.create_user(
            email=email,
            username=email,
            role='student',
            national_id=national_id,
            phone=phone or '',
            translations={'ar': {'name': name}},
            is_verified=True,
        )
        user.set_unusable_password()
        user.save(update_fields=['password'])
        return user, True

    def _import_teachers(self, rows, results):
        for row in rows:
            email = self._cell(row, 'email')
            name = self._cell(row, 'name')
            national_id = self._cell(row, 'national_id') or self._cell(row, 'رقم الهوية') or None
            phone = self._cell(row, 'phone') or self._cell(row, 'هاتف') or ''
            if not email and not national_id:
                results['errors'].append({'row': row, 'error': 'email or national_id is required for teachers'})
                continue
            # Resolve by national_id first, then email
            user = None
            if national_id:
                user = User.objects.filter(national_id=national_id, role='teacher').first()
            if user is None and email:
                user = User.objects.filter(email=email, role='teacher').first()
            if user:
                # Update existing teacher
                changed = False
                if name and not user.translations.get('ar', {}).get('name'):
                    tr = dict(user.translations or {})
                    ar = dict(tr.get('ar') or {})
                    ar['name'] = name
                    tr['ar'] = ar
                    user.translations = tr
                    changed = True
                if national_id and not user.national_id:
                    user.national_id = national_id
                    changed = True
                if phone and not user.phone:
                    user.phone = phone
                    changed = True
                if changed:
                    user.save()
                results['updated'] += 1
            else:
                # Auto-generate unique username and email when not provided
                if not email:
                    base = f'teacher.{national_id}'
                    email = f'{base}@teacher.local'
                    suffix = 1
                    while User.objects.filter(email=email).exists():
                        email = f'{base}.{suffix}@teacher.local'
                        suffix += 1
                user = User.objects.create_user(
                    email=email,
                    username=email,
                    role='teacher',
                    national_id=national_id or '',
                    phone=phone,
                    translations={'ar': {'name': name}} if name else {},
                    is_verified=True,
                )
                user.set_unusable_password()
                user.save(update_fields=['password'])
                results['created'] += 1
            self._assign_teacher(user, row, results)

    def _enroll_by_row(self, student, row, results, school_id=None, section=None):
        from apps.academics.models import Grade
        grade_level = self._row_value(row, self.STUDENT_ALIASES['grade_level'])
        section_name = self._row_value(row, self.STUDENT_ALIASES['section_name'])
        year_name = self._row_value(row, self.STUDENT_ALIASES['academic_year'])
        school_code = self._row_value(row, self.STUDENT_ALIASES['school_code'])
        if section is not None:
            school_id = section.school_id
            grade_level = grade_level or section.grade.level
            section_name = section_name or section.name
            year_name = year_name or section.academic_year.name
            school_code = section.school.school_code
        elif school_id is not None and not school_code:
            school_code = School.objects.filter(id=school_id).values_list('school_code', flat=True).first()
        if not all([grade_level, section_name, year_name, school_code]):
            return
        try:
            grade = Grade.objects.get(level=grade_level)
            year = AcademicYear.objects.get(name=year_name)
            school = School.objects.get(id=school_id) if school_id else School.objects.get(school_code=school_code)
            sec, _ = Section.objects.get_or_create(
                school=school, grade=grade, academic_year=year, name=section_name,
            )
            StudentEnrollment.objects.get_or_create(
                student=student, academic_year=year, defaults={'section': sec},
            )
        except (Grade.DoesNotExist, AcademicYear.DoesNotExist, School.DoesNotExist) as e:
            results['errors'].append({'row': row, 'error': str(e)})

    def _assign_teacher(self, teacher, row, results):
        from apps.academics.models import Subject
        subject_name = row.get('subject')
        section_name = row.get('section_name')
        year_name = row.get('academic_year')
        school_code = row.get('school_code')
        if not all([subject_name, section_name, year_name, school_code]):
            return
        try:
            subject = Subject.objects.filter(translations__ar__name=subject_name).first() or Subject.objects.create(translations={'ar': {'name': subject_name}})
            year = AcademicYear.objects.get(name=year_name)
            school = School.objects.get(school_code=school_code)
            section = Section.objects.filter(
                school=school, academic_year=year, name=section_name,
            ).first()
            if not section:
                results['errors'].append({'row': row, 'error': 'section not found'})
                return
            TeacherAssignment.objects.get_or_create(
                teacher=teacher, section=section, subject=subject, academic_year=year,
            )
        except (AcademicYear.DoesNotExist, School.DoesNotExist) as e:
            results['errors'].append({'row': row, 'error': str(e)})


class BulkExportView(APIView):
    """Bulk export (التصدير الجماعي) of students/teachers as XLSX or CSV.

    School managers and class teachers are scoped to their own school/section.
    `template=1` returns the header row plus a sample row (Eduwave-style template).
    """

    permission_classes = [IsAdminOrSchoolManager]

    STUDENT_HEADERS = [
        ('email', 'البريد الإلكتروني'),
        ('name', 'اسم الطالب'),
        ('national_id', 'الرقم الوطني'),
        ('phone', 'الهاتف'),
        ('parent_email', 'بريد ولي الأمر'),
        ('grade_level', 'الصف'),
        ('section_name', 'الشعبة'),
        ('academic_year', 'السنة الدراسية'),
        ('school_code', 'الرمز المدرسي'),
    ]
    SAMPLE_ROW = {
        'email': 'student@example.com',
        'name': 'محمد أحمد',
        'national_id': '1000000000',
        'phone': '0790000000',
        'parent_email': 'parent@example.com',
        'grade_level': '8',
        'section_name': 'أ',
        'academic_year': '2025-2026',
        'school_code': '1001',
    }

    def get(self, request):
        kind = request.query_params.get('kind', 'students')
        if not is_admin(request.user) and kind != 'students':
            return Response({'error': 'school managers and class teachers may only export students'}, status=status.HTTP_403_FORBIDDEN)
        fmt = (request.query_params.get('file_format') or 'xlsx').lower()
        if fmt not in ('xlsx', 'csv'):
            return Response({'error': 'file_format must be xlsx or csv'}, status=status.HTTP_400_BAD_REQUEST)
        template = str(request.query_params.get('template', '')).lower() in ('1', 'true', 'yes')

        scope = self._resolve_scope(request)

        if kind == 'schools':
            rows = [
                [s.school_code, s.name, s.directorate, s.governorate, s.region, s.gender, s.education_type, s.address]
                for s in School.objects.all().order_by('school_code')
            ]
            headers = ['school_code', 'name', 'directorate', 'governorate', 'region', 'gender', 'education_type', 'address']
        elif kind == 'students':
            headers = [label for _, label in self.STUDENT_HEADERS]
            if template:
                rows = [[self.SAMPLE_ROW[key] for key, _ in self.STUDENT_HEADERS]]
            else:
                enrollments = self._scoped_enrollments(scope)
                rows = []
                for en in enrollments:
                    rows.append([
                        en.student.email,
                        en.student.translations.get('ar', {}).get('name', ''),
                        en.student.national_id or '',
                        en.student.phone or '',
                        ', '.join(g.parent.email for g in en.student.linked_guardians.all()),
                        en.section.grade.level,
                        en.section.name,
                        en.section.academic_year.name,
                        en.section.school.school_code,
                    ])
        elif kind == 'teachers':
            headers = ['email', 'name', 'subject', 'school_code', 'section_name', 'academic_year']
            assignments = TeacherAssignment.objects.select_related('teacher', 'section', 'section__school', 'section__academic_year')
            if not is_admin(request.user):
                assignments = assignments.filter(section__school_id=scope.get('school_id'))
            rows = [[
                a.teacher.email,
                a.teacher.translations.get('ar', {}).get('name', ''),
                a.subject.translations.get('ar', {}).get('name', ''),
                a.section.school.school_code,
                a.section.name,
                a.section.academic_year.name,
            ] for a in assignments]
        else:
            return Response({'error': 'kind must be schools, students or teachers'}, status=status.HTTP_400_BAD_REQUEST)

        filename = 'afaq_students_template.xlsx' if template else f'afaq_{kind}.{fmt}'
        if fmt == 'csv':
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            writer = csv.writer(response)
            writer.writerow(headers)
            for row in rows:
                writer.writerow(row)
            return response
        return self._xlsx_response(headers, rows, filename)

    def _scoped_enrollments(self, scope):
        qs = StudentEnrollment.objects.select_related(
            'student', 'section', 'section__school', 'section__grade', 'section__academic_year',
        ).prefetch_related('student__linked_guardians__parent')
        if scope.get('school_id') is not None:
            qs = qs.filter(section__school_id=scope['school_id'])
        if scope.get('section') is not None:
            qs = qs.filter(section=scope['section'])
        return qs.order_by('section__name', 'student__translations')

    def _xlsx_response(self, headers, rows, filename):
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        wb = Workbook()
        ws = wb.active
        ws.title = 'Sheet1'
        header_fill = PatternFill('solid', fgColor='1F4E79')
        header_font = Font(color='FFFFFF', bold=True)
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for r, row in enumerate(rows, start=2):
            for c, value in enumerate(row, start=1):
                cell = ws.cell(row=r, column=c, value=value)
                cell.alignment = Alignment(horizontal='right')
        ws.column_dimensions['A'].width = 26
        ws.column_dimensions['B'].width = 28
        ws.column_dimensions['C'].width = 18
        for col in ws.columns:
            letter = col[0].column_letter
            if letter not in ('A', 'B', 'C'):
                ws.column_dimensions[letter].width = 16
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def _resolve_scope(self, request):
        user = request.user
        if is_admin(user):
            return {}
        if is_school_admin(user):
            school_ids = user_school_ids(user)
            raw = request.query_params.get('school_id')
            try:
                school_id = int(raw)
            except (TypeError, ValueError):
                school_id = None
            if school_id is None or not school_ids or school_id not in school_ids:
                raise serializers.ValidationError({'school_id': 'school_id is required and must be one of your schools'})
            return {'school_id': school_id}
        if is_teacher(user):
            raw = request.query_params.get('section_id')
            try:
                section = Section.objects.get(id=int(raw), class_teacher=user)
            except (Section.DoesNotExist, TypeError, ValueError) as exc:
                raise serializers.ValidationError({'section_id': 'section_id is required and must be a section you mentor'}) from exc
            return {'school_id': section.school_id, 'section': section}
        raise serializers.ValidationError('Not allowed to export')


class UserSettingsAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        setting, _ = UserAISetting.objects.get_or_create(user=request.user)
        return Response({
            "language_complexity": setting.language_complexity,
            "tone_preference": setting.tone_preference,
            "voice_type": setting.voice_type,
            "context_retrieval": setting.context_retrieval
        })

    def put(self, request):
        setting, _ = UserAISetting.objects.get_or_create(user=request.user)
        setting.language_complexity = request.data.get('language_complexity', setting.language_complexity)
        setting.tone_preference = request.data.get('tone_preference', setting.tone_preference)
        setting.voice_type = request.data.get('voice_type', setting.voice_type)
        setting.context_retrieval = request.data.get('context_retrieval', setting.context_retrieval)
        setting.save()
        return Response({
            "status": "updated",
            "language_complexity": setting.language_complexity,
            "tone_preference": setting.tone_preference,
            "voice_type": setting.voice_type,
            "context_retrieval": setting.context_retrieval
        })


class VoiceTranscribeAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        audio_file = request.FILES.get('audio')
        if not audio_file:
            return Response({'error': 'Audio file is required'}, status=status.HTTP_400_BAD_REQUEST)
        transcription = self._transcribe(audio_file)
        return Response({"text": transcription})

    def _transcribe(self, audio_file):
        """Real STT via Gemini when a key is configured; otherwise a mock fallback."""
        api_key = getattr(settings, 'GEMINI_API_KEY', '')
        if api_key:
            try:
                from google import genai
                client = genai.Client(api_key=api_key)
                response = client.models.generate_content(
                    model="gemini-2.5-flash",
                    contents=["نسخ ما يُقال في هذا الملف الصوتي إلى نص عربي.", audio_file.read()],
                )
                if response.text:
                    return response.text.strip()
            except Exception:
                pass
        return "هذا نص تجريبي تم تحويله من الصوت بنجاح عبر نظام آفاق الصوتي."


class VoiceSynthesizeAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        text = request.data.get('text', '').strip()
        if not text:
            return Response({'error': 'Text is required'}, status=status.HTTP_400_BAD_REQUEST)
        if len(text) > 5000:
            return Response({'error': 'Text exceeds 5000 character limit'}, status=status.HTTP_400_BAD_REQUEST)

        provider = request.data.get('provider', 'gemini')
        speed = float(request.data.get('speed', 1.0))
        locale = request.data.get('locale', 'ar')

        from apps.schools.tts_providers import VoiceSynthesizeAPIView as TTSProvider
        tts = TTSProvider()
        method_name = tts.PROVIDER_MAP.get(provider)
        if not method_name:
            return Response({'error': f'Unknown provider: {provider}'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            audio_bytes = getattr(tts, method_name)(text, locale, speed)
            if audio_bytes:
                from django.http import HttpResponse as DjangoHttpResponse
                response = DjangoHttpResponse(audio_bytes, content_type='audio/mpeg')
                response['Content-Disposition'] = 'inline; filename="speech.mp3"'
                response['Cache-Control'] = 'public, max-age=86400'
                return response
        except Exception as e:
            return Response({'error': f'TTS synthesis failed: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response({'error': 'TTS provider unavailable'}, status=status.HTTP_503_SERVICE_UNAVAILABLE)


class MySchoolContextAPIView(APIView):
    """Returns the school context visible to the authenticated user, scoped by their role."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        user = request.user

        school_filter = request.query_params.get('school')

        if is_admin(user):
            if school_filter:
                schools = School.objects.filter(id=school_filter)
                sections = Section.objects.select_related('school', 'grade', 'academic_year').filter(
                    school_id=school_filter,
                    # Only sections whose grade the school actually offers
                    # (excludes general/stray sections not linked to the school's structure).
                    grade__school_offers__school_id=school_filter,
                )
                sections = sections.annotate(
                    students_count_annotated=Count('students', distinct=True)
                )
                announcements = SchoolAnnouncement.objects.select_related('school', 'section').filter(school_id=school_filter)
                enrollments = StudentEnrollment.objects.select_related('student', 'section', 'section__school', 'section__grade', 'section__academic_year').filter(section__school_id=school_filter)
                tickets = ParentTeacherTicket.objects.select_related('parent', 'teacher', 'student', 'subject').filter(
                    student_id__in=StudentEnrollment.objects.filter(
                        section__school_id=school_filter
                    ).values_list('student_id', flat=True)
                )
            else:
                # No school selected: return only the schools list (for the selector);
                # the heavy context lists are scoped per school to avoid huge payloads.
                schools = list(School.objects.all()[:200])
                sections = Section.objects.none()
                announcements = SchoolAnnouncement.objects.none()
                enrollments = StudentEnrollment.objects.none()
                tickets = ParentTeacherTicket.objects.none()
        else:
            section_ids = user_section_ids(user)
            school_ids = user_school_ids(user)
            if school_filter:
                # Statistics scoped to the active school only.
                section_ids = set(Section.objects.filter(id__in=section_ids, school_id=school_filter).values_list('id', flat=True))
                school_ids = set(school_ids or []) & {int(school_filter)}
            sections = Section.objects.select_related('school', 'grade', 'academic_year')
            if section_ids:
                # Only sections whose grade the school actually offers
                # (excludes general/stray sections not linked to the school's structure).
                sections = sections.filter(
                    id__in=section_ids,
                    grade__school_offers__school_id__in=school_ids or [-1],
                )
            else:
                sections = Section.objects.none()
            sections = sections.annotate(
                students_count_annotated=Count('students', distinct=True)
            )
            schools = School.objects.filter(id__in=school_ids) if school_ids else School.objects.none()

            if is_teacher(user):
                announcements = SchoolAnnouncement.objects.filter(
                    section__teachers__teacher=user,
                ) | SchoolAnnouncement.objects.filter(
                    section__isnull=True, school__in=schools,
                )
                enrollments = StudentEnrollment.objects.filter(section_id__in=section_ids)
                tickets = ParentTeacherTicket.objects.filter(teacher=user)
            elif RoleService.has_role(user, 'student'):
                announcements = SchoolAnnouncement.objects.filter(
                    section_id__in=section_ids,
                ) | SchoolAnnouncement.objects.filter(
                    section__isnull=True, school__in=schools,
                )
                enrollments = StudentEnrollment.objects.filter(student=user)
                tickets = ParentTeacherTicket.objects.filter(
                    parent=user,
                ) | ParentTeacherTicket.objects.filter(
                    student=user,
                )
            elif RoleService.has_role(user, 'parent'):
                child_ids = FamilyLink.objects.filter(parent=user).values_list('student_id', flat=True)
                announcements = SchoolAnnouncement.objects.filter(
                    section_id__in=section_ids,
                ) | SchoolAnnouncement.objects.filter(
                    section__isnull=True, school__in=schools,
                )
                enrollments = StudentEnrollment.objects.filter(student_id__in=child_ids)
                tickets = ParentTeacherTicket.objects.filter(parent=user)
            else:
                announcements = SchoolAnnouncement.objects.none()
                enrollments = StudentEnrollment.objects.none()
                tickets = ParentTeacherTicket.objects.none()

        setting, _ = UserAISetting.objects.get_or_create(user=user)

        if is_admin(user):
            if school_filter:
                teachers = User.objects.filter(role='teacher', assignments__section__school_id=school_filter).distinct()
                students = User.objects.filter(role='student', school_enrollments__section__school_id=school_filter).distinct()
                parents = User.objects.filter(
                    role='parent',
                    parent_tickets__student__school_enrollments__section__school_id=school_filter
                ).distinct()
            else:
                teachers = User.objects.none()
                students = User.objects.none()
                parents = User.objects.none()
        else:
            section_ids = user_section_ids(user)
            teachers = User.objects.filter(assignments__section_id__in=section_ids)
            students = User.objects.filter(school_enrollments__section_id__in=section_ids)
            parents = User.objects.none()

        # Parents can manage their own children; admins see everything.
        if RoleService.has_role(user, 'parent'):
            family_links = FamilyLink.objects.filter(parent=user)
            children_ids = family_links.values_list('student_id', flat=True)
            children = User.objects.filter(id__in=children_ids)
        elif is_admin(user) and school_filter:
            family_links = FamilyLink.objects.select_related('parent', 'student').filter(
                student__school_enrollments__section__school_id=school_filter
            ).distinct()
            children = User.objects.none()
        elif is_admin(user):
            family_links = FamilyLink.objects.none()
            children = User.objects.none()
        else:
            family_links = FamilyLink.objects.filter(student=user) if RoleService.has_role(user, 'student') else FamilyLink.objects.none()
            children = User.objects.none()

        # Weekly reports visible to the user (parent) or for their own sections (teacher/admin/student).
        if RoleService.has_role(user, 'parent'):
            weekly_reports = WeeklyReport.objects.filter(parent=user)
        elif RoleService.has_role(user, 'student'):
            weekly_reports = WeeklyReport.objects.filter(student=user)
        elif is_admin(user) and school_filter:
            weekly_reports = WeeklyReport.objects.filter(
                student__school_enrollments__section__school_id=school_filter
            ).distinct()
        else:
            weekly_reports = WeeklyReport.objects.none()

        # Attendance records visible to the user (own/children/section students).
        if is_admin(user) or RoleService.has_role(user, 'school_admin'):
            attendance_qs = Attendance.objects.select_related('student', 'section', 'school')
        elif RoleService.has_role(user, 'teacher'):
            attendance_qs = Attendance.objects.select_related('student', 'section', 'school').filter(section_id__in=user_section_ids(user))
        elif RoleService.has_role(user, 'parent'):
            child_ids = FamilyLink.objects.filter(parent=user).values_list('student_id', flat=True)
            attendance_qs = Attendance.objects.select_related('student', 'section', 'school').filter(student_id__in=child_ids)
        elif RoleService.has_role(user, 'student'):
            attendance_qs = Attendance.objects.select_related('student', 'section', 'school').filter(student=user)
        else:
            attendance_qs = Attendance.objects.none()
        if school_filter:
            attendance_qs = attendance_qs.filter(school_id=school_filter)

        role_workspace = {
            'teacher': '/teacher',
            'parent': '/parent',
            'student': '/student',
        }
        if is_admin(user) or RoleService.has_role(user, 'school_admin'):
            workspace_url = '/school/admin'
        else:
            workspace_url = role_workspace.get(user.role, '/school/admin')

        attachment_qs = Attachment.objects.all() if is_admin(user) else Attachment.objects.filter(uploader=user)
        if school_filter:
            attachment_qs = attachment_qs.filter(Q(school_id=school_filter) | Q(section__school_id=school_filter))

        return Response({
            "role": user.role,
            "workspace_url": workspace_url,
            "schools": SchoolSerializer(schools, many=True).data,
            "sections": SectionSerializer(sections.distinct(), many=True).data,
            "announcements": SchoolAnnouncementSerializer(announcements.distinct(), many=True).data,
            "enrollments": StudentEnrollmentSerializer(enrollments.distinct(), many=True).data,
            "tickets": ParentTeacherTicketSerializer(tickets.distinct(), many=True).data,
            "teachers": [{"id": t.id, "email": t.email, "name": t.translations.get('ar', {}).get('name', t.email)} for t in teachers.distinct()],
            "students": [{"id": s.id, "email": s.email, "name": s.translations.get('ar', {}).get('name', s.email)} for s in students.distinct()],
            "parents": [{"id": p.id, "email": p.email, "name": p.translations.get('ar', {}).get('name', p.email)} for p in parents.distinct()],
            "family_links": FamilyLinkSerializer(family_links.distinct(), many=True).data,
            "children": [{"id": c.id, "email": c.email, "name": c.translations.get('ar', {}).get('name', c.email)} for c in children.distinct()],
            "weekly_reports": WeeklyReportSerializer(weekly_reports.distinct(), many=True).data,
            "attachments": AttachmentSerializer(
                attachment_qs.distinct(),
                many=True, context={'request': request},
            ).data,
            "attendance": AttendanceSerializer(attendance_qs.distinct(), many=True, context={'request': request}).data,
            "ai_settings": {
                "language_complexity": setting.language_complexity,
                "tone_preference": setting.tone_preference,
                "voice_type": setting.voice_type,
                "context_retrieval": setting.context_retrieval,
            },
        })


class SchoolAnalyticsAPIView(APIView):
    permission_classes = [IsAdminRole]

    def get(self, request):
        return Response({
            "total_schools": School.objects.count(),
            "total_sections": Section.objects.count(),
            "total_announcements": SchoolAnnouncement.objects.count(),
            "emergency_alerts_count": SchoolAnnouncement.objects.filter(is_emergency=True).count(),
            "whatsapp_sent_count": WhatsAppNotificationLog.objects.filter(status='sent').count(),
            "whatsapp_failed_count": WhatsAppNotificationLog.objects.filter(status='failed').count(),
            "active_tickets": ParentTeacherTicket.objects.filter(status='open').count(),
            "attachments_pending_review": Attachment.objects.filter(review_status=Attachment.ReviewStatus.PENDING).count(),
            "attachments_total": Attachment.objects.count(),
            "attendance_today": Attendance.objects.filter(date=timezone.localdate()).count(),
            "absent_today": Attendance.objects.filter(date=timezone.localdate(), status=Attendance.Status.ABSENT).count(),
            "peak_hours": "09:00 AM - 12:00 PM",
            "ai_tokens_used_estimate": 45200,
        })


class PeriodViewSet(viewsets.ModelViewSet):
    queryset = Period.objects.all()
    serializer_class = PeriodSerializer
    permission_classes = [IsAdminOrReadOnly]

    def get_queryset(self):
        qs = Period.objects.filter(is_active=True)
        school_id = self.request.query_params.get('school')
        if school_id:
            qs = qs.filter(school_id=school_id)
        if not is_admin(self.request.user):
            school_ids = user_school_ids(self.request.user)
            qs = qs.filter(school_id__in=school_ids) if school_ids else Period.objects.none()
        return qs

    def _check_school(self, serializer=None, obj=None):
        if is_admin(self.request.user):
            return
        school = (serializer.validated_data.get('school') if serializer else None) or (obj.school if obj else None)
        if not school or school.manager_id != self.request.user.id:
            raise serializers.ValidationError({'school': 'يمكنك إدارة حصص مدرستك فقط'})

    def _check_school_access(self, school_id):
        if is_admin(self.request.user):
            return
        if int(school_id) not in user_school_ids(self.request.user):
            raise serializers.ValidationError({'school': 'يمكنك إدارة حصص مدرستك فقط'})

    def perform_create(self, serializer):
        self._check_school(serializer=serializer)
        serializer.save()

    def perform_update(self, serializer):
        self._check_school(serializer=serializer, obj=self.get_object())
        serializer.save()

    def perform_destroy(self, instance):
        self._check_school(obj=instance)
        instance.delete()

    def _parse_minutes(self, data, field, default):
        try:
            value = int(data.get(field, default))
        except (TypeError, ValueError):
            value = -1
        if value < 1:
            raise serializers.ValidationError({field: 'القيمة يجب أن تكون رقماً موجباً بالدقائق'})
        return value

    def _next_generation(self, school_id):
        from django.db.models import Max
        last = Period.objects.filter(school_id=school_id).aggregate(max_gen=Max('generation'))['max_gen']
        return (last or 0) + 1

    @action(detail=False, methods=['post'], url_path='generate', permission_classes=[permissions.IsAuthenticated])
    def generate_periods(self, request):
        """Generate the full daily schedule from start time, period/break durations and period count.
        Existing active periods are archived (not deleted); a new generation is created."""
        school_id = request.data.get('school_id')
        if not school_id:
            return Response({'error': 'school_id مطلوب'}, status=status.HTTP_400_BAD_REQUEST)
        self._check_school_access(school_id)

        start_value = request.data.get('start_time')
        if not start_value:
            return Response({'error': 'start_time مطلوب'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            from datetime import datetime as _dt
            start_time = _dt.strptime(str(start_value).strip(), '%H:%M').time()
        except ValueError:
            return Response({'error': 'صيغة start_time يجب أن تكون HH:MM'}, status=status.HTTP_400_BAD_REQUEST)

        period_duration = self._parse_minutes(request.data, 'period_duration_min', 45)
        break_duration = self._parse_minutes(request.data, 'break_duration_min', 10)
        long_break_duration = self._parse_minutes(request.data, 'long_break_duration_min', 30)
        total_periods = self._parse_minutes(request.data, 'total_periods', 7)

        try:
            long_break_after = int(request.data.get('long_break_after_period', 3))
        except (TypeError, ValueError):
            long_break_after = 3
        if long_break_after < 1:
            raise serializers.ValidationError({'long_break_after_period': 'يجب أن تكون رقماً موجباً'})

        from datetime import datetime as _dt
        from datetime import timedelta as _td
        current = _dt.combine(_dt.today().date(), start_time)
        new_periods = []
        generation = self._next_generation(school_id)
        now = timezone.now()

        for i in range(1, total_periods + 1):
            start = current
            end = current + _td(minutes=period_duration)
            new_periods.append(Period(
                school_id=int(school_id),
                name=f'الحصة {i}',
                period_number=i,
                start_time=start.time(),
                end_time=end.time(),
                is_break=False,
                is_active=True,
                generation=generation,
                created_by=request.user if request.user.is_authenticated else None,
            ))
            if i == total_periods:
                break
            gap = long_break_duration if i == long_break_after else break_duration
            current = end + _td(minutes=gap)

        with transaction.atomic():
            Period.objects.filter(school_id=int(school_id), is_active=True).update(
                is_active=False,
                archived_at=now,
                archived_by=request.user if request.user.is_authenticated else None,
            )
            Period.objects.bulk_create(new_periods)

        return Response(PeriodSerializer(new_periods, many=True).data, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['get'], url_path='archives', permission_classes=[permissions.IsAuthenticated])
    def archives(self, request):
        """List archived schedule generations with who/when they were created and archived."""
        school_id = request.query_params.get('school_id') or request.query_params.get('school')
        if not school_id:
            return Response({'error': 'school_id مطلوب'}, status=status.HTTP_400_BAD_REQUEST)
        self._check_school_access(school_id)

        rows = list(Period.objects.filter(
            school_id=int(school_id), is_active=False, generation__isnull=False
        ).order_by('generation', 'period_number'))

        groups = {}
        for p in rows:
            groups.setdefault(p.generation, []).append(p)

        result = []
        for generation, items in sorted(groups.items(), reverse=True):
            ordered = sorted(items, key=lambda p: p.period_number)
            result.append({
                'generation': generation,
                'count': len(ordered),
                'created_by_name': str(ordered[0].created_by) if ordered[0].created_by else None,
                'archived_at': ordered[0].archived_at.isoformat() if ordered[0].archived_at else None,
                'archived_by_name': str(ordered[0].archived_by) if ordered[0].archived_by else None,
                'periods': PeriodSerializer(ordered, many=True).data,
            })

        return Response(result)

    @action(detail=False, methods=['post'], url_path='restore', permission_classes=[permissions.IsAuthenticated])
    def restore_generation(self, request):
        """Archive the current active schedule and reactivate a previously archived generation."""
        school_id = request.data.get('school_id')
        generation = request.data.get('generation')
        if not school_id or generation is None:
            return Response({'error': 'school_id و generation مطلوبان'}, status=status.HTTP_400_BAD_REQUEST)
        self._check_school_access(school_id)

        target_exists = Period.objects.filter(
            school_id=int(school_id), generation=int(generation), is_active=False
        ).exists()
        if not target_exists:
            return Response({'error': 'الجيل المطلوب غير موجود في الأرشيف'}, status=status.HTTP_404_NOT_FOUND)

        now = timezone.now()
        user = request.user if request.user.is_authenticated else None
        with transaction.atomic():
            Period.objects.filter(school_id=int(school_id), is_active=True).update(
                is_active=False, archived_at=now, archived_by=user,
            )
            Period.objects.filter(school_id=int(school_id), generation=int(generation)).update(
                is_active=True, archived_at=None, archived_by=None,
            )

        active = Period.objects.filter(school_id=int(school_id), is_active=True).order_by('period_number')
        return Response(PeriodSerializer(active, many=True).data)


class RoomViewSet(viewsets.ModelViewSet):
    queryset = Room.objects.all()
    serializer_class = RoomSerializer
    permission_classes = [IsAdminOrReadOnly]
    pagination_class = None

    def get_queryset(self):
        qs = Room.objects.all()
        school_id = self.request.query_params.get('school')
        if school_id:
            qs = qs.filter(school_id=school_id)
        if not is_admin(self.request.user):
            school_ids = user_school_ids(self.request.user)
            qs = qs.filter(school_id__in=school_ids) if school_ids else Room.objects.none()
        return qs

    def _check_school(self, serializer=None, obj=None):
        if is_admin(self.request.user):
            return
        school = (serializer.validated_data.get('school') if serializer else None) or (obj.school if obj else None)
        if not school or school.manager_id != self.request.user.id:
            raise serializers.ValidationError({'school': 'يمكنك إدارة قاعات مدرستك فقط'})

    def perform_create(self, serializer):
        self._check_school(serializer=serializer)
        serializer.save()

    def perform_update(self, serializer):
        self._check_school(serializer=serializer, obj=self.get_object())
        serializer.save()

    def perform_destroy(self, instance):
        self._check_school(obj=instance)
        instance.delete()


class TimetableSlotViewSet(viewsets.ModelViewSet):
    queryset = TimetableSlot.objects.all()
    serializer_class = TimetableSlotSerializer
    permission_classes = [IsAdminOrReadOnly]

    def get_queryset(self):
        qs = TimetableSlot.objects.all()
        section_id = self.request.query_params.get('section')
        if section_id:
            qs = qs.filter(section_id=section_id)
        teacher_id = self.request.query_params.get('teacher')
        if teacher_id:
            qs = qs.filter(teacher_id=teacher_id)
        room_id = self.request.query_params.get('room')
        if room_id:
            qs = qs.filter(room_id=room_id)
        academic_year_id = self.request.query_params.get('academic_year')
        if academic_year_id:
            qs = qs.filter(academic_year_id=academic_year_id)
        school_id = self.request.query_params.get('school')
        if school_id:
            qs = qs.filter(school_id=school_id)

        if not is_admin(self.request.user):
            section_ids = user_section_ids(self.request.user)
            school_ids = user_school_ids(self.request.user)
            if section_ids is not None or school_ids is not None:
                qs = qs.filter(Q(section_id__in=section_ids or []) | Q(school_id__in=school_ids or []))
            else:
                qs = TimetableSlot.objects.none()
        return qs

    def _pick_room(self, section, subject, rooms_by_type, all_rooms):
        """Pick the best room for a slot based on allocation mode and subject requirements."""
        ssp = SchoolSubjectPeriod.objects.filter(
            school=section.school, grade=section.grade, subject=subject,
        ).first()
        preferred = ssp.preferred_room_type if ssp else ''

        if preferred:
            candidates = rooms_by_type.get(preferred, [])
            for room in candidates:
                if room.capacity >= section.capacity:
                    return room
            if candidates:
                return candidates[0]

        classrooms = rooms_by_type.get('classroom', [])
        for room in classrooms:
            if room.capacity >= section.capacity:
                return room
        if classrooms:
            return classrooms[0]

        return all_rooms.first() if all_rooms else None

    @action(detail=False, methods=['post'], permission_classes=[permissions.IsAuthenticated])
    def auto_schedule(self, request):
        """Smart Auto-Scheduler with room allocation modes.

        fixed mode: each section uses its home_room.
        mobility mode: rooms are picked by subject preferred_room_type + section capacity.
        """
        school_id = request.data.get('school_id')
        academic_year_id = request.data.get('academic_year_id')
        if not school_id or not academic_year_id:
            return Response(
                {'error': 'school_id and academic_year_id are required'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not is_admin(self.request.user):
            school_ids = user_school_ids(self.request.user)
            if int(school_id) not in school_ids:
                return Response({'error': 'Permission denied for this school'}, status=status.HTTP_403_FORBIDDEN)

        try:
            school = School.objects.get(id=school_id)
        except School.DoesNotExist:
            return Response({'error': 'School not found'}, status=status.HTTP_404_NOT_FOUND)

        try:
            year = AcademicYear.objects.get(id=academic_year_id)
        except AcademicYear.DoesNotExist:
            return Response({'error': 'Academic year not found'}, status=status.HTTP_404_NOT_FOUND)

        mode = year.room_allocation_mode
        sections = Section.objects.filter(school_id=school_id, academic_year_id=academic_year_id)
        periods = Period.objects.filter(school_id=school_id, is_active=True, is_break=False).order_by('period_number')
        all_rooms = Room.objects.filter(school_id=school_id)
        week_days = school_week_days(school)

        rooms_by_type = {}
        for room in all_rooms:
            rooms_by_type.setdefault(room.room_type, []).append(room)

        created_slots = []
        errors = []

        for section in sections:
            assignments = list(TeacherAssignment.objects.filter(
                section=section, academic_year_id=academic_year_id,
            ))
            if not assignments or not periods.exists():
                continue

            period_idx = 0
            for day in week_days:
                for period in periods:
                    if period_idx >= len(assignments):
                        break
                    assignment = assignments[period_idx % len(assignments)]

                    exists = TimetableSlot.objects.filter(
                        section=section, day_of_week=day, period=period,
                    ).exists()
                    if not exists:
                        if mode == AcademicYear.ALLOC_FIXED:
                            room = section.home_room
                        else:
                            room = self._pick_room(section, assignment.subject, rooms_by_type, all_rooms)

                        try:
                            slot = TimetableSlot.objects.create(
                                school_id=school_id,
                                academic_year_id=academic_year_id,
                                section=section,
                                day_of_week=day,
                                period=period,
                                subject=assignment.subject,
                                teacher=assignment.teacher,
                                room=room,
                            )
                            created_slots.append(slot.id)
                        except Exception as e:
                            errors.append(str(e))
                    period_idx += 1

        return Response({
            'success': True,
            'mode': mode,
            'created_count': len(created_slots),
            'errors': errors,
        })

    @action(detail=False, methods=['post'], permission_classes=[permissions.IsAuthenticated])
    def setup_fixed_rooms(self, request):
        """Auto-create classrooms for each section and assign as home_room.

        Used when switching to fixed room allocation mode.
        Creates one Room per section (name = section name, capacity = section capacity)
        and links it as the section's home_room.
        """
        school_id = request.data.get('school_id')
        academic_year_id = request.data.get('academic_year_id')
        if not school_id or not academic_year_id:
            return Response(
                {'error': 'school_id and academic_year_id are required'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not is_admin(self.request.user):
            school_ids = user_school_ids(self.request.user)
            if int(school_id) not in school_ids:
                return Response({'error': 'Permission denied for this school'}, status=status.HTTP_403_FORBIDDEN)

        sections = Section.objects.filter(
            school_id=school_id,
            academic_year_id=academic_year_id,
            home_room__isnull=True,
            grade__school_offers__school_id=school_id,
        ).select_related('grade')

        created_rooms = []
        linked_sections = 0

        for section in sections:
            grade_name = section.grade.translations.get('ar', {}).get('name', str(section.grade.level))
            room_name = f"{grade_name} - {section.name}"

            room, _ = Room.objects.get_or_create(
                school_id=school_id,
                name=room_name,
                defaults={
                    'code': f"S{section.grade.level}-{section.name}",
                    'capacity': section.capacity,
                    'room_type': 'classroom',
                },
            )
            created_rooms.append(room.id)

            if section.home_room_id is None:
                section.home_room = room
                section.save(update_fields=['home_room'])
                linked_sections += 1

        return Response({
            'success': True,
            'rooms_created': len(created_rooms),
            'sections_linked': linked_sections,
        })

    @action(detail=False, methods=['get'], permission_classes=[permissions.IsAuthenticated])
    def export_pdf(self, request):
        """Export timetable for a section or teacher as PDF."""
        section_id = request.query_params.get('section')
        teacher_id = request.query_params.get('teacher')
        school_id = request.query_params.get('school')
        academic_year_id = request.query_params.get('academic_year')
        locale = request.query_params.get('locale', 'ar')

        qs = TimetableSlot.objects.select_related(
            'section', 'section__grade', 'subject', 'teacher', 'period', 'room',
        )
        if section_id:
            qs = qs.filter(section_id=section_id)
        elif teacher_id:
            qs = qs.filter(teacher_id=teacher_id)
        elif school_id:
            qs = qs.filter(school_id=school_id)
        if academic_year_id:
            qs = qs.filter(academic_year_id=academic_year_id)

        slots = list(qs.order_by('day_of_week', 'period__period_number'))

        if not slots:
            return Response({'error': 'No timetable slots found'}, status=status.HTTP_404_NOT_FOUND)

        school_name = slots[0].school.name if slots else ''
        section_name = str(slots[0].section) if section_id and slots else ''
        teacher_name = slots[0].teacher.translations.get('ar', {}).get('name', slots[0].teacher.email) if teacher_id and slots else ''

        day_names = {
            1: 'الإثنين', 2: 'الثلاثاء', 3: 'الأربعاء',
            4: 'الخميس', 5: 'الجمعة', 6: 'السبت', 7: 'الأحد',
        }
        if locale == 'en':
            day_names = {1: 'Monday', 2: 'Tuesday', 3: 'Wednesday', 4: 'Thursday', 5: 'Friday', 6: 'Saturday', 7: 'Sunday'}

        # Group by day
        grid: dict[int, list] = {d: [] for d in range(1, 8)}
        for slot in slots:
            grid[slot.day_of_week].append(slot)

        title = f'الجدول الدراسي — {school_name}'
        if section_name:
            title += f' | {section_name}'
        if teacher_name:
            title += f' | {teacher_name}'

        rows_html = ''
        for day_num in range(1, 8):
            day_slots = sorted(grid[day_num], key=lambda s: s.period.period_number)
            cells = ''
            for s in day_slots:
                cells += '<td style="border:1px solid #ccc;padding:6px;text-align:center;font-size:12px">'
                cells += f'<b>{s.subject.name}</b><br>{s.teacher.translations.get("ar",{}).get("name",s.teacher.email)}<br>'
                if s.room:
                    cells += f'<small>{s.room.name}</small>'
                cells += '</td>'
            rows_html += f'<tr><td style="border:1px solid #ccc;padding:6px;font-weight:bold;background:#f0f0f0">{day_names.get(day_num, day_num)}</td>{cells}</tr>'

        html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>@page{{size:A4 landscape;margin:1cm}}body{{font-family:sans-serif;direction:rtl}}</style>
</head><body>
<h2 style="text-align:center">{title}</h2>
<table style="width:100%;border-collapse:collapse;margin-top:20px">
<thead><tr style="background:#333;color:#fff">
<th style="border:1px solid #ccc;padding:8px">اليوم</th>
<th style="border:1px solid #ccc;padding:8px">الحصة</th><th style="border:1px solid #ccc;padding:8px">الحصة</th><th style="border:1px solid #ccc;padding:8px">الحصة</th><th style="border:1px solid #ccc;padding:8px">الحصة</th><th style="border:1px solid #ccc;padding:8px">الحصة</th>
</tr></thead>
<tbody>{rows_html}</tbody>
</table>
<p style="text-align:center;color:#888;margin-top:20px">آفاق تكنولوجي — نظام المتابعة المدرسية الذكية</p>
</body></html>"""

        from django.http import HttpResponse
        try:
            from weasyprint import HTML as WeasyHTML
            pdf_bytes = WeasyHTML(string=html).write_pdf()
            response = HttpResponse(pdf_bytes, content_type='application/pdf')
            filename = f'timetable_{section_name or teacher_name or school_name}.pdf'.replace(' ', '_')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
        except ImportError:
            # Fallback: return HTML if WeasyPrint is not installed
            return HttpResponse(html, content_type='text/html')

    @action(detail=False, methods=['get'], permission_classes=[permissions.IsAuthenticated])
    def export_excel(self, request):
        """Export timetable for a section or teacher as XLSX."""
        section_id = request.query_params.get('section')
        teacher_id = request.query_params.get('teacher')
        school_id = request.query_params.get('school')
        academic_year_id = request.query_params.get('academic_year')

        qs = TimetableSlot.objects.select_related(
            'section', 'section__grade', 'subject', 'teacher', 'period', 'room',
        )
        if section_id:
            qs = qs.filter(section_id=section_id)
        elif teacher_id:
            qs = qs.filter(teacher_id=teacher_id)
        elif school_id:
            qs = qs.filter(school_id=school_id)
        if academic_year_id:
            qs = qs.filter(academic_year_id=academic_year_id)

        slots = list(qs.order_by('day_of_week', 'period__period_number'))

        if not slots:
            return Response({'error': 'No timetable slots found'}, status=status.HTTP_404_NOT_FOUND)

        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'الجدول الدراسي'

        day_names = {
            1: 'الإثنين', 2: 'الثلاثاء', 3: 'الأربعاء',
            4: 'الخميس', 5: 'الجمعة', 6: 'السبت', 7: 'الأحد',
        }

        ws.append(['اليوم', 'الحصة', 'الوقت', 'المادة', 'المعلم', 'الشعبة', 'القاعة'])
        for s in slots:
            ws.append([
                day_names.get(s.day_of_week, str(s.day_of_week)),
                s.period.name,
                f'{s.period.start_time} - {s.period.end_time}',
                s.subject.name,
                s.teacher.translations.get('ar', {}).get('name', s.teacher.email),
                str(s.section),
                s.room.name if s.room else '',
            ])

        from django.http import HttpResponse
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f'timetable_{section_id or teacher_id or school_id}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    @action(detail=True, methods=['patch'], permission_classes=[permissions.IsAuthenticated])
    def move(self, request, pk=None):
        """Move a timetable slot to a new day/period via drag-and-drop.

        PATCH /timetable-slots/{id}/move/
        Body: { "day_of_week": 1, "period_id": 5 }
        Validates triple conflict (section, teacher, room) before applying.
        """
        slot = self.get_object()
        new_day = request.data.get('day_of_week')
        new_period_id = request.data.get('period_id')

        if new_day is None or new_period_id is None:
            return Response(
                {'error': 'day_of_week and period_id are required'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        new_day = int(new_day)
        new_period_id = int(new_period_id)

        try:
            new_period = Period.objects.get(id=new_period_id, school=slot.school, is_active=True)
        except Period.DoesNotExist:
            return Response({'error': 'Period not found'}, status=status.HTTP_404_NOT_FOUND)

        if not is_admin(self.request.user):
            school_ids = user_school_ids(self.request.user)
            if slot.school_id not in (school_ids or []):
                return Response({'error': 'Permission denied'}, status=status.HTTP_403_FORBIDDEN)

        existing_same = TimetableSlot.objects.filter(
            section=slot.section, academic_year=slot.academic_year,
            day_of_week=new_day, period=new_period,
        ).exclude(pk=slot.pk).exists()
        if existing_same:
            return Response(
                {'error': 'الشعبة الصفية لديها حصة مسجلة بالفعل في هذا الوقت.',
                 'conflict_type': 'section'},
                status=status.HTTP_409_CONFLICT,
            )

        teacher_conflict = TimetableSlot.objects.filter(
            teacher=slot.teacher, academic_year=slot.academic_year,
            day_of_week=new_day, period=new_period,
        ).exclude(pk=slot.pk).exists()
        if teacher_conflict:
            return Response(
                {'error': 'المعلم مرتبط بحصة أخرى في نفس هذا الوقت لشعبة أخرى.',
                 'conflict_type': 'teacher'},
                status=status.HTTP_409_CONFLICT,
            )

        if slot.room:
            room_conflict = TimetableSlot.objects.filter(
                room=slot.room, academic_year=slot.academic_year,
                day_of_week=new_day, period=new_period,
            ).exclude(pk=slot.pk).exists()
            if room_conflict:
                return Response(
                    {'error': 'القاعة / المختبر محجوزة بالفعل في نفس هذا الوقت.',
                     'conflict_type': 'room'},
                    status=status.HTTP_409_CONFLICT,
                )

        slot.day_of_week = new_day
        slot.period = new_period
        slot.save(update_fields=['day_of_week', 'period'])

        return Response(TimetableSlotSerializer(slot, context={'request': request}).data)


def check_school_subscription(user, school=None):
    if is_admin(user):
        return True
    if not (user and user.is_authenticated):
        return False

    level = user.get_subscription_level() if hasattr(user, 'get_subscription_level') else 0
    if level >= 2 or getattr(user, 'subscription_plan', '') in ('school', 'enterprise'):
        return True

    try:
        from apps.subscriptions.services import get_user_plan
        plan = get_user_plan(user)
        if plan and (plan.level >= 2 or plan.code in ('school', 'enterprise')):
            return True
    except Exception:
        pass

    if school and school.manager:
        mgr = school.manager
        mgr_level = mgr.get_subscription_level() if hasattr(mgr, 'get_subscription_level') else 0
        if mgr_level >= 2 or getattr(mgr, 'subscription_plan', '') in ('school', 'enterprise'):
            return True
        try:
            from apps.subscriptions.services import get_user_plan
            mgr_plan = get_user_plan(mgr)
            if mgr_plan and (mgr_plan.level >= 2 or mgr_plan.code in ('school', 'enterprise')):
                return True
        except Exception:
            pass
    return False


class AuxiliaryModuleGatingMixin:
    def _check_gating(self, request, school=None):
        if request.method in permissions.SAFE_METHODS:
            return None
        if not school:
            school = self._resolve_school(request)
        if not check_school_subscription(request.user, school):
            return Response(
                {
                    'detail': 'يتطلب اشتراك باقة المدرسة (School) أو الريادة (Enterprise) لإنشاء أو إدارة الوحدات الإضافية (الرسوم، النقل، المكتبة). يرجى الترقية إلى باقة المدرسة أو الريادة.',
                    'error': 'subscription_required',
                    'upgrade_url': '/subscriptions',
                },
                status=status.HTTP_403_FORBIDDEN,
            )
        return None

    def _resolve_school(self, request):
        raw_school = request.data.get('school') if hasattr(request, 'data') else None
        if raw_school:
            try:
                return School.objects.get(pk=raw_school)
            except School.DoesNotExist:
                pass

        for key in ('bus', 'route', 'book', 'fee', 'student'):
            val = request.data.get(key) if hasattr(request, 'data') else None
            if val:
                try:
                    if key == 'bus':
                        return SchoolBus.objects.get(pk=val).school
                    elif key == 'route':
                        return BusRoute.objects.get(pk=val).bus.school
                    elif key == 'book':
                        return Book.objects.get(pk=val).school
                    elif key == 'fee':
                        return SchoolFee.objects.get(pk=val).school
                    elif key == 'student':
                        en = StudentEnrollment.objects.filter(student_id=val).select_related('section__school').first()
                        if en:
                            return en.section.school
                except Exception:
                    pass

        try:
            obj = self.get_object()
            if obj:
                return self._get_school_from_obj(obj)
        except Exception:
            pass
        return None

    def _get_school_from_obj(self, obj):
        if hasattr(obj, 'school') and obj.school:
            return obj.school
        if hasattr(obj, 'bus') and obj.bus and obj.bus.school:
            return obj.bus.school
        if hasattr(obj, 'route') and obj.route and obj.route.bus and obj.route.bus.school:
            return obj.route.bus.school
        if hasattr(obj, 'book') and obj.book and obj.book.school:
            return obj.book.school
        if hasattr(obj, 'fee') and obj.fee and obj.fee.school:
            return obj.fee.school
        return None

    def create(self, request, *args, **kwargs):
        res = self._check_gating(request)
        if res:
            return res
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        res = self._check_gating(request)
        if res:
            return res
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        res = self._check_gating(request)
        if res:
            return res
        return super().partial_update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        res = self._check_gating(request)
        if res:
            return res
        return super().destroy(request, *args, **kwargs)


class SchoolFeeViewSet(AuxiliaryModuleGatingMixin, viewsets.ModelViewSet):
    queryset = SchoolFee.objects.all()
    serializer_class = SchoolFeeSerializer
    permission_classes = [IsAdminOrReadOnly]

    def get_queryset(self):
        qs = SchoolFee.objects.all()
        school_id = self.request.query_params.get('school')
        if school_id:
            qs = qs.filter(school_id=school_id)
        return qs


class StudentFeeAssignmentViewSet(AuxiliaryModuleGatingMixin, viewsets.ModelViewSet):
    queryset = StudentFeeAssignment.objects.all()
    serializer_class = StudentFeeAssignmentSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        qs = StudentFeeAssignment.objects.all()
        student_id = self.request.query_params.get('student')
        if student_id:
            qs = qs.filter(student_id=student_id)
        if not is_admin(self.request.user):
            qs = qs.filter(student=self.request.user)
        return qs


class StudentReportCardPDFView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, student_id):
        student = get_object_or_404(User, pk=student_id)
        enrollment = StudentEnrollment.objects.filter(student=student).select_related('school', 'section').first()
        school = enrollment.school if enrollment else None
        attendance_count = Attendance.objects.filter(student=student, status='present').count()
        absent_count = Attendance.objects.filter(student=student, status='absent').count()

        html_content = f"""
        <!DOCTYPE html>
        <html lang="ar" dir="rtl">
        <head>
        <meta charset="utf-8">
        <style>
            @import url('https://fonts.googleapis.com/css2?family=Cairo:wght@400;700&display=swap');
            body {{ font-family: 'Cairo', Arial, sans-serif; direction: rtl; text-align: right; color: #1e293b; padding: 20px; }}
            .header {{ text-align: center; border-bottom: 2px solid #4f46e5; padding-bottom: 15px; margin-bottom: 20px; }}
            .school-name {{ font-size: 20pt; font-weight: bold; color: #4f46e5; }}
            .title {{ font-size: 16pt; font-weight: bold; margin-top: 5px; }}
            .info-box {{ background: #f8fafc; border: 1px solid #e2e8f0; padding: 15px; border-radius: 8px; margin-bottom: 20px; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 15px; }}
            th, td {{ border: 1px solid #cbd5e1; padding: 10px; text-align: center; font-size: 11pt; }}
            th {{ background: #f1f5f9; font-weight: bold; }}
            .footer {{ margin-top: 40px; text-align: center; font-size: 10pt; color: #64748b; }}
        </style>
        </head>
        <body>
            <div class="header">
                <div class="school-name">{school.name if school else 'المدرسة النموذجية'}</div>
                <div class="title">كشف الجلاء المدرسي الرسمي</div>
            </div>
            <div class="info-box">
                <p><strong>اسم الطالب:</strong> {student.get_full_name() or student.email}</p>
                <p><strong>الشعبة:</strong> {enrollment.section.name if enrollment and enrollment.section else 'غير محدد'}</p>
            </div>
            <h3>ملخص الحضور والغياب</h3>
            <table>
                <tr>
                    <th>الأيام الحاضرة</th>
                    <th>الأيام الغائبة</th>
                </tr>
                <tr>
                    <td style="color: #059669; font-weight: bold;">{attendance_count}</td>
                    <td style="color: #dc2626; font-weight: bold;">{absent_count}</td>
                </tr>
            </table>
            <div class="footer">
                <p>تم استخراج هذا التقرير إلكترونياً من منصة آفاق تكنولوجي (Afaq Tech SIS)</p>
            </div>
        </body>
        </html>
        """
        from weasyprint import HTML
        pdf_bytes = HTML(string=html_content).write_pdf()
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="report-card-{student_id}.pdf"'
        return response


class BiometricWebhookAPIView(APIView):
    """Receives attendance webhooks from biometric devices / RFID turnstiles."""
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        student_id = request.data.get('student_id')
        request.data.get('device_id')
        status_val = request.data.get('status', 'present')
        request.data.get('timestamp')

        if not student_id:
            return Response({'error': 'student_id is required'}, status=status.HTTP_400_BAD_REQUEST)

        student = get_object_or_404(User, pk=student_id)

        attendance, created = Attendance.objects.update_or_create(
            student=student,
            date=timezone.localdate(),
            defaults={'status': status_val}
        )

        try:
            from .whatsapp import send_whatsapp_alert
            family_links = FamilyLink.objects.filter(student=student)
            for link in family_links:
                if link.parent and link.parent.phone:
                    msg = f"إشعار حضور مدرسي: تم تسجيل حضور الطالب {student.get_full_name() or student.email} في المدرسة بتاريخ {attendance.date}."
                    send_whatsapp_alert(link.parent.phone, msg)
        except Exception:
            pass

        return Response({'success': True, 'attendance_id': attendance.id})


class StudentPredictiveAnalyticsAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, student_id):
        student = get_object_or_404(User, pk=student_id)

        attendance_records = Attendance.objects.filter(student=student)
        total_days = attendance_records.count()
        absent_days = attendance_records.filter(status='absent').count()
        absent_ratio = (absent_days / total_days) if total_days > 0 else 0.0

        grade_entries = GradeEntry.objects.filter(student=student).select_related('category')
        grades_summary = []
        for entry in grade_entries:
            grades_summary.append({
                'subject': entry.category.subject.translations.get('ar', {}).get('name', str(entry.category.subject_id)) if hasattr(entry.category.subject, 'translations') else str(entry.category.subject_id),
                'category': entry.category.name,
                'score': float(entry.score),
                'max_score': float(entry.category.max_score),
                'percentage': round(float(entry.percentage), 2) if entry.category.max_score else 0,
            })

        assignments = AssignmentSubmission.objects.filter(student=student).select_related('assignment')
        assignment_stats = {
            'total': assignments.count(),
            'submitted': assignments.filter(status__in=['submitted', 'graded']).count(),
            'graded': assignments.filter(status='graded').count(),
            'late': 0,
        }
        for sub in assignments.select_related('assignment'):
            if sub.submitted_at and sub.assignment.due_date and sub.submitted_at > sub.assignment.due_date:
                assignment_stats['late'] += 1

        from django.db.models import Avg
        avg_score = grade_entries.aggregate(avg=Avg('score'))['avg']

        context_payload = {
            'student': {
                'id': student.id,
                'name': student.get_full_name() or student.email,
                'email': student.email,
            },
            'attendance': {
                'total_days': total_days,
                'absent_days': absent_days,
                'absence_ratio': round(absent_ratio * 100, 2),
            },
            'grades': {
                'entries': grades_summary[:20],
                'average_score': round(float(avg_score), 2) if avg_score else None,
                'total_entries': grade_entries.count(),
            },
            'assignments': assignment_stats,
        }

        rule_risk = "low"
        rule_recommendations = ["الاستمرار في الأداء المنتظم والمشاركة الصفية."]
        if absent_ratio > 0.15:
            rule_risk = "high"
        elif absent_ratio > 0.08:
            rule_risk = "medium"

        ai_analysis = None
        try:
            from apps.ai.router import ProviderRouter
            router = ProviderRouter()

            prompt = (
                "أنت محلل تعليمي متخصص. حلل بيانات الطالب التالية وأعد تحليلاً تنبؤياً بالشكل JSON التالي:\n"
                '{\n'
                '  "risk_level": "high|medium|low",\n'
                '  "risk_score": 0-100,\n'
                '  "summary": "ملخص مختصر بالعربية",\n'
                '  "strengths": ["نقطة قوة 1", "نقطة قوة 2"],\n'
                '  "weaknesses": ["نقطة ضعف 1", "نقطة ضعف 2"],\n'
                '  "recommendations": [\n'
                '    {"action": "إجراء مقترح", "target": "المعلم/ولي الأمر/الطالب", "priority": "high|medium|low"}\n'
                '  ],\n'
                '  "predicted_trend": "improving|stable|declining"\n'
                '}\n\n'
                f"بيانات الطالب:\n{json.dumps(context_payload, ensure_ascii=False, default=str)}"
            )

            response = router.generate(prompt, feature='analytics')
            if response and response.success:
                ai_text = response.content.strip()
                if ai_text.startswith('```'):
                    ai_text = ai_text.split('\n', 1)[1].rsplit('```', 1)[0]
                ai_analysis = json.loads(ai_text)
        except Exception:
            ai_analysis = None

        return Response({
            'student_id': student.id,
            'student_name': student.get_full_name() or student.email,
            'attendance': context_payload['attendance'],
            'grades_summary': context_payload['grades'],
            'assignments': context_payload['assignments'],
            'risk_level': ai_analysis.get('risk_level', rule_risk) if ai_analysis else rule_risk,
            'risk_score': ai_analysis.get('risk_score') if ai_analysis else None,
            'summary': ai_analysis.get('summary') if ai_analysis else None,
            'strengths': ai_analysis.get('strengths', []) if ai_analysis else [],
            'weaknesses': ai_analysis.get('weaknesses', []) if ai_analysis else [],
            'recommendations': ai_analysis.get('recommendations', [{'action': r, 'target': 'المعلم', 'priority': 'medium'} for r in rule_recommendations]) if ai_analysis else [{'action': r, 'target': 'المعلم', 'priority': 'medium'} for r in rule_recommendations],
            'predicted_trend': ai_analysis.get('predicted_trend', 'stable') if ai_analysis else 'stable',
            'ai_powered': ai_analysis is not None,
        })


class SchoolBusViewSet(AuxiliaryModuleGatingMixin, viewsets.ModelViewSet):
    queryset = SchoolBus.objects.all()
    serializer_class = SchoolBusSerializer
    permission_classes = [IsAdminOrReadOnly]

    def get_queryset(self):
        qs = SchoolBus.objects.all()
        school_id = self.request.query_params.get('school')
        if school_id:
            qs = qs.filter(school_id=school_id)
        return qs


class BusRouteViewSet(AuxiliaryModuleGatingMixin, viewsets.ModelViewSet):
    queryset = BusRoute.objects.all()
    serializer_class = BusRouteSerializer
    permission_classes = [IsAdminOrReadOnly]

    def get_queryset(self):
        qs = BusRoute.objects.all()
        bus_id = self.request.query_params.get('bus')
        if bus_id:
            qs = qs.filter(bus_id=bus_id)
        return qs


class StudentBusAssignmentViewSet(AuxiliaryModuleGatingMixin, viewsets.ModelViewSet):
    queryset = StudentBusAssignment.objects.all()
    serializer_class = StudentBusAssignmentSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        qs = StudentBusAssignment.objects.all()
        student_id = self.request.query_params.get('student')
        if student_id:
            qs = qs.filter(student_id=student_id)
        if not is_admin(self.request.user):
            qs = qs.filter(student=self.request.user)
        return qs


class BookViewSet(AuxiliaryModuleGatingMixin, viewsets.ModelViewSet):
    queryset = Book.objects.all()
    serializer_class = BookSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        qs = Book.objects.all()
        school_id = self.request.query_params.get('school')
        if school_id:
            qs = qs.filter(school_id=school_id)
        return qs


class LibraryLendingViewSet(AuxiliaryModuleGatingMixin, viewsets.ModelViewSet):
    queryset = LibraryLending.objects.all()
    serializer_class = LibraryLendingSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        qs = LibraryLending.objects.select_related('book', 'borrower').all()
        school_id = self.request.query_params.get('school')
        if school_id:
            qs = qs.filter(book__school_id=school_id)
        borrower_id = self.request.query_params.get('borrower')
        if borrower_id:
            qs = qs.filter(borrower_id=borrower_id)
        if not is_admin(self.request.user) and RoleService.has_role(self.request.user, 'student'):
            qs = qs.filter(borrower=self.request.user)
        return qs

    @action(detail=False, methods=['get'], url_path='people')
    def people(self, request):
        """List school students/teachers/parents to pick a borrower for a lending record."""
        school_id = request.query_params.get('school')
        role = request.query_params.get('role', '')
        if not school_id:
            return Response({'students': [], 'teachers': [], 'parents': []})

        def qs(role_filter):
            return User.objects.filter(
                role=role_filter,
                school_enrollments__section__school_id=school_id,
            ).distinct()

        def qs_teachers():
            return User.objects.filter(
                role='teacher',
                assignments__section__school_id=school_id,
            ).distinct()

        def qs_parents():
            return User.objects.filter(
                role='parent',
                parent_tickets__student__school_enrollments__section__school_id=school_id,
            ).distinct()

        def serialize(users):
            return [
                {
                    'id': u.id,
                    'email': u.email,
                    'name': u.translations.get('ar', {}).get('name', u.email),
                }
                for u in users
            ]

        if role == 'student':
            return Response({'students': serialize(qs('student'))})
        if role == 'teacher':
            return Response({'teachers': serialize(qs_teachers())})
        if role == 'parent':
            return Response({'parents': serialize(qs_parents())})
        return Response({
            'students': serialize(qs('student')),
            'teachers': serialize(qs_teachers()),
            'parents': serialize(qs_parents()),
        })


class FAQCopilotAPIView(APIView):
    """FAQ Copilot — AI-powered auto-reply for parent questions.

    Accepts a question and optional school context, searches the FAQ database
    for relevant entries, and uses AI to generate a helpful response. Returns
    both the AI answer and any matching FAQ entries.
    """
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        question = (request.data.get('question') or '').strip()
        school_id = request.data.get('school')
        if not question:
            return Response({'error': 'question is required'}, status=status.HTTP_400_BAD_REQUEST)

        # Gather context
        faqs = list(FAQ.objects.filter(is_active=True).values_list('question', 'answer')[:20])
        faq_context = '\n'.join(f'- س: {q}\n  ج: {a}' for q, a in faqs) if faqs else 'لا توجد أسئلة شائعة مسجلة.'

        school_context = ''
        if school_id:
            try:
                school = School.objects.get(id=school_id)
                school_context = f'المدرسة: {school.name} ({school.directorate})'
            except School.DoesNotExist:
                pass

        # Check if the question matches any FAQ closely
        matched_faq = None
        q_lower = question.lower()
        for fq, fa in faqs:
            if q_lower in fq.lower() or fq.lower() in q_lower:
                matched_faq = {'question': fq, 'answer': fa}
                break

        # If a close FAQ match is found, return it directly without AI
        if matched_faq:
            return Response({
                'answer': matched_faq['answer'],
                'source': 'faq_match',
                'matched_faq': matched_faq,
                'faqs': [{'question': fq, 'answer': fa} for fq, fa in faqs[:10]],
            })

        # Use AI to generate a response
        system_prompt = f"""أنت مساعد ذكي لإدارة المدرسة. أجب على سؤال ولي الأمر بأسلوب مهني وواضح.
إذا كان السؤال متعلقاً ب规则 أو سياسات المدرسة، استخدم معلومات الأسئلة الشائعة التالية كمرجع:
{faq_context}
{school_context}
إذا لم تجد إجابة واضحة، قل ذلك بأدب وانصح بالتواصل مع الإدارة المدرسية.
أجب بالعربية."""

        try:
            from apps.ai.router import ProviderRouter
            router = ProviderRouter()
            result = router.generate(
                prompt=question,
                feature='general',
                system_instruction=system_prompt,
                use_cache=False,
            )
            answer = result.text if result and result.text else 'عذراً، لم أتمكن من إيجاد إجابة مناسبة.'
        except Exception:
            answer = 'عذراً،خدمة الذكاء الاصطناعي غير متاحة حالياً. يرجى التواصل مع الإدارة المدرسية مباشرة.'

        return Response({
            'answer': answer,
            'source': 'ai_copilot',
            'faqs': [{'question': fq, 'answer': fa} for fq, fa in faqs[:10]],
        })


# ---------------------------------------------------------------------------
# Grade Book ViewSets
# ---------------------------------------------------------------------------

class GradeCategoryViewSet(viewsets.ModelViewSet):
    serializer_class = GradeCategorySerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        qs = GradeCategory.objects.select_related('subject', 'school').all()
        school_id = self.request.query_params.get('school')
        if school_id:
            qs = qs.filter(school_id=school_id)
        subject = self.request.query_params.get('subject')
        if subject:
            qs = qs.filter(subject_id=subject)
        if not is_admin(self.request.user):
            if RoleService.has_role(self.request.user, 'school_admin'):
                school_ids = user_school_ids(self.request.user)
                qs = qs.filter(school_id__in=school_ids) if school_ids else qs.none()
            else:
                qs = qs.none()
        return qs

    def perform_create(self, serializer):
        serializer.save()


class GradeEntryViewSet(viewsets.ModelViewSet):
    serializer_class = GradeEntrySerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        qs = GradeEntry.objects.select_related(
            'category', 'student', 'section', 'graded_by'
        ).all()
        user = self.request.user
        if is_admin(user):
            pass
        elif RoleService.has_role(user, 'school_admin'):
            school_ids = user_school_ids(user)
            qs = qs.filter(section__school_id__in=school_ids) if school_ids else qs.none()
        elif RoleService.has_role(user, 'teacher'):
            section_ids = user_section_ids(user)
            qs = qs.filter(section_id__in=section_ids) if section_ids else qs.none()
        elif RoleService.has_role(user, 'parent'):
            child_ids = FamilyLink.objects.filter(parent=user).values_list('student_id', flat=True)
            qs = qs.filter(student_id__in=child_ids)
        elif RoleService.has_role(user, 'student'):
            qs = qs.filter(student=user)
        else:
            qs = qs.none()

        section = self.request.query_params.get('section')
        if section:
            qs = qs.filter(section_id=section)
        student = self.request.query_params.get('student')
        if student:
            qs = qs.filter(student_id=student)
        category = self.request.query_params.get('category')
        if category:
            qs = qs.filter(category_id=category)
        subject = self.request.query_params.get('subject')
        if subject:
            qs = qs.filter(category__subject_id=subject)
        return qs

    def perform_create(self, serializer):
        serializer.save(graded_by=self.request.user)

    @action(detail=False, methods=['post'])
    def bulk_create(self, request):
        """Bulk create/update grade entries for a section+category."""
        category_id = request.data.get('category')
        section_id = request.data.get('section')
        grades = request.data.get('grades', [])  # [{student, score, notes}, ...]
        if not category_id or not section_id or not grades:
            return Response({'error': 'category, section, and grades are required'},
                            status=status.HTTP_400_BAD_REQUEST)
        try:
            category = GradeCategory.objects.get(id=category_id)
        except GradeCategory.DoesNotExist:
            return Response({'error': 'Category not found'}, status=status.HTTP_404_NOT_FOUND)
        created, updated = 0, 0
        with transaction.atomic():
            for g in grades:
                student_id = g.get('student')
                score = g.get('score')
                notes = g.get('notes', '')
                if student_id is None or score is None:
                    continue
                obj, was_created = GradeEntry.objects.update_or_create(
                    category=category, student_id=student_id,
                    defaults={
                        'section_id': section_id,
                        'score': score,
                        'notes': notes,
                        'graded_by': request.user,
                    }
                )
                if was_created:
                    created += 1
                else:
                    updated += 1
        return Response({'created': created, 'updated': updated})


# ---------------------------------------------------------------------------
# Assignment ViewSets
# ---------------------------------------------------------------------------

class AssignmentViewSet(viewsets.ModelViewSet):
    serializer_class = AssignmentSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        qs = Assignment.objects.select_related('section', 'subject', 'teacher').all()
        user = self.request.user
        if is_admin(user):
            pass
        elif RoleService.has_role(user, 'school_admin'):
            school_ids = user_school_ids(user)
            qs = qs.filter(section__school_id__in=school_ids) if school_ids else qs.none()
        elif RoleService.has_role(user, 'teacher'):
            qs = qs.filter(teacher=user)
        elif RoleService.has_role(user, 'parent'):
            child_ids = FamilyLink.objects.filter(parent=user).values_list('student_id', flat=True)
            child_section_ids = StudentEnrollment.objects.filter(student_id__in=child_ids).values_list('section_id', flat=True)
            qs = qs.filter(section_id__in=child_section_ids)
        elif RoleService.has_role(user, 'student'):
            enrolled_section_ids = StudentEnrollment.objects.filter(student=user).values_list('section_id', flat=True)
            qs = qs.filter(section_id__in=enrolled_section_ids)
        else:
            qs = qs.none()

        section = self.request.query_params.get('section')
        if section:
            qs = qs.filter(section_id=section)
        subject = self.request.query_params.get('subject')
        if subject:
            qs = qs.filter(subject_id=subject)
        return qs

    def perform_create(self, serializer):
        serializer.save(teacher=self.request.user)


class AssignmentSubmissionViewSet(viewsets.ModelViewSet):
    serializer_class = AssignmentSubmissionSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        qs = AssignmentSubmission.objects.select_related(
            'assignment', 'student', 'graded_by'
        ).all()
        user = self.request.user
        if is_admin(user):
            pass
        elif RoleService.has_role(user, 'school_admin'):
            school_ids = user_school_ids(user)
            qs = qs.filter(assignment__section__school_id__in=school_ids) if school_ids else qs.none()
        elif RoleService.has_role(user, 'teacher'):
            qs = qs.filter(assignment__teacher=user)
        elif RoleService.has_role(user, 'parent'):
            child_ids = FamilyLink.objects.filter(parent=user).values_list('student_id', flat=True)
            qs = qs.filter(student_id__in=child_ids)
        elif RoleService.has_role(user, 'student'):
            qs = qs.filter(student=user)
        else:
            qs = qs.none()

        assignment = self.request.query_params.get('assignment')
        if assignment:
            qs = qs.filter(assignment_id=assignment)
        student = self.request.query_params.get('student')
        if student:
            qs = qs.filter(student_id=student)
        return qs

    def perform_create(self, serializer):
        serializer.save(student=self.request.user)

    @action(detail=True, methods=['post'])
    def grade(self, request, pk=None):
        """Grade a submission."""
        submission = self.get_object()
        if not is_admin(request.user) and not RoleService.has_role(request.user, 'teacher'):
            return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)
        score = request.data.get('score')
        feedback = request.data.get('feedback', '')
        if score is None:
            return Response({'error': 'score is required'}, status=status.HTTP_400_BAD_REQUEST)
        submission.score = score
        submission.feedback = feedback
        submission.status = AssignmentSubmission.Status.GRADED
        submission.graded_at = timezone.now()
        submission.graded_by = request.user
        submission.save()
        return Response(AssignmentSubmissionSerializer(submission).data)


class SchoolManagerRequestViewSet(viewsets.ModelViewSet):
    """Request to transfer school ownership. Current manager creates, admin reviews."""
    serializer_class = SchoolManagerRequestSerializer
    permission_classes = [IsAdminOrReadOnly]

    def get_queryset(self):
        qs = SchoolManagerRequest.objects.select_related('school', 'current_manager', 'reviewed_by').all()
        if is_admin(self.request.user):
            return qs
        return qs.filter(current_manager=self.request.user)

    def get_serializer_class(self):
        if self.action == 'create':
            return SchoolManagerRequestCreateSerializer
        if self.action == 'review':
            return SchoolManagerRequestReviewSerializer
        return SchoolManagerRequestSerializer

    def perform_create(self, serializer):
        serializer.save()

    @action(detail=True, methods=['post'], url_path='review')
    def review(self, request, pk=None):
        ro = self.get_object()
        ser = SchoolManagerRequestReviewSerializer(data=request.data)
        ser.is_valid(raise_exception=True)

        ro.status = ser.validated_data['status']
        ro.admin_notes = ser.validated_data.get('admin_notes', '')
        ro.reviewed_by = request.user
        ro.reviewed_at = timezone.now()
        ro.save()

        if ro.status == 'approved' and ro.new_manager_id:
            from apps.users.services import RoleService
            School.objects.filter(id=ro.school_id).update(manager_id=ro.new_manager_id)
            try:
                from apps.users.models import User as UserModel
                new_mgr = UserModel.objects.get(id=ro.new_manager_id)
                RoleService.assign_role(new_mgr, 'school_admin')
            except Exception:
                pass

        return Response(SchoolManagerRequestSerializer(ro, context={'request': request}).data)


class SchoolDeviceViewSet(AuxiliaryModuleGatingMixin, viewsets.ModelViewSet):
    """CRUD for school devices (GPS trackers, RFID readers, cameras, mobile apps)."""
    serializer_class = SchoolDeviceSerializer
    permission_classes = [IsAdminOrReadOnly]

    def get_queryset(self):
        qs = SchoolDevice.objects.select_related('school', 'assigned_bus').all()
        school_id = self.request.query_params.get('school')
        if school_id:
            qs = qs.filter(school_id=school_id)
        device_type = self.request.query_params.get('device_type')
        if device_type:
            qs = qs.filter(device_type=device_type)
        if not is_admin(self.request.user):
            school_ids = user_school_ids(self.request.user)
            if school_ids is not None:
                qs = qs.filter(school_id__in=school_ids)
        return qs

    @action(detail=True, methods=['post'], url_path='regenerate-token')
    def regenerate_token(self, request, pk=None):
        """Generate a fresh API token for a device."""
        device = self.get_object()
        import secrets
        device.api_token = secrets.token_hex(32)
        device.save(update_fields=['api_token'])
        return Response({'api_token': device.api_token})

    @action(detail=True, methods=['post'], url_path='heartbeat')
    def heartbeat(self, request, pk=None):
        """Device calls this to report it is online."""
        device = self.get_object()
        device.status = SchoolDevice.Status.ONLINE
        device.last_seen_at = timezone.now()
        device.save(update_fields=['status', 'last_seen_at'])
        return Response({'status': 'ok'})


class DeviceTelemetryAPIView(APIView):
    """Receives GPS telemetry from bus tracking devices or mobile driver apps.

    Expected payload:
    {
        "device_identifier": "IMEI or MAC",
        "bus_number": "optional",
        "latitude": 31.95,
        "longitude": 35.93,
        "speed": 45.0,
        "heading": 180.0,
        "timestamp": "2026-08-20T07:30:00Z"
    }
    """
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        device_identifier = request.data.get('device_identifier')
        if not device_identifier:
            return Response({'error': 'device_identifier is required'}, status=status.HTTP_400_BAD_REQUEST)

        device = SchoolDevice.objects.filter(device_identifier=device_identifier).first()
        if not device:
            return Response({'error': 'Unknown device'}, status=status.HTTP_404_NOT_FOUND)

        device.status = SchoolDevice.Status.ONLINE
        device.last_seen_at = timezone.now()
        device.save(update_fields=['status', 'last_seen_at'])

        bus = device.assigned_bus
        if not bus:
            bus_number = request.data.get('bus_number')
            if bus_number:
                bus = SchoolBus.objects.filter(bus_number=bus_number, school=device.school).first()

        if not bus:
            return Response({'error': 'No bus assigned to this device'}, status=status.HTTP_400_BAD_REQUEST)

        latitude = request.data.get('latitude', 0.0)
        longitude = request.data.get('longitude', 0.0)
        speed = request.data.get('speed', 0.0)
        heading = request.data.get('heading', 0.0)
        timestamp = request.data.get('timestamp', timezone.now())

        log = BusLocationLog.objects.create(
            bus=bus,
            device=device,
            latitude=latitude,
            longitude=longitude,
            speed=speed,
            heading=heading,
            timestamp=timestamp,
        )

        return Response({'success': True, 'log_id': log.id})


class DeviceScanAPIView(APIView):
    """Receives RFID tap / facial recognition events from devices.

    Expected payload:
    {
        "device_identifier": "IMEI or MAC",
        "student_id": 123,
        "event_type": "rfid_tap" | "facial_recognition",
        "direction": "board" | "exit",
        "timestamp": "2026-08-20T07:30:00Z",
        "raw_payload": {}
    }
    """
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        device_identifier = request.data.get('device_identifier')
        if not device_identifier:
            return Response({'error': 'device_identifier is required'}, status=status.HTTP_400_BAD_REQUEST)

        device = SchoolDevice.objects.filter(device_identifier=device_identifier).first()
        if not device:
            return Response({'error': 'Unknown device'}, status=status.HTTP_404_NOT_FOUND)

        device.status = SchoolDevice.Status.ONLINE
        device.last_seen_at = timezone.now()
        device.save(update_fields=['status', 'last_seen_at'])

        student_id = request.data.get('student_id')
        student = None
        if student_id:
            student = get_object_or_404(User, pk=student_id)

        event_type = request.data.get('event_type', 'rfid_tap')
        direction = request.data.get('direction', '')
        timestamp = request.data.get('timestamp', timezone.now())
        raw_payload = request.data.get('raw_payload', {})

        event = DeviceEvent.objects.create(
            device=device,
            event_type=event_type,
            student=student,
            direction=direction,
            raw_payload=raw_payload,
            timestamp=timestamp,
        )

        if student and event_type in ('rfid_tap', 'facial_recognition') and direction:
            from apps.schools.models import Attendance, FamilyLink
            status_val = 'present' if direction == 'board' else 'absent'
            attendance, _ = Attendance.objects.update_or_create(
                student=student,
                date=timezone.localdate(),
                defaults={'status': status_val, 'school': device.school}
            )
            try:
                from .whatsapp import send_whatsapp_alert
                family_links = FamilyLink.objects.filter(student=student)
                for link in family_links:
                    if link.parent and link.parent.phone:
                        msg = f"إشعار حضور مدرسي: تم تسجيل {'صعود' if direction == 'board' else 'نزول'} الطالب {student.get_full_name() or student.email} من الحافلة بتاريخ {attendance.date}."
                        send_whatsapp_alert(link.parent.phone, msg)
            except Exception:
                pass

        return Response({'success': True, 'event_id': event.id})


class BusLiveLocationAPIView(APIView):
    """Return the latest known location for all buses in a school (or a single bus)."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        bus_id = request.query_params.get('bus')
        school_id = request.query_params.get('school')

        qs = BusLocationLog.objects.select_related('bus', 'bus__school').all()
        if bus_id:
            qs = qs.filter(bus_id=bus_id)
        elif school_id:
            qs = qs.filter(bus__school_id=school_id)
        else:
            school_ids = user_school_ids(request.user)
            if school_ids is not None:
                qs = qs.filter(bus__school_id__in=school_ids)

        from django.db.models import Max
        latest_locations = (
            qs.values('bus')
            .annotate(latest=Max('timestamp'))
            .order_by('bus')
        )

        result = []
        for entry in latest_locations:
            log = BusLocationLog.objects.select_related('bus').filter(
                bus_id=entry['bus'], timestamp=entry['latest']
            ).first()
            if log:
                result.append({
                    'bus_id': log.bus.id,
                    'bus_number': log.bus.bus_number,
                    'driver_name': log.bus.driver_name,
                    'latitude': log.latitude,
                    'longitude': log.longitude,
                    'speed': log.speed,
                    'heading': log.heading,
                    'timestamp': log.timestamp.isoformat(),
                })

        return Response(result)

